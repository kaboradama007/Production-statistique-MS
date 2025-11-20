# app_annuaire.py
# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import re
import streamlit as st
import pandas as pd
import requests
from io import StringIO
import numpy as np
import ast
from io import BytesIO
from openpyxl.utils import coordinate_to_tuple
# -------------------------------------------------------------------------
# CONFIGURATION STREAMLIT
# -------------------------------------------------------------------------
st.set_page_config(page_title="ANNUAIRE STATISTIQUE MS V1", layout="wide")
st.title("📘 ANNUAIRE STATISTIQUE MS")
# -------------------------------------------------------------------------
#----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
# PARAMÈTRES UTILISATEUR DE ENDOS BF

# ----------------------------------------------------------------------------------------------
#utilisateur="annuaire"
#passe="Annuaire@25"
utilisateur = st.text_input("Username de Endos-BF", type="default")
passe = st.text_input("Mot de passe de Endos-BF", type="password")
url_base = "https://endos.minsante.bf/api"
annee_annuaire = st.text_input("Annuaire de l'année de :", "2024")
trimestres=annee_annuaire+"Q4"

# CHARGEMENT DES FICHIERS DE DEPENDANCES
#----------------------------------------------------------------------------------------
#adresse="dependances"
Listes_indicateurs_annuaire_stat=pd.read_excel("Listes_indicateurs_annuaire_stat.xlsx",header=0,sheet_name="indicateurs")
Listes_personnes_agees_annuaire=pd.read_excel("Listes_indicateurs_annuaire_stat.xlsx",header=0,sheet_name="personnes_agees")
correspondanceUID=pd.read_excel("CorrespondanceUID_Indicateur_DE.xlsx",header=0)
# code de rangement annuaire DS ET CH
code_annuaire_global=pd.read_excel("Code_annuaireDS_CH.xlsx",header=0, sheet_name="nouvelle")
code_annuaire_ch = pd.DataFrame(code_annuaire_global, columns=["UO_annuaire", "type_uo", "cod_ann"])
code_annuaire_ds=code_annuaire_ch[code_annuaire_ch["type_uo"].isin (["DRS","DS","Total_bfa"])]
# Ordre de rangement des centres hospitaliers
code_ch_annuaire=pd.read_excel("Code_annuaireDS_CH.xlsx",header=0, sheet_name="CH")
# Ordre de présentations des actes de chirurgie dans les centres hospitaliers
chirurgie_ch=pd.read_excel("Chirurgie_ch.xlsx",header=0)
# Correspondance de chaque CH avec son distrit d'implantation
ds_implantation_CH=pd.read_excel("district_d_implantation CH.xlsx",header=0)
# nosologie de onsultation et d'hospitalisation, liste indicateurs
indicateurs_nosologies=pd.read_excel("liste_indicateurs_nosologies.xlsx",header=0)
indicateurs_noso_hospitalisation=pd.read_excel("Liste_nosologies_hospitalisations.xlsx",header=0)
# code de rangement nosologie de consultation
code_nosologie_consultation=pd.read_excel("code_nosologie_annuaire.xlsx",header=0)
code_nosologie_hospitalisation=pd.read_excel("code_nosologie_annuaire.xlsx",header=0, sheet_name="noso_hospitalisation")


# ----------------------------------------------------------------------------------------------
# FONCTION D'EXTRACTION ENDOS BF
# ----------------------------------------------------------------------------------------------
def extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire, niveau, listesindicateurs, nom_extraction=""):
    
    # --- Vérifier les indicateurs ---
    if not listesindicateurs:
        st.warning(f"⚠️ Aucun indicateur pour : {nom_extraction}")
        return None

    # --- Récupérer les secrets DHIS2 ---
    try:
        utilisateur = utilisateur
        passe = passe
        base_url_api = url_base
    except Exception as e:
        st.error("❌ Secrets DHIS2 non trouvés. Veuillez créer un secrets.toml valide.")
        return None

    # --- Conteneurs Streamlit ---
    with st.expander(f"📊 {nom_extraction} ({len(listesindicateurs)} indicateurs)", expanded=False):
        progress_bar = st.progress(0)
        status_text = st.empty()

        # URL de base
        base_url = f"{base_url_api}/29/analytics.csv?dimension=ou:LEVEL-{niveau};zmSNCYjqQGj"

        # Gestion des années/périodes
        if isinstance(annee_annuaire, (list, tuple)):
            dates = ";".join(map(str, annee_annuaire))
        else:
            dates = str(annee_annuaire)

        FINAL = "&displayProperty=NAME&ignoreLimit=TRUE&hierarchyMeta=true&showHierarchy=true&hideEmptyRows=true&rows=ou;pe;dx"

        body = ""
        succes, erreurs = 0, 0

        # --- Téléchargement des indicateurs ---
        for i, var in enumerate(listesindicateurs):
            url = f"{base_url}&dimension=pe:{dates}&dimension=dx:{var}{FINAL}"

            try:
                r = requests.get(url, auth=(utilisateur, passe), timeout=(30,None))
                if r.status_code == 200:
                    succes += 1
                    # Concatène sans l'entête sauf pour le premier indicateur
                    if i == 0:
                        body += r.text
                    else:
                        body += r.text.split("\n", 1)[1]
                else:
                    erreurs += 1
            except Exception:
                erreurs += 1

            # Mise à jour de la progression
            progress_bar.progress((i + 1) / len(listesindicateurs))
            status_text.text(f"Traitement: {i+1}/{len(listesindicateurs)} - ✅ {succes} | ⚠️ {erreurs}")

        # --- Conversion en DataFrame ---
        try:
            df = pd.read_csv(StringIO(body), low_memory=False)
            df["date_downloaded"] = datetime.today().strftime('%Y-%m-%d')

            st.success(f"✅ {succes} indicateurs téléchargés | {len(df)} lignes")
            progress_bar.empty()
            status_text.empty()

            return df
        except Exception as e:
            st.error(f"❌ Erreur lors de la conversion en DataFrame: {str(e)}")
            return None


#FONCTION DE TRAITEMENT DES DONNEES AVEC LES CHR ET CHU

def traitement_donnees_ds_ch(df_donnees,code_rangement_annuaire,uo_correspondance_unitech,CorrespondanceUID):
  df_extract= df_donnees.copy()
  df_extract=pd.merge(df_extract,CorrespondanceUID,left_on="dataid", right_on="Uid_endosBF", how="left")
  df_extract=df_extract[['orgunitlevel1', 'orgunitlevel2','orgunitlevel4','IndicateursAnnuaire','Total']].copy()
  df_extract=pd.merge(df_extract,uo_correspondance_unitech,
                           left_on=["orgunitlevel4"], right_on="structures", how="left")
  df_extract=df_extract[['orgunitlevel1', 'orgunitlevel2','correspondance','IndicateursAnnuaire','Total']]
  df_extract=df_extract.rename(columns={"correspondance":"DS_CH"})
  df=df_extract.copy()

  ##Dépivotage des données afin d'avoir les DE en colonne
  df = df.pivot_table(
      index=['orgunitlevel1', 'orgunitlevel2','DS_CH'],
      columns="IndicateursAnnuaire",
      values="Total",
      aggfunc="sum"
  ).reset_index()

  ###
  #structure district
  df_ds = df.groupby(["DS_CH"], as_index=False).sum()
  df_ds=df_ds.drop(['orgunitlevel1', 'orgunitlevel2'], axis=1)
  df_ds=df_ds.rename(columns={"DS_CH":"structures"})

  #structure région
  df_region = df.groupby(["orgunitlevel2"], as_index=False).sum()
  df_region=df_region.drop(['orgunitlevel1', 'DS_CH'], axis=1)
  df_region=df_region.rename(columns={"orgunitlevel2":"structures"})

  #structure pays
  df_pays = df.groupby(["orgunitlevel1"], as_index=False).sum()
  df_pays=df_pays.drop(['orgunitlevel2', 'DS_CH'], axis=1)
  df_pays=df_pays.rename(columns={"orgunitlevel1":"structures"})

   # fusion
  data=pd.concat([df_ds, df_region,df_pays], axis=0, ignore_index=True)
  data=pd.merge(code_rangement_annuaire,data, left_on="UO_annuaire", right_on="structures", how="left")
  # les detailles districts, region , pays
  data_detaille=data.drop(["UO_annuaire","cod_ann"], axis=1)

  data["pays"]="Burkina Faso"
   # total DS
  total_DS = data[data['type_uo'] == 'DS']
  total_DS = total_DS.groupby(["pays"], as_index=False).sum()
  total_DS=total_DS.drop(["UO_annuaire","cod_ann","structures"], axis=1)
  total_DS = total_DS.rename(columns={"pays": "structures"})
  total_DS['structures'] = total_DS['structures'].replace("Burkina Faso", "Total District")

     # total ch
  total_ch = data[data['type_uo'] == 'CH']
  total_ch = total_ch.groupby(["pays"], as_index=False).sum()
  total_ch=total_ch.drop(["UO_annuaire","cod_ann","structures"], axis=1)
  total_ch = total_ch.rename(columns={"pays": "structures"})
  total_ch['structures'] = total_ch['structures'].replace("Burkina Faso", "Total Hôpital")


 # fusion
  data=pd.concat([data_detaille, total_DS,total_ch], axis=0, ignore_index=True)
  data=pd.merge(code_rangement_annuaire,data, left_on="UO_annuaire", right_on="structures", how="left")
  data= data.sort_values(by="cod_ann", ignore_index=True)
  data=data.drop(['structures', 'cod_ann'],axis=1)
  data=data.rename(columns={"UO_annuaire":"structures"})
  df=data.copy()

  return df


#//////////////////////////////
# Traitement des donnees DS sans les hopitaux
#//////////////////////////////

def traitement_donnees_ds(df_donnees,code_rangement_annuaire,correspondanceUID):
  df_extract= df_donnees.copy()
  # mapping avec les uid des indicateurs et choix des variables d'interet
  df_extract=pd.merge(df_extract,correspondanceUID,left_on="dataid", right_on="Uid_endosBF", how="left")
  df_extract=df_extract[['orgunitlevel1', 'orgunitlevel2','orgunitlevel4','IndicateursAnnuaire','Total']].copy()

  ##Dépivotage des données afin d'avoir les DE en colonne
  df = df_extract.pivot_table(
      index=['orgunitlevel1', 'orgunitlevel2','orgunitlevel4'],
      columns="IndicateursAnnuaire",
      values="Total",
      aggfunc="sum"
  ).reset_index()


#structure district
  df_ds = df.groupby(["orgunitlevel4"], as_index=False).sum()
  df_ds=df_ds.drop(['orgunitlevel1', 'orgunitlevel2'], axis=1)
  df_ds=df_ds.rename(columns={"orgunitlevel4":"structures"})

  #structure région
  df_region = df.groupby(["orgunitlevel2"], as_index=False).sum()
  df_region=df_region.drop(['orgunitlevel1', 'orgunitlevel4'], axis=1)
  df_region=df_region.rename(columns={"orgunitlevel2":"structures"})

  #structure pays
  df_pays = df.groupby(["orgunitlevel1"], as_index=False).sum()
  df_pays=df_pays.drop(['orgunitlevel2', 'orgunitlevel4'], axis=1)
  df_pays=df_pays.rename(columns={"orgunitlevel1":"structures"})

   # fusion
  data=pd.concat([df_ds, df_region,df_pays], axis=0, ignore_index=True)
  data=pd.merge(code_rangement_annuaire,data, left_on="UO_annuaire", right_on="structures", how="left")
  data= data.sort_values(by="cod_ann", ignore_index=True)
  data=data.drop(['structures', 'cod_ann',"type_uo"],axis=1)
  data=data.rename(columns={"UO_annuaire":"structures"})

  return data


#-----------------------------------------------------------------------------------------------------
#FONCTION D'EXTRACTION DES UNITES D'ORGANISATION
#----------------------------------------------------------------------------------------------------

def extraction_UO(utilisateur, passe, url_base):
    # 🔒 Étape 1 : Charger les identifiants depuis secrets.toml ou entrée manuelle
    if "dhis2" in st.secrets:
        utilisateur = utilisateur
        passe = passe
        url_base = url_base
        st.info("🔐 Identifiants DHIS2 chargés depuis secrets.toml")
    else:
        st.warning("⚠️ Aucun fichier secrets.toml détecté. Veuillez saisir vos identifiants manuellement.")
        with st.form("login_form"):
            utilisateur = st.text_input("Nom d'utilisateur DHIS2")
            passe = st.text_input("Mot de passe DHIS2", type="password")
            url_base = st.text_input("URL de base", value="https://endos.minsante.bf/api")
            submitted = st.form_submit_button("Se connecter")
        if not submitted:
            st.stop()  # on attend la soumission du formulaire

    # 🔗 Étape 2 : Appel à l’API DHIS2
    url = f"{url_base}/organisationUnits.json"
    params = {"paging": "false", "fields": "id,name,path,level"}

    try:
        with st.spinner("⏳ Récupération des unités d’organisation..."):
            r = requests.get(url, auth=(utilisateur, passe), params=params)
    except Exception as e:
        st.error(f"Erreur de connexion : {e}")
        return None, None

    if not r.ok:
        st.error(f"❌ Erreur HTTP {r.status_code}: {r.text}")
        return None, None

    data = r.json().get("organisationUnits", [])
    if not data:
        st.warning("⚠️ Aucun résultat renvoyé par l’API.")
        return None, None

    # 🧩 Étape 3 : Transformation des données
    df_org = pd.json_normalize(data, sep="_")
    df = df_org.copy()

    df['path_clean'] = df['path'].str.lstrip('/')
    df_path_split = df['path_clean'].str.split('/', expand=True)
    df_path_split.columns = [f'level_{i+1}' for i in range(df_path_split.shape[1])]
    df_final = pd.concat([df.drop(columns=['path_clean']), df_path_split], axis=1)

    df_final.columns = ['name', 'path', 'id', "level", 'pays', 'region', 'province', 'district', 'commune', 'FS']

    df_structure = df_final[["id", "name"]]
    mappings = {col: df_structure.set_index('id')['name'].to_dict()
                for col in ['pays', 'region', 'province', 'district', 'commune', 'FS']}

    for col, mp in mappings.items():
        df_final[col] = df_final[col].map(mp).fillna(df_final[col])

    df_uo = df_final[['id', "level", 'pays', 'region', 'province', 'district', 'commune', 'name']]
    df_uo.columns = ['id', "level", 'pays', 'region', 'province', 'district', 'commune', 'structures']

    uo_correspondance_unitech = df_uo[df_uo["level"].isin([4])]
    uo_correspondance_unitech = uo_correspondance_unitech[['region', 'province', 'district', 'structures']]

    uo_correspondance_unitech["correspondance"] = uo_correspondance_unitech.apply(
        lambda row: row["district"] if str(row["district"]).startswith("DS ") else row["province"], axis=1
    )

    return df_uo, uo_correspondance_unitech

def total_ds_a_partirde_totalbf(df,code_annuaire_ch, col_ref='UO_annuaire', valeur_source='Burkina Faso', valeur_cible='Total District'):
  df=pd.merge(code_annuaire_ch,df, left_on="UO_annuaire",right_on="structures",how="left")
  colonnes_numeriques = df.select_dtypes(include='number').columns
  valeurs_bf = df.loc[df[col_ref] == valeur_source, colonnes_numeriques].iloc[0]
  df.loc[df[col_ref] == valeur_cible, colonnes_numeriques] = valeurs_bf.values
  df = df.drop(["type_uo", "cod_ann", "structures"], axis=1, errors='ignore')
  df = df.rename(columns={col_ref: "structures"})
  return df

# Ajout de colonne pour completer le dataframe avec des variables à valeurs nulles
def ajout_variables_nulle(df, colonnes):
    for col in colonnes:
        if col not in df.columns:
            df[col] = np.nan
    return df
#-----------------------------------------------------------------------------------------------------
#FONCTION D'EXTRACTION DES GROUPES UNITES D'ORGANISATION
#----------------------------------------------------------------------------------------------------

def extraction_groupe_uo(utilisateur, passe, df_uo):
    # --- Identifiants DHIS2 depuis secrets.toml ---
    utilisateur = utilisateur
    passe = passe

    # --- URL et paramètres ---
    url = "https://endos.minsante.bf/api/organisationUnitGroups.json"
    auth = (utilisateur, passe)
    params = {
        "paging": "false",
        "fields": "id,name,organisationUnits"
    }

    # --- Requête API ---
    r = requests.get(url, auth=auth, params=params)
    if not r.ok:
        st.error(f"❌ Erreur HTTP {r.status_code} : {r.text}")
        return pd.DataFrame(), pd.DataFrame()  # ← évite l'erreur

    data = r.json().get("organisationUnitGroups", [])
    df_groupe = pd.json_normalize(data, sep="_")

    # --- Fonction utilitaire ---
    def parse_units(x):
        if x is None:
            return []
        if isinstance(x, (list, np.ndarray)):
            return list(x)
        if isinstance(x, str):
            if x.strip() == "[]":
                return []
            try:
                return ast.literal_eval(x)
            except Exception:
                return []
        if pd.isna(x):
            return []
        return []

    # --- Nettoyage ---
    df_groupe["organisationUnits"] = df_groupe["organisationUnits"].apply(parse_units)
    df_eclate = df_groupe.explode("organisationUnits", ignore_index=True)
    df_eclate["organisationUnits"] = df_eclate["organisationUnits"].apply(
        lambda d: d["id"] if isinstance(d, dict) and "id" in d else None
    )

    # --- Jointure avec df_uo ---
    df_groupe = pd.merge(df_uo, df_eclate, left_on="id", right_on="organisationUnits", how="left")
    df_groupe = df_groupe[
        ["id_x", 'pays', 'region', 'province', 'district', 'commune',
         'structures', 'level', 'name']
    ]

    # --- Types et statuts ---
    type_statut = [
        "Cabinet de soins infirmiers", "Cabinet dentaire", "Cabinet médicaux",
        "CHR", "CHU", "Clinique", "Clinique d'accouchement", "CM", "CMA",
        "CREN", "CRST", "CSPS", "Dispensaire", "FS Privee", "FS Public",
        "Infirmerie", "Maternités isolées", "Polyclinique", "Unités CHR/CHU",
        "Services_CMA", "FS  définitivement Fermées"
    ]

    df_groupe_type = df_groupe[df_groupe['name'].isin(type_statut)]
    df_groupe_global = df_groupe_type[['id_x', 'structures', 'level', 'name']]

    # --- Pivot pour obtenir les types en colonnes ---
    df_pivot = df_groupe_global.pivot_table(
        index=['id_x', 'structures', 'level'],
        columns='name',
        aggfunc='size',
        fill_value=0
    ).reset_index()

    df_pivot = pd.merge(df_pivot, df_uo, left_on="id_x", right_on="id", how="left")

    # --- Sélection des colonnes finales ---
    colonnes_finales = [
        'id', 'pays', 'region', 'province', 'district', 'commune', 'structures_x',
        'level_x', 'FS Public', 'FS Privee', 'CMA', 'CM', 'CSPS', 'Dispensaire',
        'Maternités isolées', 'CREN', 'Infirmerie', 'Cabinet de soins infirmiers',
        'Cabinet dentaire', 'Cabinet médicaux', 'Clinique',
        'Clinique d\'accouchement', 'Polyclinique', "CHR", "CHU",
        "Unités CHR/CHU", 'Services_CMA', "FS  définitivement Fermées"
    ]

    for col in colonnes_finales:
        if col not in df_pivot.columns:
            df_pivot[col] = 0

    df_type = df_pivot[colonnes_finales]

    # --- Statut ---
    statut = ["FS Privee", "FS Public"]
    df_statut = df_groupe[df_groupe['name'].isin(statut)]
    df_groupe_statut = df_statut[['id_x', 'structures', 'level', 'name']]

    return df_type, df_groupe_statut

#fonction de convertion en colonne et num de ligne
def to_row_col(cell_ref):
    """
    Convertit différentes formes en (row, col):
    - "A1" -> (1,1)
    - " A1 " -> (1,1)
    - ("A1") -> (1,1)
    - (r,c) -> (r,c) si tuple de 2 ints
    - "A_1" -> (1,1) (underscore toléré)
    Retourne (row, col) ou lève ValueError.
    """
    if isinstance(cell_ref, tuple) and len(cell_ref) == 2:
        r, c = cell_ref
        if isinstance(r, int) and isinstance(c, int):
            return r, c
        else:
            raise ValueError(f"Tuple non-entiers: {cell_ref}")

    if not isinstance(cell_ref, str):
        raise ValueError(f"cellule n'est pas une chaîne: {repr(cell_ref)} (type={type(cell_ref)})")

    s = cell_ref.strip()
    # remplacer underscore ou espaces par rien
    s = s.replace("_", "").replace(" ", "")

    # recherche lettre(s)+chiffres
    m = re.match(r"^([A-Za-z]+)(\d+)$", s)
    if not m:
        raise ValueError(f"Format de cellule invalide: {repr(cell_ref)}")
    col_letters, row_str = m.group(1), m.group(2)

    # convertir col letters -> numéro de colonne
    col = 0
    for ch in col_letters.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    row = int(row_str)
    return row, col




# -------------------------------------------------------------------------
# CANEVAS
# -------------------------------------------------------------------------
st.header("CANEVAS DE BASE DE L'ANNUAIRE")

canevas_path = st.text_input(
    "Chemin du fichier canevas (ex : CanevasAnnuaire241025_fin.xlsx)",
    value="CanevasAnnuaire241025_fin.xlsx"
)

if not os.path.exists(canevas_path):
    st.warning(f"⚠️ Le fichier canevas '{canevas_path}' n’existe pas dans le répertoire courant.")
    st.stop()

# -------------------------------------------------------------------------
# BOUTON DE GÉNÉRATION
# -------------------------------------------------------------------------
if st.button("🚀 Lancer la génération du fichier Annuaire"):
    try:
        st.info("⏳ Génération du fichier en cours...")

        # 1️ Extraction des UO et correspondances CHR/CHU
        df_uo, uo_correspondance_unitech = extraction_UO(utilisateur, passe, url_base)
        st.success("✅ Unités d'organisation extraites")

        # 2️ Extraction des groupes UO
        df_type, df_groupe_statut = extraction_groupe_uo(utilisateur, passe,df_uo)
        st.success("✅ Groupes d'UO extraits")

        # 3️ Extraction des officines
        st.info("📦 Extraction des officines")
        listind0_officines = Listes_indicateurs_annuaire_stat[
            Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind0_officines"
        ]['Uid_endosBF'].dropna().unique().tolist()
        df_officines_extract = extractionendos_dhis(utilisateur, passe, url_base,trimestres, 4, listind0_officines)
        st.success("✅ Officines extraites")

        # 4️ Complétude districts
        st.info("📊 Extraction complétude districts")
        listind0_completude_ds = Listes_indicateurs_annuaire_stat[
            Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind0_completude_ds"
        ]['Uid_endosBF'].dropna().unique().tolist()
        ds_completude = extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire, 6, listind0_completude_ds)
        ds_completude["type_report"] = ds_completude["dataname"].apply(lambda x: x.split(" - ")[-1].strip())
        st.success("✅ Complétude districts extraite")

        # 5️ Complétude CHR/CHU
        st.info("📊 Extraction complétude CHR/CHU")
        listind0_completude_ch = Listes_indicateurs_annuaire_stat[
            Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind0_completude_ch"
        ]['Uid_endosBF'].dropna().unique().tolist()
        dfcompletude = extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire, 4, listind0_completude_ch)
        dfcompletude["type_report"] = dfcompletude["dataname"].apply(lambda x: x.split(" - ")[-1].strip())
        st.success("✅ Complétude CHR/CHU extraite")
        
        #6-------------- Données par districts et unités des centres hospitaliers par an-------------------------------------
        st.info("📊 Données par districts et unités des centres hospitaliers par an")
        listind1_niveau4_an=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind1_niveau4_an"]['Uid_endosBF'].dropna().unique().tolist()
        df_data_level4=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,4,listind1_niveau4_an)
        data=df_data_level4.copy()
        st.success("✅ Données par districts et unités des centres hospitaliers par an")
        
        #7-----------------------Population et populations cibles----------------------------------------------
        st.info("📊 Population et populations cibles")
        listind2_population=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind2_population"]['Uid_endosBF'].dropna().unique().tolist()
        df_population=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,4,listind2_population)
        st.success("✅ Population et populations cibles")
        

        #8------------------------DRD et rupture dans les DRD par région----------------------------------------
        st.info("📊 DRD et rupture dans les DRD par région")
        listind3_drd=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind3_drd"]['Uid_endosBF'].dropna().unique().tolist()
        df_drd_extract=extractionendos_dhis(utilisateur, passe, url_base,trimestres,2,listind3_drd)
        st.success("✅ DRD et rupture dans les DRD par région")

        #9-----------------------indicateurs Q4 (DMEG, Lepre, VIH, normes en personnel)-------------------------
        st.info("📊 indicateurs Q4 (DMEG, Lepre, VIH, normes en personnel)")
        listind4_q4_dmeg_vih=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind4_q4_dmeg_vih"]['Uid_endosBF'].dropna().unique().tolist()
        df_indicateurs_q4_extract=extractionendos_dhis(utilisateur, passe, url_base,trimestres,4,listind4_q4_dmeg_vih)
        st.success("✅ indicateurs Q4 (DMEG, Lepre, VIH, normes en personnel)")

        #510--------Indicateurs annuels (Vaccin anti palu (VAP) et causes de décès maternel par cause)-------------
        st.info("📊 Indicateurs annuels (Vaccin anti palu (VAP)")
        listind5_vap_deces_mat=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind5_vap_deces_mat"]['Uid_endosBF'].dropna().unique().tolist()
        vap_deces_mat_extract=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,4,listind5_vap_deces_mat)
        st.success("✅ Indicateurs annuels (Vaccin anti palu (VAP)")

        #11------------- indicateurs semestriels (JVA): donnees de campagne----------------------------------------
        st.info("📊 indicateurs semestriels (JVA)")
        semestres=f"{annee_annuaire}Q1;{annee_annuaire}Q2;{annee_annuaire}Q3;{annee_annuaire}Q4"
        listind6_jva=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind6_jva"]['Uid_endosBF'].dropna().unique().tolist()
        df_jva_extract=extractionendos_dhis(utilisateur, passe, url_base,semestres,4,listind6_jva)
        st.success("✅ indicateurs semestriels (JVA)")

        #12-------------CONSULTANT, REFERENCE----------------------------------------------------------------------
        st.info("📊 CONSULTANT, REFERENCE")
        listind7_consult_ref_fs=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind7_consult_ref_fs"]['Uid_endosBF'].dropna().unique().tolist()
        #Pour les FS de niveau 6: type et statut (Consultant, PCIME, evacuation, reference, sortie)
        df_consult_ref_fs_extract=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,6,listind7_consult_ref_fs)
        #pour les unites des hopitaux
        df_consult_ref_ch_extract=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,4,listind7_consult_ref_fs)
        st.success("✅ CONSULTANT, REFERENCE")

        #13------------------Lits Q4 (lits des CMA et centres hospitaliers)---------------------------------------
        st.info("📊 Lits Q4 (lits des CMA et centres hospitaliers)")
        listind8_lits_cma_ch=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind8_lits_cma_ch"]['Uid_endosBF'].dropna().unique().tolist()
        # FS niveau 6
        df_lits_fs_extract=extractionendos_dhis(utilisateur, passe, url_base,trimestres,6,listind8_lits_cma_ch)
        # Unité des hopitaux
        df_lits_ch_extract=extractionendos_dhis(utilisateur, passe, url_base,trimestres,4,listind8_lits_cma_ch)
        st.success("✅ Lits Q4 (lits des CMA et centres hospitaliers)")

        #14-------------------------------WHOPEN, MPR, Integration des services-------------------------------------
        st.info("📊 WHOPEN, MPR, Integration des services")
        listind9_whopen_mpr_int=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind9_whopen_mpr_int"]['Uid_endosBF'].dropna().unique().tolist()
        #pour les indicateurs additionnels niveau FS
        indicateurs_additionnels_extract_fs=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,6,listind9_whopen_mpr_int)
        #pour les indicateurs additionnels niveau CH
        indicateurs_additionnels_extract_ch=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,4,listind9_whopen_mpr_int)
        st.success("✅ WHOPEN, MPR, Integration des services")

        #15--------Indicateurs personnes agees-------------
        st.info("📊 Indicateurs personnes agees")
        listind10_personnes_agees=Listes_personnes_agees_annuaire['dataid'].dropna().unique().tolist()
        personnes_agees_extract=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,2,listind10_personnes_agees)
        st.success("✅ Indicateurs personnes agees")

        #16--------Indicateurs nosologie_consultation-------------
        st.info("📊 Indicateurs nosologie_consultation")
        liste_nosologie_consultation=indicateurs_nosologies[indicateurs_nosologies['type']=="noso_consultation"]['uid'].dropna().unique().tolist()
        df_nosologie_extract=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,2,liste_nosologie_consultation)
        st.success("✅ Indicateurs nosologie_consultation")

        #17--------Indicateurs noso_hospitalisation-------------
        st.info("📊 Indicateurs noso_hospitalisation")
        liste_noso_hospitalisation=indicateurs_noso_hospitalisation["uid"].dropna().unique().tolist()
        df_noso_hospitalisation=extractionendos_dhis(utilisateur, passe, url_base,annee_annuaire,2,liste_noso_hospitalisation) #40mn
        st.success("✅ Indicateurs noso_hospitalisation")
        

        #-------------------Traitement------------------------------------
        # CONSTRUCTION DE LA LISTE DES FS PAR TYPE POUR LE MAPPING AVEC DES DONNEES EXTRAITES PAR FS
        fs_fonctionnelle_type=df_type[df_type["FS  définitivement Fermées"]==0]
        fs_fonctionnelle_type=fs_fonctionnelle_type.drop(
            ["pays","region","province","district","commune","level_x","FS  définitivement Fermées","FS Public","FS Privee"], axis=1)
        fs_fonctionnelle_type=fs_fonctionnelle_type.melt(id_vars=["id","structures_x"], var_name="types_fs", value_name="Valeur")
        fs_fonctionnelle_type=fs_fonctionnelle_type[fs_fonctionnelle_type["Valeur"]!=0]
        fs_fonctionnelle_type.rename(columns={"id":"id_fs","structures_x":"FS" }, inplace=True)
        fs_fonctionnelle_type=fs_fonctionnelle_type.drop("Valeur", axis=1)

        # CONSTRUCTION DE LA LISTE DES FS PAR STATUT POUR LE MAPPING AVEC DES DONNEES EXTRAITES PAR FS
        fs_fonctionnelle_statut=df_type[df_type["FS  définitivement Fermées"]==0]
        fs_fonctionnelle_statut=fs_fonctionnelle_statut[["id","structures_x","FS Public","FS Privee"]]
        fs_fonctionnelle_statut=fs_fonctionnelle_statut.melt(id_vars=["id","structures_x"], var_name="statut_fs", value_name="Valeur")
        fs_fonctionnelle_statut=fs_fonctionnelle_statut[fs_fonctionnelle_statut["Valeur"]!=0]
        fs_fonctionnelle_statut.rename(columns={"id":"id_fs","structures_x":"FS" }, inplace=True)
        fs_fonctionnelle_statut=fs_fonctionnelle_statut.drop("Valeur", axis=1)

        # création de dataframe des formations sanitairs avec le type et statut
        type_statut_fs=pd.merge(fs_fonctionnelle_statut,fs_fonctionnelle_type,left_on="id_fs",right_on="id_fs",how="left")
        type_statut_fs.drop("FS_y", axis=1,inplace=True)
        type_statut_fs.rename(columns={"FS_x":"FS","statut_fs":"statut","types_fs":"types"}, inplace=True)



        #------------------------------TRAITEMENT DES DONNEES SUR LES INFRASTRUCTURE DE L'ANNUAIRE-----------------
        data_corr_ch=ds_implantation_CH.copy()

        level=[5,6]
        df_type_annuaire = df_type[
            (((df_type["Unités CHR/CHU"] == 0) &
            (df_type["Services_CMA"] == 0)) &
            (df_type["FS  définitivement Fermées"] == 0))
            ].copy()


        # Création d'un dictionnaire de correspondance pour le mapping des champs vide de district
        dict_corr_ch = dict(zip(data_corr_ch["uid"], data_corr_ch["ds"]))
        df_type_annuaire["district"] = df_type_annuaire["district"].fillna(df_type_annuaire["id"].map(dict_corr_ch))

        #----------------------------------------------------------------------------------------
        # STRUCTURES PUBLIQUES
        #----------------------------------------------------------------------------------------
        fs_publique=df_type_annuaire[df_type_annuaire["FS Public"]==1]
        fs_publique=fs_publique.copy()
        sommes_infirmeries = ['CREN', 'Infirmerie','Cabinet de soins infirmiers','Cabinet dentaire', 'Cabinet médicaux',
                            'Clinique', "Clinique d'accouchement", 'Polyclinique']

        fs_publique["Infirmerie"] = fs_publique[sommes_infirmeries].sum(axis=1)
        fs_publique=fs_publique[['id', 'pays', 'region', 'province', 'district','CHU','CHR','CMA', 'CM','CSPS', 'Dispensaire', 'Maternités isolées',"Infirmerie"]].copy()
        col_to_totalpublique=['CHU','CHR','CMA', 'CM','CSPS', 'Dispensaire', 'Maternités isolées',"Infirmerie"]
        fs_publique["total"] = fs_publique[col_to_totalpublique].sum(axis=1)

        #structure district
        fs_publique_ds = fs_publique.groupby(["district"], as_index=False).sum()
        fs_publique_ds=fs_publique_ds.drop(['id', 'pays', 'region', 'province'], axis=1)
        fs_publique_ds=fs_publique_ds.rename(columns={"district":"structures"})

        #structure région
        fs_publique_region = fs_publique.groupby(["region"], as_index=False).sum()
        fs_publique_region=fs_publique_region.drop(['id', 'pays', 'district', 'province'], axis=1)
        fs_publique_region=fs_publique_region.rename(columns={"region":"structures"})

        #structure pays
        fs_publique_pays=fs_publique.copy()
        fs_publique_pays["pays"]="Burkina Faso"
        fs_publique_pays=fs_publique_pays.groupby(["pays"], as_index=False).sum()
        fs_publique_pays=fs_publique_pays.drop(['id', 'region', 'province', 'district'], axis=1)
        fs_publique_pays=fs_publique_pays.rename(columns={"pays":"structures"})

        # fusion FS publique
        fs_publique_uo=pd.concat([fs_publique_ds, fs_publique_region,fs_publique_pays], axis=0, ignore_index=True)
        fs_publique_uo=pd.merge(fs_publique_uo,code_annuaire_ds,left_on="structures", right_on="UO_annuaire", how="left")
        fs_publique_uo= fs_publique_uo.sort_values(by="cod_ann", ignore_index=True)
        fs_publique_uo=fs_publique_uo.drop(['UO_annuaire', 'type_uo', 'cod_ann'],axis=1)



        #----------------------------------------------------------------------------
        #OFFICINES ET DES DEPOTS PRIVEES
        #-----------------------------------------------------------------------------
        officine_df=traitement_donnees_ds(df_officines_extract,code_annuaire_ds,correspondanceUID)

        #----------------------------------------------------------------------------------------
        # STRUCTURES PRIVEE
        #----------------------------------------------------------------------------------------
        # Structures publiques
        fs_privee=df_type_annuaire[df_type_annuaire["FS Public"]==0]
        fs_privee=fs_privee.copy()
        sommes_autresprivees = ['Maternités isolées', 'CREN']
        fs_privee["Autre privé"] = fs_privee[sommes_autresprivees].sum(axis=1)
        fs_privee=fs_privee[['id', 'pays', 'region', 'province', 'district','CHR','Polyclinique','Clinique','CMA',
        'CM','Cabinet médicaux','Cabinet dentaire', 'CSPS',"Clinique d'accouchement", 'Cabinet de soins infirmiers','Dispensaire','Infirmerie',"Autre privé"]].copy()
        fs_privee=fs_privee.rename(columns={"CHR":"Hopital"})

        col_to_totalprivee=['Hopital','Polyclinique','Clinique','CMA','CM','Cabinet médicaux','Cabinet dentaire', 'CSPS',"Clinique d'accouchement", 'Cabinet de soins infirmiers','Dispensaire','Infirmerie',"Autre privé"]
        fs_privee["total"] = fs_privee[col_to_totalprivee].sum(axis=1)

        #structure district
        fs_privee_ds = fs_privee.groupby(["district"], as_index=False).sum()
        fs_privee_ds=fs_privee_ds.drop(['id', 'pays', 'region', 'province'], axis=1)
        fs_privee_ds=fs_privee_ds.rename(columns={"district":"structures"})

        #structure région
        fs_privee_region = fs_privee.groupby(["region"], as_index=False).sum()
        fs_privee_region=fs_privee_region.drop(['id', 'pays', 'district', 'province'], axis=1)
        fs_privee_region=fs_privee_region.rename(columns={"region":"structures"})

        #structure pays
        fs_privee_pays=fs_privee.copy()
        fs_privee_pays["pays"]="Burkina Faso"
        fs_privee_pays=fs_privee_pays.groupby(["pays"], as_index=False).sum()
        fs_privee_pays=fs_privee_pays.drop(['id', 'region', 'province', 'district'], axis=1)
        fs_privee_pays=fs_privee_pays.rename(columns={"pays":"structures"})

        # fusion privee
        fs_privee_uo=pd.concat([fs_privee_ds, fs_privee_region,fs_privee_pays], axis=0, ignore_index=True)
        fs_privee_uo=pd.merge(code_annuaire_ds,fs_privee_uo, left_on="UO_annuaire", right_on="structures", how="left")
        fs_privee_uo= fs_privee_uo.sort_values(by="cod_ann", ignore_index=True)
        fs_privee_uo=fs_privee_uo.drop(['structures', 'type_uo', 'cod_ann'],axis=1)
        fs_privee_uo.rename(columns={"UO_annuaire":"structures"})
        fs_privee_uo=pd.merge(fs_privee_uo, officine_df, left_on="UO_annuaire", right_on="structures", how="left")
        fs_privee_annuaire=fs_privee_uo.drop("structures", axis=1)
        fs_privee_annuaire=fs_privee_annuaire.rename(columns={"UO_annuaire":"structures"})

        #fs_publique_uo.to_excel("fs_publique_annuaire.xlsx", index=0)
        #fs_privee_annuaire.to_excel("fs_privee_annuaire.xlsx", index=0)
        df_fs_publique=fs_publique_uo.drop("structures", axis=1)
        df_fs_privee=fs_privee_annuaire.drop("structures", axis=1)
        df_fs_privee1=df_fs_privee[['Hopital', 'Polyclinique', 'Clinique', 'CMA', 'CM', 'Cabinet médicaux','Cabinet dentaire', 'CSPS']]
        df_fs_privee2=df_fs_privee[["Clinique d'accouchement",'Cabinet de soins infirmiers', 'Dispensaire', 'Infirmerie','Autre privé', 'total',
                                    'Effectif Officines','Effectif Dépôts pharmaceutiques']]

        code_annuaire_unite_ch=uo_correspondance_unitech.copy()




        #--------------------------------------------------------------------------------------------------------------------------------------------
        #       COMPLETUDE DE LA SAISIE DES RAPPORTS
        #---------------------------------------------------------------------------------------------------------------------------------------------
        ch_completude =dfcompletude.copy()
        masque_ch=ch_completude["organisationunitname"].str.startswith("DS", na=False)
        ch_completude=ch_completude[~masque_ch]

        completudeds_ch=pd.concat([ds_completude,ch_completude], ignore_index=True)
        df_completude=pd.merge(completudeds_ch,df_groupe_statut, left_on="organisationunitid", right_on="id_x", how="left")

        formulaire_fs=[

        "2021-CH 01 - Consultants et mouvements des hospitalisés et références - Actual reports",
        "2021-CH 01 - Consultants et mouvements des hospitalisés et références - Actual reports on time",
        "2021-CH 01 - Consultants et mouvements des hospitalisés et références - Expected reports",
        "2021-CM-CMA 08 - Consultants, mouvements des malades et personnes âgées - Actual reports",
        "2021-CM-CMA 08 - Consultants, mouvements des malades et personnes âgées - Actual reports on time",
        "2021-CM-CMA 08 - Consultants, mouvements des malades et personnes âgées - Expected reports",
        "2021-CSPS 05 - Morbidité 2 - Actual reports",
        "2021-CSPS 05 - Morbidité 2 - Actual reports on time",
        "2021-CSPS 05 - Morbidité 2 - Expected reports"
        ]
        df_completude=df_completude[df_completude["dataname"].isin(formulaire_fs)]


        ## COMPLETUDE DU PUBLIQUE
        completude_publique=df_completude[df_completude["name"].isin(["FS Public"])]
        completude_publique=completude_publique.drop(['orgunitlevel1', 'orgunitlevel3','orgunitlevel5', 'orgunitlevel6', 'organisationunitid',
            'organisationunitname', 'organisationunitcode',
            'organisationunitdescription', 'periodid', 'periodcode',
            'perioddescription', 'dataid','datacode',
            'datadescription', 'date_downloaded', 'id_x',
            'structures', 'level',"name","periodname","dataname"],axis=1)
        colonnes=['orgunitlevel2', 'orgunitlevel4', 'type_report', 'Total']
        completude_publique=completude_publique[colonnes]

        ##Dépivotage des données afin d'avoir les DE en colonne
        completude_publique = completude_publique.pivot_table(
            index=["orgunitlevel2","orgunitlevel4"],
            columns="type_report",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        completude_publique = completude_publique.rename(columns={"type_report":"indexe","orgunitlevel2": "region","orgunitlevel4":"unitech_ds",
                                                                'Actual reports':"Publique-Rapports_saisis", 'Actual reports on time':"Publique-Rapports_saisis_a_temps",
                                                                    'Expected reports':"Publique-Rapports_attendus"})


        # COMPLETUDE FS PRIVEE
        completude_privee=df_completude[df_completude["name"].isin(["FS Privee"])]

        completude_privee=completude_privee.drop(['orgunitlevel1', 'orgunitlevel3','orgunitlevel5', 'orgunitlevel6', 'organisationunitid',
            'organisationunitname', 'organisationunitcode',
            'organisationunitdescription', 'periodid', 'periodcode',
            'perioddescription', 'dataid','datacode',
            'datadescription', 'date_downloaded', 'id_x',
            'structures', 'level',"name","periodname","dataname"],axis=1)


        colonnes=['orgunitlevel2', 'orgunitlevel4', 'type_report', 'Total']
        completude_privee=completude_privee[colonnes]

        ##Dépivotage des données afin d'avoir les DE en colonne
        completude_privee = completude_privee.pivot_table(
            index=["orgunitlevel2","orgunitlevel4"],
            columns="type_report",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        completude_privee = completude_privee.rename(columns={"type_report":"indexe","orgunitlevel2": "region","orgunitlevel4":"unitech_ds",
                                                'Actual reports':"Privee-Rapports_saisis", 'Actual reports on time':"Privee-Rapports_saisis_a_temps",
                                                                    'Expected reports':"Privee-Rapports_attendus"})


        completude_global=pd.merge(completude_publique,completude_privee,
                                left_on=["region",	"unitech_ds"],
                                right_on=["region",	"unitech_ds"], how="outer")

        completude_global=pd.merge(completude_global,uo_correspondance_unitech,
                                left_on=["unitech_ds"], right_on="structures", how="left")


        #######################################################################################
        completude_finale=completude_global[['region_x', 'correspondance','Publique-Rapports_saisis',
            'Publique-Rapports_saisis_a_temps', 'Publique-Rapports_attendus',
            'Privee-Rapports_saisis', 'Privee-Rapports_saisis_a_temps',
            'Privee-Rapports_attendus' ]]

        completude_finale=completude_finale.groupby(['region_x',"correspondance"], as_index=False).sum()
        completude_finale["type_uo"] = completude_finale["correspondance"].apply(
            lambda x: "DS" if str(x).startswith("DS ") else "CH"
        )
        completude_finale["pays"]="Burkina Faso"

        completude_finale=completude_finale[['pays','region_x','type_uo', 'correspondance', 'Publique-Rapports_saisis',
            'Publique-Rapports_saisis_a_temps', 'Publique-Rapports_attendus',
            'Privee-Rapports_saisis', 'Privee-Rapports_saisis_a_temps',
            'Privee-Rapports_attendus']]

        completude_finale = completude_finale.rename(columns={"region_x": "region", "correspondance":"structures"})
        completude_ds_ch=completude_finale.drop(["pays",	"region",	"type_uo"], axis=1)

        # total pays
        completude_pays=completude_finale.groupby(['pays'], as_index=False).sum()
        completude_pays=completude_pays.drop(["region","structures","type_uo"], axis=1)
        completude_pays = completude_pays.rename(columns={"pays": "structures"})

        # total région

        completude_region=completude_finale.groupby(['region'], as_index=False).sum()
        completude_region=completude_region.drop(["pays","structures","type_uo"], axis=1)
        completude_region = completude_region.rename(columns={"region": "structures"})


        # total DS
        completude_District = completude_finale[completude_finale['type_uo'] == 'DS']
        completude_District = completude_District.groupby(["pays"], as_index=False).sum()
        completude_District=completude_District.drop(["region","structures","type_uo"], axis=1)
        completude_District = completude_District.rename(columns={"pays": "structures"})
        completude_District['structures'] = completude_District['structures'].replace("Burkina Faso", "Total District")

        # total  hopital "Total Hôpital"
        completude_ch = completude_finale[completude_finale['type_uo'] == "CH"]
        completude_ch = completude_ch.groupby(["pays"], as_index=False).sum()
        completude_ch=completude_ch.drop(["region","structures","type_uo"], axis=1)
        completude_ch = completude_ch.rename(columns={"pays": "structures"})
        completude_ch['structures'] = completude_ch['structures'].replace("Burkina Faso", "Total Hôpital")

        completude_fusion = pd.concat([completude_ds_ch,completude_region,completude_District,completude_ch,completude_pays], axis=0, ignore_index=True)
        completude_fusion=pd.merge(code_annuaire_ch,completude_fusion,left_on="UO_annuaire", right_on="structures", how="left")
        completude_fusion= completude_fusion.sort_values(by="cod_ann", ignore_index=True)
        completude_fs=completude_fusion.drop(['UO_annuaire', 'type_uo', 'cod_ann'],axis=1)


        #_________________________COMPLETUDE COMMUNAUTAIRE___________________________________________

        completudeds_ch=pd.concat([ds_completude,ch_completude], ignore_index=True)
        dfc_asbc=pd.merge(completudeds_ch,df_groupe_statut, left_on="organisationunitid", right_on="id_x", how="left")

        formulaire_fs=[
        "2021-ASBC - Formulaire communautaire - Actual reports",
        "2021-ASBC - Formulaire communautaire - Actual reports on time",
        "2021-ASBC - Formulaire communautaire - Expected reports"
        ]
        dfc_asbc=dfc_asbc[dfc_asbc["dataname"].isin(formulaire_fs)]

        dfc_asbc=dfc_asbc.drop(['orgunitlevel1', 'orgunitlevel3','orgunitlevel5', 'orgunitlevel6', 'organisationunitid',
            'organisationunitname', 'organisationunitcode',
            'organisationunitdescription', 'periodid', 'periodcode',
            'perioddescription', 'dataid','datacode',
            'datadescription', 'date_downloaded', 'id_x',
            'structures', 'level',"name","periodname","dataname"],axis=1)
        colonnes=['orgunitlevel2', 'orgunitlevel4', 'type_report', 'Total']
        dfc_asbc=dfc_asbc[colonnes]

        ##Dépivotage des données afin d'avoir les DE en colonne
        dfc_asbc = dfc_asbc.pivot_table(
            index=["orgunitlevel2","orgunitlevel4"],
            columns="type_report",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        dfc_asbc = dfc_asbc.rename(columns={"type_report":"indexe","orgunitlevel2": "region","orgunitlevel4":"unitech_ds",
                                                                'Actual reports':"ASBC-Rapports_saisis", 'Actual reports on time':"ASBC-Rapports_saisis_a_temps",
                                                                    'Expected reports':"ASBC-Rapports_attendus"})


        dfc_asbc["pays"]="Burkina Faso"
        dfc_asbc = dfc_asbc.rename(columns={"unitech_ds": "structures"})
        dfc_asbc=dfc_asbc[["pays","region","structures","ASBC-Rapports_saisis","ASBC-Rapports_saisis_a_temps","ASBC-Rapports_attendus"]]


        # total pays
        dfc_asbc_pays=dfc_asbc.groupby(['pays'], as_index=False).sum()
        dfc_asbc_pays=dfc_asbc_pays.drop(["region","structures"], axis=1)
        dfc_asbc_pays = dfc_asbc_pays.rename(columns={"pays": "structures"})

        # total région

        dfc_asbc_region=dfc_asbc.groupby(['region'], as_index=False).sum()
        dfc_asbc_region=dfc_asbc_region.drop(["pays","structures"], axis=1)
        dfc_asbc_region = dfc_asbc_region.rename(columns={"region": "structures"})
        dfc_asbc=dfc_asbc.drop(["pays","region"], axis=1)
        dfc_asbc_fusion = pd.concat([dfc_asbc,dfc_asbc_region,dfc_asbc_pays], axis=0, ignore_index=True)


        df_ds=code_annuaire_ch[code_annuaire_ch["type_uo"].isin(["DS","DRS","Total_bfa"])]

        dfc_asbc_fusion=pd.merge(df_ds,dfc_asbc_fusion,left_on="UO_annuaire", right_on="structures", how="left")
        dfc_asbc_fusion= dfc_asbc_fusion.sort_values(by="cod_ann", ignore_index=True)
        dfc_asbc_fusion=dfc_asbc_fusion.drop(['structures', 'type_uo', 'cod_ann'],axis=1)

        #------AUTRES DONNEES
        # liste des indicateurs annuaire
        #-----------------------------------------------------------------------------------------------------------------------------
        #1-------------- Données par districts et unités des centres hospitaliers par an-------------------------------------
        annauire_ind1_niveau4_an=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind1_niveau4_an"]['IndicateursAnnuaire'].dropna().unique().tolist()

        #2-----------------------Population et populations cibles----------------------------------------------
        annauire_ind2_population=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind2_population"]['IndicateursAnnuaire'].dropna().unique().tolist()

        #3------------------------DRD et rupture dans les DRD par région----------------------------------------
        annauire_ind3_drd=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind3_drd"]['IndicateursAnnuaire'].dropna().unique().tolist()

        #4-----------------------indicateurs Q4 (DMEG, Lepre, VIH, normes en personnel)-------------------------
        annauire_ind4_q4_dmeg_vih=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind4_q4_dmeg_vih"]['IndicateursAnnuaire'].dropna().unique().tolist()

        #5--------Indicateurs annuels (Vaccin anti palu (VAP) et causes de décès maternel par cause)-------------
        annauire_ind5_vap_deces_mat=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind5_vap_deces_mat"]['IndicateursAnnuaire'].dropna().unique().tolist()
        #6------------- indicateurs semestriels (JVA): donnees de campagne----------------------------------------
        annauire_ind6_jva=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind6_jva"]['IndicateursAnnuaire'].dropna().unique().tolist()
        #annauire_ind6_jva=["UO_annuaire"] + annauire_ind6_jva
        #7-------------CONSULTANT, REFERENCE----------------------------------------------------------------------
        annauire_ind7_consult_ref_fs=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind7_consult_ref_fs"]['IndicateursAnnuaire'].dropna().unique().tolist()
        #8------------------Lits Q4 (lits des CMA et centres hospitaliers)---------------------------------------
        annauire_ind8_lits_cma_ch=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind8_lits_cma_ch"]['IndicateursAnnuaire'].dropna().unique().tolist()
        #9-------------------------------WHOPEN, MPR, Integration des services-------------------------------------
        annauire_ind9_whopen_mpr_int=Listes_indicateurs_annuaire_stat[Listes_indicateurs_annuaire_stat['listes_indicateurs']=="listind9_whopen_mpr_int"]['IndicateursAnnuaire'].dropna().unique().tolist()



        #--------------------------------------------------------------------------------------------------------------------------------
        #                TRAITEMENT DES AUTRES DONNEES
        #--------------------------------------------------------------------------------------------------------------------------------

        df_jva=df_jva_extract.copy()
        #--------------Liste des régions--------------
        region_bf=code_annuaire_ch[code_annuaire_ch["type_uo"].isin(["DRS","Total_bfa"])].reset_index(drop=True)
        region_bf.drop(["type_uo","cod_ann"],axis=1, inplace=True)

        #--------------Traitement des données de niveau district et unité ch------------------------------------
        df_data_level4=data.copy()
        df_data_level_concat_pop=pd.concat([df_data_level4,df_population], ignore_index=0)
        data_ds=traitement_donnees_ds_ch(df_data_level_concat_pop,code_annuaire_ch,uo_correspondance_unitech,correspondanceUID)
        data_ds=data_ds.drop("type_uo_y", axis=1)
        data_ds.rename(columns={"type_uo_x":"type_uo"}, inplace=1)
        data_ds=ajout_variables_nulle(data_ds,annauire_ind1_niveau4_an)
        df=data_ds.copy()


        #------------------POPULATION---------------------------------------------
        data_population=traitement_donnees_ds(df_population,code_annuaire_ds,correspondanceUID)
        data_population=ajout_variables_nulle(data_population,annauire_ind2_population)
        data_populationfinal=pd.merge(code_annuaire_ch,data_population,left_on="UO_annuaire", right_on="structures", how="left")
        # affectation des données du Burkina a total district
        colonnes_numeriques = data_populationfinal.select_dtypes(include='number').columns

        valeurs_bf = data_populationfinal.loc[data_populationfinal['UO_annuaire'] == 'Burkina Faso', colonnes_numeriques].iloc[0]
        data_populationfinal.loc[data_populationfinal['UO_annuaire'] == 'Total District', colonnes_numeriques] = valeurs_bf.values
        data_populationfinal.drop(["type_uo","cod_ann","structures"], axis=1, inplace=True)
        data_populationfinal.rename(columns={"UO_annuaire":"structures"}, inplace=True)

        #------------------DRD---------------------------------------------
        df_drd=df_drd_extract.copy()
        df_drd=pd.merge(df_drd,correspondanceUID,left_on="dataid", right_on="Uid_endosBF", how="left")
        df_drd=df_drd[['orgunitlevel1', 'orgunitlevel2','IndicateursAnnuaire','Total']].copy()

        ##Dépivotage des données afin d'avoir les DE en colonne
        df_drd = df_drd.pivot_table(
            index=['orgunitlevel1', 'orgunitlevel2'],
            columns="IndicateursAnnuaire",
            values="Total",
            aggfunc="sum"
        ).reset_index()
        # regions
        df_drd_region=df_drd.drop("orgunitlevel1", axis=1)
        df_drd_region.rename(columns={"orgunitlevel2":"structures"}, inplace=1)
        # pays
        df_drd_pays=df_drd.groupby("orgunitlevel1",as_index=False ).sum()
        df_drd_pays=df_drd_pays.drop("orgunitlevel2", axis=1)
        df_drd_pays.rename(columns={"orgunitlevel1":"structures"}, inplace=1)
        # fusion
        df_drd_final=pd.concat([df_drd_region,df_drd_pays], ignore_index=1).rename_axis(None, axis=1)
        df_drd_final_merge=pd.merge(region_bf,df_drd_final, left_on="UO_annuaire",right_on="structures",how="left")
        df_drd_final_merge=ajout_variables_nulle(df_drd_final_merge,annauire_ind3_drd)
        df_drd_final=df_drd_final_merge.drop(["structures","UO_annuaire"], axis=1)
        #------------------indicateurs_q4---------------------------------------------
        df_indicateurs_q4_extract1=df_indicateurs_q4_extract
        df_indicateurs_q4=traitement_donnees_ds_ch(df_indicateurs_q4_extract1,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_indicateurs_q4.drop(["type_uo_y"], axis=1, inplace=True)
        df_indicateurs_q4.rename(columns={"type_uo_x":"type_uo"}, inplace=True)
        df_specifique=df_indicateurs_q4.copy()
        df_specifique=ajout_variables_nulle(df_specifique,annauire_ind4_q4_dmeg_vih)
        masque_ds=df_specifique["type_uo"].isin(["Total_bfa","DS","DRS"])
        df_specifique_ds=df_specifique[masque_ds]

        #------------------VAP deces_mat---------------------------------------------
        df_vap_deces_mat_extract=vap_deces_mat_extract
        df_vap_deces_mat=traitement_donnees_ds_ch(df_vap_deces_mat_extract,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_vap_deces_mat.drop(["type_uo_y"], axis=1, inplace=True)
        df_vap_deces_mat.rename(columns={"type_uo_x":"type_uo"}, inplace=True)
        df_vap_deces_mat=ajout_variables_nulle(df_vap_deces_mat,annauire_ind5_vap_deces_mat)
        #------------------JVA---------------------------------------------
        df_jva_brut=df_jva.copy()
        df_jva_brut["trimestre"]=df_jva_brut["periodid"].str[-2:]
        masque_jva_p1=df_jva_brut["trimestre"].isin(["Q1","Q2"])
        df_jva_brut_p1=df_jva_brut[masque_jva_p1]
        df_jva_brut_p2=df_jva_brut[~masque_jva_p1]
        #----------------Application de la fonction de traitement-------------------------
        #pour p1
        df_jva_p1=traitement_donnees_ds(df_jva_brut_p1,code_annuaire_ds,correspondanceUID)
        df_jva_p1_merge=pd.merge(code_annuaire_ch,df_jva_p1,left_on="UO_annuaire",right_on="structures", how="left")
        df_jva_p1_merge=df_jva_p1_merge.drop(["type_uo", "cod_ann","structures"],axis=1)
        df_jva_p1_merge=ajout_variables_nulle(df_jva_p1_merge,annauire_ind6_jva)

        #pour p2
        df_jva_p2=traitement_donnees_ds(df_jva_brut_p2,code_annuaire_ds,correspondanceUID)
        df_jva_p2_merge=pd.merge(code_annuaire_ch,df_jva_p2,left_on="UO_annuaire",right_on="structures", how="left")
        df_jva_p2_merge=df_jva_p2_merge.drop(["type_uo", "cod_ann","structures"],axis=1)
        df_jva_p2_merge=ajout_variables_nulle(df_jva_p2_merge,annauire_ind6_jva)
        #------------------------consultant---------------------------------------
        #données CH
        df_consult_ref_ch=df_consult_ref_ch_extract.copy()
        masque_ch=~df_consult_ref_ch['organisationunitname'].str.startswith("DS",na=False)
        df_consult_ref_ch=df_consult_ref_ch[masque_ch]
        #données DS
        df_consult_ref_fs=df_consult_ref_fs_extract.copy()
        df_consult_ref_fs.drop(['orgunitlevel5', 'orgunitlevel6'], axis=1, inplace=True)
        #fusion des tableau CH et DS
        df_consultant_reference=pd.concat([df_consult_ref_fs,df_consult_ref_ch], axis=0, ignore_index=True)
        df_consultant_reference_merge=pd.merge(df_consultant_reference,type_statut_fs,left_on="organisationunitid", right_on="id_fs", how="left")
        df_consultant_reference_merge=ajout_variables_nulle(df_consultant_reference_merge,annauire_ind7_consult_ref_fs)

        #------------------------consultant privee---------------------------------
        #traitement tableau consultant
        df_prive=df_consultant_reference_merge.copy()
        masque_prive=df_prive["statut"]=="FS Privee"
        df_prive=df_prive[masque_prive]
        df_prive=traitement_donnees_ds_ch(df_prive,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_prive=ajout_variables_nulle(df_prive,annauire_ind7_consult_ref_fs)
        #--------------------Tableau PCIME------------------------------------------------------
        df_premier_echelon=df_consultant_reference_merge.copy()
        premier_echelon=["CM","Clinique d'accouchement","Cabinet de soins infirmiers","Maternités isolées",
                    "CSPS","CREN","Cabinet dentaire","Infirmerie","Cabinet médicaux","Dispensaire"]
        df_premier_echelon=df_premier_echelon[df_premier_echelon["types"].isin(premier_echelon)]
        df_pcime=traitement_donnees_ds_ch(df_premier_echelon,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_pcime=df_pcime[["Nouveaux consultants Moins de 5 ans","Nombre d'enfants pris en charge selon l'approche PCIME"]]
        #--------------------Tableau TETU------------------------------------------------------
        df_reference_tetu=df_consultant_reference_merge.copy()
        cma_ch=["Services_CMA","Unités CHR/CHU"]
        df_reference_tetu=df_reference_tetu[df_reference_tetu["types"].isin(cma_ch)]
        df_reference_tetu=traitement_donnees_ds_ch(df_reference_tetu,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)

        df_tetu_cma_ch=df_reference_tetu[["Nouveaux consultants moins de cinq ans (auto orienté, évacué, référé)",
                                        "Nombre d'enfants pris en charge selon l'approche TETU"]]

        #--------------------------Tableau reference premier echelon----------------------------------
        df_ref_contreref_echelon1=traitement_donnees_ds(df_premier_echelon,code_annuaire_ds,correspondanceUID)
        df_ref_contreref_echelon1=df_ref_contreref_echelon1[["FS 1er échelon Références/Evacuations réalisées",
                                                    "FS 1er échelon Contre-référence reçue"]]
        #------------------------------Tableau reference CMA et CH---------------------------------
        df_reference_cma_ch=df_consultant_reference_merge.copy()
        cma_ch=["Services_CMA","Unités CHR/CHU"]
        df_reference_cma_ch=df_reference_cma_ch[df_reference_cma_ch["types"].isin(cma_ch)]
        df_reference_cma_ch= traitement_donnees_ds_ch(df_reference_cma_ch,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_ref_contreref_cma_ch=df_reference_cma_ch[["FS 1er échelon Références/Evacuations réalisées",
                                                    "FS 1er échelon Contre-référence reçue"]]
        #---------------------------------Tableau evacuation CMA et CH------------------------------------
        df_evacuation_cma_ch=df_reference_cma_ch[["Evacuations /references réçues",
                                                    "Nombre de Contre reference réalisées"]]
        #---------------------------------Tableau MEO et deces------------------------------------
        df_meo=df_consultant_reference_merge.copy()
        premier_echelon_meo=["Clinique d'accouchement","Cabinet de soins infirmiers","Maternités isolées",
                    "CSPS","CREN","Infirmerie","Dispensaire"]
        df_meo=df_meo[df_meo["types"].isin(premier_echelon_meo)]
        df_meo=traitement_donnees_ds(df_meo,code_annuaire_ds,correspondanceUID)
        df_mise_en_observation=df_meo[["Nombre de malades mis en observation","Nombre de décès","Décès d'enfants de moins de 5 ans",
                                    "Dont nombre de décès des enfants de moins de 5 ans dans les 24 heures après admission"]]
        #--------------------Tableau mode d'entrée en hospitalisation------------------------------------------------------
        df_deuxieme_echelon_ch=df_consultant_reference_merge.copy()
        deuxieme_echelon_ch=["Services_CMA","Unités CHR/CHU","CM","Clinique","Cabinet médicaux","Polyclinique"]
        df_deuxieme_echelon_ch=df_deuxieme_echelon_ch[df_deuxieme_echelon_ch["types"].isin(deuxieme_echelon_ch)]
        df_entre_hospi=traitement_donnees_ds_ch(df_deuxieme_echelon_ch,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_entre_hospi=df_entre_hospi[["Malades entrés Auto orientés","Malades entrés Référés","Malades entrés Evacués"]]

        #--------------------Tableau mode de sortie en hospitalisation------------------------------------------------------
        df_deuxieme_echelon_ch=df_consultant_reference_merge.copy()
        deuxieme_echelon_ch=["Services_CMA","Unités CHR/CHU","CM","Clinique","Cabinet médicaux","Polyclinique"]
        df_deuxieme_echelon_ch=df_deuxieme_echelon_ch[df_deuxieme_echelon_ch["types"].isin(deuxieme_echelon_ch)]
        df_sortie_hospi=traitement_donnees_ds_ch(df_deuxieme_echelon_ch,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_sortie_hospi=df_sortie_hospi[["Malades sortis Guéris","Malades sortis Sans avis medical","Malades sortis Référés",
                                        "Malades sortis Evacués","Malades sortis Décédés","Malades sortis Contre avis médical"]]


        #------------------------consultant CHR CHU ---------------------------------
        #consultants chr_chu
        df_consultant_chr_chu=df_consultant_reference_merge.copy()
        df_consultant_chr_chu=pd.merge(df_consultant_chr_chu,correspondanceUID, left_on="dataid", right_on="Uid_endosBF", how="left")

        df_consultant_chr_chu=df_consultant_chr_chu[df_consultant_chr_chu["types"]=="Unités CHR/CHU"]
        df_consultant_chr_chu=df_consultant_chr_chu[["orgunitlevel3","IndicateursAnnuaire","Total"]]
        df_consultant_chr_chu=df_consultant_chr_chu.pivot_table(index="orgunitlevel3",
                                                                columns="IndicateursAnnuaire",
                                                                values="Total",aggfunc="sum"
                                                                ).reset_index()
        df_consultant_chr_chu=df_consultant_chr_chu[["orgunitlevel3","Nouveaux consultants Ensemble","Consultants Référés","Consultants Evacués"]]
        df_consultant_chr_chu=pd.merge(df_consultant_chr_chu,code_ch_annuaire,left_on="orgunitlevel3", right_on="hopital", how="left")
        df_consultant_chr_chu.sort_values("code",inplace=True)
        df_consultant_chr_chu.drop(["orgunitlevel3","code", "hopital"], axis=1, inplace=True)

        #------------------------chirurgie ds ---------------------------------
        df_chirurgie_ds=data[data["organisationunitname"].str.startswith("DS", na=False)]
        df_chirurgie_ds=traitement_donnees_ds(df_chirurgie_ds,code_annuaire_ds,correspondanceUID)
        df_chirurgie_ds=df_chirurgie_ds[["structures","Cas de Appendicectomie","Cesarienne","Chirurgie-Cure de hernies","Chirurgie-Hydrocèle","G.E.U",
        "Laparotomie pour occlusion intestinale","Autres laparotomies","Autres interventions essentielles en chirurgie"]]
        df_chirurgie_ds=total_ds_a_partirde_totalbf(df_chirurgie_ds,code_annuaire_ch, col_ref='UO_annuaire',
                                                    valeur_source='Burkina Faso', valeur_cible='Total District')
        df_chirurgie_ds=df_chirurgie_ds.drop("structures", axis=1)

        #------------------------chirurgie CH ---------------------------------
        df_donnees_ch=data[~data["organisationunitname"].str.startswith("DS", na=False)]
        df_donnees_ch=pd.merge(df_donnees_ch,correspondanceUID, left_on="dataid", right_on="Uid_endosBF", how="left")
        df_donnees_ch=df_donnees_ch[["orgunitlevel3","IndicateursAnnuaire","Total"]]

        chirurgie_ch_liste=["Chirurgie-Parage","Chirurgie-Ostéosynthèse","Chirurgie-Amputations","Chirurgie-Ablation du matériel d’ostéosynthèse",
        "Chirurgie-Pose de prothèse","Interventions en Traumatologie","Chirurgie-Chirurgie de l’abdomen","Interventions en chirurgie viscérale",
        "Chirurgie-Néphrectomie","Chirurgie-Cure de fistule vésico-vaginale","Chirurgie-Cure de cystocèle","Interventions en Urologie",
        "Chirurgie-Cure d’hématomes","Chirurgie-Cure de spina bifida","Chirurgie-Plaies cranio cérébrales","Interventions en Neuro Chirurgie",
        "Interventions sur utérus","Interventions sur annexes","Interventions sur le sein","Chirurgie-Autres interventions en gynéco-obstétrique",
        "Extraction dentaire","Blocage/contention maxillaire","Exerese de epilus","Chirurgie-Autres chirurgies dentaires maxillo-faciales",
        "Autres interventions chirurgicales CHR/CHU","Chirurgie-Cure de hernies","Chirurgie-Hydrocèle","Cesarienne","Chirurgie-Cataracte",
        "Interventions de Glaucome","Interventions pour Evisceration","Chirurgie-Trichiasis trachomateux","Intervention-Autres trichiasis",
        "Intervention pour Cure de Pterygion","Intervention pour Cure de Chalazion","Interventions pour autres Ophtalmologie",
        "Trachéotomie","Cheloidectomie cou et face","Polypeptectomie des fosses nasales","Chirurgie-Adenoidectomie","Chirurgie-Adeno-amygdalectomie",
        "Chirurgie-Amygdalectomie","Chirurgie-Thyroidectomie","Chirurgie-Autres interventions en ORL"]
        df_chirurgie_ch=df_donnees_ch[df_donnees_ch["IndicateursAnnuaire"].isin(chirurgie_ch_liste)]

        df_chirurgie_ch=df_chirurgie_ch.pivot_table(index="IndicateursAnnuaire",
                                                                columns="orgunitlevel3",
                                                                values="Total",aggfunc="sum"
                                                                ).reset_index()


        df_chirurgie_ch["total_chirurgie_ch"]=df_chirurgie_ch.select_dtypes(include='number').sum(axis=1)
        df_chirurgie_ch["CHR Manga"]=np.nan
        df_chirurgie_ch=df_chirurgie_ch[["IndicateursAnnuaire",'CHR Banfora','CHR Dédougou','CHU Bogodogo', 'CHU Pédiatrique CDG','CHU Tengandogo','CHU Yalgado','Hôpital Paul VI',
            'Hôpital Saint Camille de Ouagadougou', 'Hôpital SCHIPHRA','CHR Tenkodogo', 'CHR Kaya','CHR Koudougou',"CHR Manga","CHR Fada N'Gourma",
                                'CHU Sanou Souro','CHUR Ouahigouya', 'CHR Ziniaré','CHR Dori', 'CHR Gaoua','total_chirurgie_ch']]
        df_chirurgie_ch=pd.merge(chirurgie_ch, df_chirurgie_ch, left_on="chirurgie",right_on="IndicateursAnnuaire", how="left")

        #------------------------Soins d'urgences ---------------------------------
        df_donnees_ch_pivot=df_donnees_ch.pivot_table(index="orgunitlevel3",
                                                                columns="IndicateursAnnuaire",
                                                                values="Total",aggfunc="sum"
                                                                ).reset_index()

        df_donnees_ch_pivot=pd.merge(df_donnees_ch_pivot,code_ch_annuaire,left_on="orgunitlevel3", right_on="hopital", how="left")
        df_donnees_ch_pivot.sort_values("code",inplace=True)
        df_donnees_ch_pivot.drop(["code", "hopital"], axis=1, inplace=True)

        df_soins_urgence=df_donnees_ch_pivot[["orgunitlevel3","Patients reçus en urgence",
        "Patients reçus en urgence et mis en observation (MEO)",
        "Nombre de journees de mise en observation",
        "Séjour moyen (en jour) en MEO",
        "Patients référés/reçus en urgence",
        "Dont patients reçus en urgence/évacués d'une formation sanitaire privée",
        "Nombre total de patients reçus en urgence: transféré",
        "Autres (circuit non spécifié)",
        "Normale (exéat)",
        "Contre avis médical",
        "Sans avis médical",
        "Décédé en urgence",
        "Sortie Transféré dans un autre service",
        "Sortie Evacué vers un autre hôpital",]].copy()
        df_soins_urgence["patients mis en observation (%)"]=round((df_soins_urgence["Patients reçus en urgence et mis en observation (MEO)"])/(df_soins_urgence["Patients reçus en urgence"])*100,2)

        #------------------------------Données Tuberculose-----------------------------------------------------------------
        df_tb_region=data.copy()
        df_tb_region=traitement_donnees_ds_ch(df_tb_region,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_tb_region=df_tb_region[["structures","type_uo_x","Malades soumis au dépistage de TB","Malades dépistés positifs TB","Malades orientés par les acteurs communautaires soumis au dépistage de TB",
        "Malades orientés par les acteurs communautaires dépistés positifs TB","TB Lames lues pour le dépistage: Négatif","TB Lames lues pour le dépistage: Positif 1-9 B",
        "TB Lames lues pour le dépistage: Positif 1 à 3+","TB Lames lues pour le contrôle M2-M3: Négatif","TB Lames lues pour le contrôle M2-M3: Positif 1-9 B",
        "TB Lames lues pour le contrôle M2-M3: Positif 1 à 3+","TB Lames lues pour le contrôle M5-M6: Négatif","TB Lames lues pour le contrôle M5-M6: Positif 1-9 B",
        "TB Lames lues pour le contrôle M5-M6: Positif 1 à 3+","Total lames lues TB",

        "TB Total notifiés (Nouveaux cas et Rechutes)","Nbre de cas TB (tous les cas, toutes formes) testés au VIH","Nombre de cas TB (tous les cas, toutes formes) testé VIH+",
        "Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri","Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV",

        "Nbre de cas TB (tous les cas, toutes formes) testés au VIH Cohorte","Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte",
        "Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri Cohorte","Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV Cohorte" ]]
        # dataframe tb avec merge region
        df_tb_region=df_tb_region[df_tb_region["type_uo_x"].isin(["DRS","Total_bfa"])]
        df_tb_region_merge=pd.merge(region_bf,df_tb_region, left_on="UO_annuaire",right_on="structures",how="left")



        #-------------LITS D'HOSPITALISATION, JOURNEE HOSPI---------------------------------------------
        #FS
        df_lits_fs=df_lits_fs_extract.copy()
        df_lits_fs=df_lits_fs.drop(["orgunitlevel5","orgunitlevel6"], axis=1)
        #ch
        df_lits_ch=df_lits_ch_extract.copy()
        df_lits_ch=df_lits_ch[~df_lits_ch["organisationunitname"].str.startswith("DS", na=False)]
        df_lits_ch.head()
        # fusion
        df_lits_cma_ch=pd.concat([df_lits_fs,df_lits_ch], ignore_index=True)
        df_lits_cma_ch=ajout_variables_nulle(df_lits_cma_ch,annauire_ind8_lits_cma_ch)
        # merge avec les types de structures
        df_lits_cma_ch_merge=pd.merge(df_lits_cma_ch, type_statut_fs,
                                    left_on="organisationunitid",
                                    right_on="id_fs",
                                    how="left")

        df_lits_cma_ch=df_lits_cma_ch_merge[df_lits_cma_ch_merge["types"].isin(["Services_CMA","Unités CHR/CHU"])]
        df_lits_cma_ch=df_lits_cma_ch[['orgunitlevel1', 'orgunitlevel2', 'orgunitlevel4',"dataid",'dataname','Total']]
        df_lits_hospitalisation_cma_ch=traitement_donnees_ds_ch(df_lits_cma_ch,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_lits_hospitalisation_cma_ch.drop(["type_uo_y","type_uo_x"], axis=1, inplace=True)
        df_lits_cma_ch_chu=df_lits_hospitalisation_cma_ch[['Lits en suites de couches',"Lits d'hospitalisation des maternités","Lits d'hospitalisation des autres services"]]
        df_hospitalisation_cm_ch=df_lits_hospitalisation_cma_ch[["MaladesTotal sortie","Malades sortis Décédés","Nombre total des journées d'hospitalisations cm_ch"]]

        #----------------------Base de données additionnelles---------------------------------
        #données CH
        df_indicateurs_additionnels_ch=indicateurs_additionnels_extract_ch.copy()
        masque_ch=~df_indicateurs_additionnels_ch['organisationunitname'].str.startswith("DS",na=False)
        df_indicateurs_additionnels_ch=df_indicateurs_additionnels_ch[masque_ch]

        #données DS
        df_indicateurs_additionnels_fs=indicateurs_additionnels_extract_fs.copy()
        df_indicateurs_additionnels_fs.drop(['orgunitlevel5', 'orgunitlevel6'], axis=1, inplace=True)

        #fusion des tableau CH et DS
        df_indicateurs_additionnels=pd.concat([df_indicateurs_additionnels_fs,df_indicateurs_additionnels_ch], axis=0, ignore_index=True)

        df_indicateurs_additionnels_merge=pd.merge(df_indicateurs_additionnels,type_statut_fs,left_on="organisationunitid", right_on="id_fs", how="left")

        #traitement tableau integration des services et whopen
        df_integration_whopen=df_indicateurs_additionnels_merge.copy()
        masque_publique=df_integration_whopen["statut"]=="FS Public"
        df_integration_whopen=df_integration_whopen[masque_publique]

        df_integration_whopen=traitement_donnees_ds_ch(df_integration_whopen,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)

        # 3 TRAITEMENT NOSOLOGIES
        #--------------------------------------------------------------------------------------------------------------------------------------
        # LES NOSOLOGIES DE CONSULTATION ET D'HOSPITALISATION
        #--------------------------------------------------------------------------------------------------------------------------------------
        # nosologie de consultation

        df_nosologie=df_nosologie_extract.copy()
        df_nosologie_merge=pd.merge(df_nosologie, indicateurs_nosologies, left_on="dataid", right_on="uid", how="left")
        df_nosologie_merge = df_nosologie_merge.dropna(subset=["pathologies"])


        df_nosologie_merge["tranche_age"] = df_nosologie_merge["pathologies"].apply(
            lambda x: "< 5 ans" if "< 5 ans" in x
            else "5-14 ans" if "5-14 ans" in x
            else "15 ans et +" if "15 ans et +" in x
            else None
        )

        # On enlève les 3 motifs d'âges
        df_nosologie_merge["data_element"] = df_nosologie_merge["pathologies"].str.replace(r'\s*(< 5 ans|5-14 ans|15 ans et \+)', '', regex=True)

        df_nosologie_merge=df_nosologie_merge[['organisationunitname', 'data_element',  'tranche_age', 'Total']]

        ##Dépivotage des données afin d'avoir les DE en colonne
        data_consultation_age = df_nosologie_merge.pivot_table(
            index=["data_element"],
            columns="tranche_age",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        data_consultation_age=data_consultation_age[["data_element","< 5 ans","5-14 ans","15 ans et +"]]
        data_consultation_age_merge=pd.merge(code_nosologie_consultation,data_consultation_age,left_on="nosologie_consultation",right_on="data_element",how="left")
        data_consultation_age_merge=data_consultation_age_merge.drop(["ORDRE","code","data_element"], axis=1)
        data_consultation_age_merge.rename(columns={"nosologie_consultation":"Affection"}, inplace=True)
        df_nosologie_consul_tranche_age=data_consultation_age_merge.copy()

        # -----------------------nosologie par région-------------------------------------------------------------
        data_consultation_region=df_nosologie_merge[["organisationunitname","data_element","Total"]]

        #Dépivotage des données afin d'avoir les DE en colonne
        data_consultation_region_pivot = data_consultation_region.pivot_table(
            index=["data_element"],
            columns="organisationunitname",
            values="Total",
            aggfunc="sum"
        ).reset_index()
        data_consultation_region_pivot=pd.merge(code_nosologie_consultation,data_consultation_region_pivot,left_on="nosologie_consultation", right_on="data_element", how="left")
        df_nosologie_consul_region1=data_consultation_region_pivot[["nosologie_consultation",'Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako',
            'Nakambé', 'Nando']]
        df_nosologie_consul_region2=data_consultation_region_pivot[["nosologie_consultation",'Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou',
            'Tannounyan', 'Tapoa', 'Yaadga']]





        #-----------------les principaux motifs de consultation-------------------------------------------------
        principauxmotif_consultation=data_consultation_age.copy()
        principauxmotif_consultation["data_element"]=principauxmotif_consultation["data_element"].replace({"Paludisme simple":"Paludisme", "Paludisme grave":"Paludisme"})
        principauxmotif_consultation=principauxmotif_consultation.groupby("data_element", as_index=False).sum()

        principauxmotif_consultation["total"]=principauxmotif_consultation["< 5 ans"]+principauxmotif_consultation["5-14 ans"]+principauxmotif_consultation["15 ans et +"]
        total_noso_cons=principauxmotif_consultation["total"].sum()
        principauxmotif_consultation = principauxmotif_consultation[~principauxmotif_consultation["data_element"].str.contains("Autr", case=False, na=False)]

        principauxmotif_consultation["frequence"]=round(principauxmotif_consultation["total"]*100/total_noso_cons,1)
        principauxmotif_consultation=principauxmotif_consultation.sort_values("frequence",ascending=False)
        principauxmotif_consultation=principauxmotif_consultation.iloc[0:10,:].reset_index(drop=True)

        principauxmotif_consultation.head()

        #------------------Synthese nosologies consultation----------------------------------
        df_nosologie_consul_tranche_age=df_nosologie_consul_tranche_age.drop("Affection", axis=1)
        df_nosologie_consul_region1=data_consultation_region_pivot[['Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako',
            'Nakambé', 'Nando']]
        df_nosologie_consul_region2=data_consultation_region_pivot[['Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou',
            'Tannounyan', 'Tapoa', 'Yaadga']]


        #----------------------------------------------------------------------------------------------------
        #NOSOLOGIE DES HOSPITALISATIONS
        #----------------------------------------------------------------------------------------------------
        df_noso_hospitalisation=df_noso_hospitalisation.copy()
        df_noso_hospitalisation_merge=pd.merge(df_noso_hospitalisation, indicateurs_noso_hospitalisation, left_on="dataid", right_on="uid", how="left")


        # Traitement nosologie hospitalisation
        #-------------------------------------------------------------------------
        def detect_tranche(x):
            if pd.isna(x):
                return None
            elif "15 ans et +, Nombre de cas" in x:
                return "15 ans et +, Nombre de cas"
            elif "15 ans et +, Nombre de deces" in x:
                return "15 ans et +, Nombre de deces"
            elif "5-14 ans, Nombre de cas" in x:
                return "5-14 ans, Nombre de cas"
            elif "5-14 ans, Nombre de deces" in x:
                return "5-14 ans, Nombre de deces"
            elif "< 5 ans, Nombre de cas" in x:
                return "< 5 ans, Nombre de cas"
            elif "< 5 ans, Nombre de deces" in x:
                return "< 5 ans, Nombre de deces"
            else:
                return None
        df_noso_hospitalisation_merge["tranche_age"] = df_noso_hospitalisation_merge["noso_hospitalisation"].apply(detect_tranche)

        # On enlève les 3 motifs d'âges
        patterns = [
            r"< 5 ans, Nombre de cas",
            r"< 5 ans, Nombre de deces",
            r"5-14 ans, Nombre de cas",
            r"5-14 ans, Nombre de deces",
            r"15 ans et \+, Nombre de cas",
            r"15 ans et \+, Nombre de deces"
        ]

        df_noso_hospitalisation_merge["data_element"] = df_noso_hospitalisation_merge["noso_hospitalisation"]
        for pat in patterns:
            df_noso_hospitalisation_merge["data_element"] = df_noso_hospitalisation_merge["data_element"].str.replace(pat, "", regex=True)
        df_noso_hospitalisation_merge["data_element"] = df_noso_hospitalisation_merge["data_element"].str.strip()
        df_noso_hospit=df_noso_hospitalisation_merge[['organisationunitname', 'data_element',  'tranche_age', 'Total']]

        ##Dépivotage des données afin d'avoir les tranche d'age en colonne
        data_noso_hospit_age = df_noso_hospit.pivot_table(
            index=["data_element"],
            columns="tranche_age",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        data_noso_hospit_age=data_noso_hospit_age[["data_element","< 5 ans, Nombre de cas","< 5 ans, Nombre de deces",
        "5-14 ans, Nombre de cas","5-14 ans, Nombre de deces","15 ans et +, Nombre de cas","15 ans et +, Nombre de deces"]]
        df_noso_hospit_motif=data_noso_hospit_age.copy()
        data_noso_hospit_age=pd.merge(code_nosologie_hospitalisation, data_noso_hospit_age, left_on="nosologie_consultation", right_on="data_element", how="left")
        data_noso_hospit_age=data_noso_hospit_age.drop(["code","data_element","ORDRE"],axis=1)
        data_noso_hospit_age.rename(columns={"nosologie_consultation":"pathologies"}, inplace=True)


        #-------------------Principaux motif d'hospitalisation
        principauxmotif_hospitalisation=df_noso_hospit_motif[["data_element","< 5 ans, Nombre de cas","5-14 ans, Nombre de cas","15 ans et +, Nombre de cas"]].copy()
        principauxmotif_hospitalisation["total"]=principauxmotif_hospitalisation["< 5 ans, Nombre de cas"]+principauxmotif_hospitalisation["5-14 ans, Nombre de cas"]+principauxmotif_hospitalisation["15 ans et +, Nombre de cas"]
        total_noso_hospit_cas=principauxmotif_hospitalisation["total"].sum()
        principauxmotif_hospitalisation = principauxmotif_hospitalisation[~principauxmotif_hospitalisation["data_element"].str.contains("Autr", case=False, na=False)]
        principauxmotif_hospitalisation["frequence"]=round(principauxmotif_hospitalisation["total"]*100/total_noso_hospit_cas,1)
        principauxmotif_hospitalisation=principauxmotif_hospitalisation.sort_values("frequence",ascending=False)
        principauxmotif_hospitalisation=principauxmotif_hospitalisation.iloc[0:10,:].reset_index(drop=True)

        #-------------------Principaux motif de deces------------------------------------------------
        principauxmotif_hospi_deces=df_noso_hospit_motif[["data_element","< 5 ans, Nombre de deces","5-14 ans, Nombre de deces","15 ans et +, Nombre de deces"]].copy()
        principauxmotif_hospi_deces["total"] = principauxmotif_hospi_deces.select_dtypes(include="number").sum(axis=1)
        total_noso_hospit_deces=principauxmotif_hospi_deces["total"].sum()
        principauxmotif_hospi_deces = principauxmotif_hospi_deces[~principauxmotif_hospi_deces["data_element"].str.contains("Autr", case=False, na=False)]
        principauxmotif_hospi_deces["frequence"]=round(principauxmotif_hospi_deces["total"]*100/total_noso_hospit_deces,1)
        principauxmotif_hospi_deces=principauxmotif_hospi_deces.sort_values("frequence",ascending=False)
        principauxmotif_hospi_deces=principauxmotif_hospi_deces.iloc[0:10,:].reset_index(drop=True)




        #---------------------hospitalisation CAS par région
        data_hospit_cas=df_noso_hospit[df_noso_hospit["tranche_age"].isin (["< 5 ans, Nombre de cas","5-14 ans, Nombre de cas","15 ans et +, Nombre de cas"])]
        data_hospit_cas=data_hospit_cas.drop("tranche_age", axis=1)

        # depivoter et renomme
        data_hospit_cas = data_hospit_cas.pivot_table(
            index="data_element",
            columns="organisationunitname",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        regions=['Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako','Nakambé', 'Nando','Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou','Tannounyan', 'Tapoa', 'Yaadga']
        data_hospit_cas = data_hospit_cas.rename(
            columns={col: f"{col}-Cas" for col in regions})

        data_hospit_cas=pd.merge(code_nosologie_hospitalisation, data_hospit_cas, left_on="nosologie_consultation", right_on="data_element", how="left")
        data_hospit_cas=data_hospit_cas.drop(["code","data_element","ORDRE"],axis=1)
        data_hospit_cas.rename(columns={"nosologie_consultation":"pathologies"}, inplace=True)

        #---------------------hospitalisation deces par région
        data_hospit_deces=df_noso_hospit.query('tranche_age not in ["< 5 ans, Nombre de cas","5-14 ans, Nombre de cas","15 ans et +, Nombre de cas"]')
        data_hospit_deces=data_hospit_deces.drop("tranche_age", axis=1)
        # depivoter et renommer
        data_hospit_deces = data_hospit_deces.pivot_table(
            index="data_element",
            columns="organisationunitname",
            values="Total",
            aggfunc="sum"
        ).reset_index()

        regions=['Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako','Nakambé', 'Nando','Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou','Tannounyan', 'Tapoa', 'Yaadga']
        data_hospit_deces = data_hospit_deces.rename(
            columns={col: f"{col}-Deces" for col in regions})

        data_hospit_deces=pd.merge(code_nosologie_hospitalisation, data_hospit_deces, left_on="nosologie_consultation", right_on="data_element", how="left")
        data_hospit_deces=data_hospit_deces.drop(["code","data_element","ORDRE"],axis=1)
        data_hospit_deces.rename(columns={"nosologie_consultation":"pathologies"}, inplace=True)

        #---------------------hospitalisation CAS-DECES par région
        hospitregion_cas_deces=pd.merge(data_hospit_cas,data_hospit_deces, left_on="pathologies", right_on="pathologies", how="left")


        colonnes_fixees = ["pathologies"]
        colonnes_triees = sorted([c for c in hospitregion_cas_deces.columns if c not in colonnes_fixees])
        # Réorganisation
        hospitregion_cas_deces = hospitregion_cas_deces[colonnes_fixees + colonnes_triees]
        hospitregion_cas_deces1=hospitregion_cas_deces[['pathologies', 'Bankui-Cas', 'Bankui-Deces', 'Djoro-Cas', 'Djoro-Deces','Goulmou-Cas',
                                                        'Goulmou-Deces', 'Guiriko-Cas', 'Guiriko-Deces','Kadiogo-Cas', 'Kadiogo-Deces',
                                                        'Koulsé-Cas', 'Koulsé-Deces']]
        hospitregion_cas_deces2=hospitregion_cas_deces[['pathologies', 'Liptako-Cas', 'Liptako-Deces', 'Nakambé-Cas', 'Nakambé-Deces',
                                                    'Nando-Cas', 'Nando-Deces', 'Nazinon-Cas', 'Nazinon-Deces', 'Oubri-Cas',
                                                    'Oubri-Deces', 'Sirba-Cas', 'Sirba-Deces',]]

        hospitregion_cas_deces3=hospitregion_cas_deces[['pathologies','Soum-Cas', 'Soum-Deces','Sourou-Cas', 'Sourou-Deces',
                                                        'Tannounyan-Cas', 'Tannounyan-Deces','Tapoa-Cas', 'Tapoa-Deces',
                                                        'Yaadga-Cas', 'Yaadga-Deces']]

        # Traitement nosologie hospitalisation
        df_personnes_agees=personnes_agees_extract.copy()
        df_personnes_agees_merge=pd.merge(df_personnes_agees,Listes_personnes_agees_annuaire,left_on="dataid",right_on="dataid", how="left")
        df_personnes_agees_merge=df_personnes_agees_merge[['organisationunitname', 'data_element',  'tranche_age', 'Total']]

        # selon le sexe
        df_personnes_agees_sexe = df_personnes_agees_merge.pivot_table(
            index=["data_element"],
            columns="tranche_age",
            values="Total",
            aggfunc="sum"
        ).reset_index()
        df_personnes_agees_sexe=df_personnes_agees_sexe[['data_element',  'Cas Masculin','Déces Masculin','Cas Féminin','Déces Féminin']]
        


        # selon la région
        df_personnes_agees_region=df_personnes_agees_merge[['organisationunitname', 'data_element',  'tranche_age', 'Total']]
        df_personnes_agees_region_cas=df_personnes_agees_region[df_personnes_agees_region["tranche_age"].isin(['Cas Masculin','Cas Féminin'])]
        df_personnes_agees_region_cas = df_personnes_agees_region_cas.pivot_table(
            index=["data_element"],
            columns="organisationunitname",
            values="Total",
            aggfunc="sum"
        ).reset_index()
        regions=['Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako','Nakambé', 'Nando','Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou','Tannounyan', 'Tapoa', 'Yaadga']
        df_personnes_agees_region_cas = df_personnes_agees_region_cas.rename(
            columns={col: f"{col}-Cas" for col in regions})

        df_personnes_agees_region_deces=df_personnes_agees_region[~df_personnes_agees_region["tranche_age"].isin(['Cas Masculin','Cas Féminin'])]
        df_personnes_agees_region_deces = df_personnes_agees_region_deces.pivot_table(
            index=["data_element"],
            columns="organisationunitname",
            values="Total",
            aggfunc="sum"
        ).reset_index()
        regions=['Bankui','Djoro', 'Goulmou', 'Guiriko', 'Kadiogo', 'Koulsé', 'Liptako','Nakambé', 'Nando','Nazinon', 'Oubri', 'Sirba', 'Soum', 'Sourou','Tannounyan', 'Tapoa', 'Yaadga']
        df_personnes_agees_region_deces = df_personnes_agees_region_deces.rename(
            columns={col: f"{col}-Deces" for col in regions})

        df_personnes_agees_region_cas_deces=pd.merge(df_personnes_agees_region_cas,df_personnes_agees_region_deces,left_on="data_element",right_on="data_element", how="left" )

        colonnes_fixees = ["data_element"]
        colonnes_triees = sorted([c for c in df_personnes_agees_region_cas_deces.columns if c not in colonnes_fixees])
        # Réorganisation
        df_personnes_agees_region_cas_deces = df_personnes_agees_region_cas_deces[colonnes_fixees + colonnes_triees]

        df_personnes_agees_region_cas_deces1=df_personnes_agees_region_cas_deces[['data_element', 'Bankui-Cas', 'Bankui-Deces', 'Djoro-Cas', 'Djoro-Deces','Goulmou-Cas',
                                                        'Goulmou-Deces', 'Guiriko-Cas', 'Guiriko-Deces','Kadiogo-Cas', 'Kadiogo-Deces',
                                                        'Koulsé-Cas', 'Koulsé-Deces']]
        df_personnes_agees_region_cas_deces2=df_personnes_agees_region_cas_deces[['data_element', 'Liptako-Cas', 'Liptako-Deces', 'Nakambé-Cas', 'Nakambé-Deces',
                                                    'Nando-Cas', 'Nando-Deces', 'Nazinon-Cas', 'Nazinon-Deces', 'Oubri-Cas',
                                                    'Oubri-Deces', 'Sirba-Cas', 'Sirba-Deces',]]

        df_personnes_agees_region_cas_deces3=df_personnes_agees_region_cas_deces[['data_element','Soum-Cas', 'Soum-Deces','Sourou-Cas', 'Sourou-Deces',
                                                        'Tannounyan-Cas', 'Tannounyan-Deces','Tapoa-Cas', 'Tapoa-Deces',
                                                        'Yaadga-Cas', 'Yaadga-Deces']]
        

        #----INDICATEURS CALCULES ANNUAIRE STATISTIQUE

        #--------------------------------CALCUL DES INDICATEURS
        df=data_ds.copy()
        indicateurs=pd.DataFrame({
            "Taux (%) d'abandon DTC-HepB-Hib1 et DTC-HepB-Hib3":np.where(df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib1"] != 0,
            round((df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib1"]-df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib3"])
            /df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib1"]*100,2),np.nan),

            "Couverture vaccinale en BCG":np.where(df["Naissances vivantes attendues"] != 0,
                                                round(df["Nombre d'enfants de 0-11 mois ayant reçu le BCG"]/df["Naissances vivantes attendues"]*100,2),np.nan),

            "Couverture vaccinale en Hep B":np.where(df["Naissances vivantes attendues"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le HepB"]/df["Naissances vivantes attendues"]*100,2),np.nan),


            "Proportion d'enfants ayant reçu Hep B dans les 24H":np.where(df["Naissances vivantes attendues"] != 0,
                                                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le Hep B dans les 24H"]/df["Naissances vivantes attendues"]*100,2),np.nan),

            "Couverture vaccinale en VPO 1":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le VPO 1"]/df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en DTC-HepB-Hib1":np.where(df["Population 0 à 11 mois"] != 0,
                                                            round(df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib1"]/df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en Pneumo 1":np.where(df["Population 0 à 11 mois"] != 0,
                                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le Pneumo 1"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en Rota 1":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le Rota 1"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en VPO 3":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le VPO 3"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en DTC-HepB-Hib3":np.where(df["Population 0 à 11 mois"] != 0,
                                                            round(df["Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib3"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en Pneumo 3":np.where(df["Population 0 à 11 mois"] != 0,
                                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le Pneumo 3"] / df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en Rota 3":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le Rota 3"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en VPI1":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le VPI1"] / df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en VPI2":np.where(df["Population 0 à 11 mois"] != 0,
                                                    round(df["Nombre d'enfants de 0-11 mois ayant reçu le VPI2"] / df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en RR1":np.where(df["Population 0 à 11 mois"] != 0,
                                                round(df["Nombre d'enfants de 0-11 mois ayant reçu le RR1"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en VAA":np.where(df["Population 0 à 11 mois"] != 0,
                                                round(df["Nombre d'enfants de 0-11 mois ayant reçu le VAA"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture vaccinale en MenA":np.where(df["Population 12 à 23 Mois"] != 0,
                                                round(df["Nombre d'enfants de 12-23 mois ayant reçu le MenA"]/ df["Population 12 à 23 Mois"]*100,2),np.nan),

            "Couverture vaccinale en RR2":np.where(df["Population 12 à 23 Mois"] != 0,
                                                round(df["Nombre d'enfants de 12-23 mois ayant reçu le RR2"]/ df["Population 12 à 23 Mois"]*100,2),np.nan),

            "Couverture vaccinale en HPV1":np.where(df["Fille de 9 ans"] != 0,
                                                round(df["Nombre de filles de 9 ans ayant reçu le HPV1"]/ df["Fille de 9 ans"]*100,2),np.nan),

            "Couverture vaccinale des femmes enceintes en Td 2 et +":np.where(df["Grossesses attendues"] != 0,
                                                                            round(df["Nombre de femmes enceintes ayant recu le Td 2 et plus et femmes completement vaccinées"]
                                                                            /df["Grossesses attendues"]*100,2),np.nan),
            "Couverture VAP1":np.where(df["Population 0 à 11 mois"] != 0,
                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le VAP1"]/df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture VAP2":np.where(df["Population 0 à 11 mois"] != 0,
                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le VAP2"]/ df["Population 0 à 11 mois"]*100,2),np.nan),

            "Couverture VAP3":np.where(df["Population 0 à 11 mois"] != 0,
                                        round(df["Nombre d'enfants de 0-11 mois ayant reçu le VAP3"]/df["Population 0 à 11 mois"]*100,2),np.nan),

            "Taux d'utilisation des méthodes contraceptives":np.where(df["Femmes en Age de procréer"] != 0,
                                                                    round(df["Utilisatrices PF totales"] /df["Femmes en Age de procréer"]*100,2),np.nan),

            "Couverture en CPN1 (%)":np.where(df["Grossesses attendues"] != 0,
                                            round(df["Femmes vues en CPN1"]/ df["Grossesses attendues"]*100,2),np.nan),

            "Couverture en CPN4 (%)":np.where(df["Grossesses attendues"] != 0,
                                            round(df["Femmes vues en CPN4"]/ df["Grossesses attendues"]*100,2),np.nan),

            "Proportion des femmes ayant bénéficié d'un TPI3 (%)":np.where(df["Femmes vues en CPN1"] != 0,
                                                                        round(df["Nombre de femmes ayant bénéficié d'un TPI3"]
                                                                        / df["Femmes vues en CPN1"]*100,2),np.nan),

            "Proportion de femmes vues en CPN ayant reçu un counseling nutritionnel (%)":np.where(df["femmes vues en CPN"] != 0,
                                                                                                round(df["Femmes vues en CPN ayant reçu un counseling nutritionnel"]
                                                                                                / df["femmes vues en CPN"]*100,2),np.nan),

            "Proportion de CNS de 0 à 23 mois dont la mère/accompagnant a reçu un counseling sur l’ANJE":np.where(df["Enfants (nouveaux +anciens) vus en consultation nourrissons sains 0 à 23 mois"] != 0,
                                                                                                round(df["Enfant de 0 à 23 mois dont la mère/accompagnant a reçu un counseling sur l’ANJE"]
                                                                                                / df["Enfants (nouveaux +anciens) vus en consultation nourrissons sains 0 à 23 mois"]*100,2),np.nan),

            "Proportion (%) des grossesses référées":np.where(df["Grossesses attendues"] != 0,
                                                            round(df["Grossesses référées"] / df["Grossesses attendues"]*100,2),np.nan),

            "Taux d'accouchement dans les FS (%)":np.where(df["Accouchements attendus"] != 0,
                                                        round(df["Total accouchement"]/df["Accouchements attendus"]*100,2),np.nan),

            "Proportion (%) des accouchements réalisés avec partogramme":np.where(df["Total accouchement"] != 0,
                                                                                round(df["Accouchement avec partogramme"]/ df["Total accouchement"]*100,2),np.nan),

            "Proportion (%) de mort-nés frais parmi les mort-nés":np.where((df["Mort né Frais"]+df["Mort né Macéré"]) != 0,
                                                                        round(df["Mort né Frais"]/ (df["Mort né Frais"]+df["Mort né Macéré"])*100,2),np.nan),

            "Proportion de faible poids de naissance":np.where(df["Total Naissance vivante"] != 0,
                                                            round(df["Moins de 2500 g"]/ df["Total Naissance vivante"]*100,2),np.nan),

            "Proportion (%) d'enfants mis au sein à la première heure":np.where(df["Total Naissance vivante"] != 0,
                                                                                round(df["Enfants mis au sein à la première heure"]/ df["Total Naissance vivante"]*100,2),np.nan),

            "Proportion de décès néonatal parmis les naissances vivantes":np.where(df["Total Naissance vivante"] != 0,
                                                                                round(df["Total décès néo-natal"]/ df["Total Naissance vivante"]*100,2),np.nan),

            "Couverture (%) en consultation posnatale 7 ème jour":np.where(df["Femmes ayant accouché"] != 0,
                                                                        round(df["Femmes vues en consultation postnatale 7 ème jour"]/ df["Femmes ayant accouché"]*100,2),np.nan),

            "Couverture (%) en consultation posnatale 42 ème jour":np.where(df["Femmes ayant accouché"] != 0,
                                                                            round(df["Femmes vues en consultation postnatale 42 ème jour"]/df["Femmes ayant accouché"]*100,2),np.nan),

            "Taux de décès pour 100 000 parturientes":np.where(df["Femmes ayant accouché"] != 0,
                                                            round(df["Total décès maternels toutes causes confondues"]/ df["Femmes ayant accouché"]*100000,2),np.nan),

            "Proportion (%) de décès maternel audité":np.where(df["Total décès maternels toutes causes confondues"] != 0,
                                                            round(df["Nombre de décès maternel audité"]/ df["Total décès maternels toutes causes confondues"]*100,2),np.nan),

            "Taux de dépistage PTME VIH":np.where(df["Femmes vues en CPN1"] != 0,
                                                round(df["Nombre de femmes vues en CPN au cours du mois et ayant bénéficié d’un test VIH"]/ df["Femmes vues en CPN1"]*100,2),np.nan),

            "Taux de séropositivité PTME":np.where(df["Nombre de femmes vues en CPN au cours du mois et ayant bénéficié d’un test VIH"] != 0,
                                                round(df["Nombre de femmes enceintes dépistées VIH+"]
                                                / df["Nombre de femmes vues en CPN au cours du mois et ayant bénéficié d’un test VIH"]*100,2),np.nan),

            "% Enfants nés vivant qui reçoivent une prophylaxie par ARV":np.where(df["Enfants nés vivant de mères séropositives"] != 0,
                                                                                round(df["Enfants nés vivant qui reçoivent une prophylaxie par ARV"]
                                                                                / df["Enfants nés vivant de mères séropositives"]*100,2),np.nan),

            "Taux de positivité à la PCR des nourrissons nés de mère séropositive à la naissance":np.where(df["Nourrissons nés de mères séropositives ayant subi un test virologique du VIH (PCR) à la naissance"] != 0,
                                                                                                        round(df["Nourrissons ayant un test virologique du VIH (PCR) positif à la naissance"]
                                                                                                        / df["Nourrissons nés de mères séropositives ayant subi un test virologique du VIH (PCR) à la naissance"]*100,2),np.nan),

            "Taux de positivité au TDR des enfants nés de mère séropositive":np.where(df["Enfants nés de mères séropositives ayant subi un test sérologique du VIH (TDR)"] != 0,
                                                                                    round(df["Enfants ayant un test sérologique du VIH (TDR) positif"] / df["Enfants nés de mères séropositives ayant subi un test sérologique du VIH (TDR)"]*100,2),np.nan),

            "% Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs":np.where(df["Femmes vues en CPN1"] != 0,
                                                                                        round(df["Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs"]
                                                                                    / df["Femmes vues en CPN1"]*100,2),np.nan),

            "Taux de positivité des femmes vues en CPN testées à l'hépatite B":np.where(df["Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs"] != 0,
                                                                                    round(df["Femmes vues en CPN testées positives à l'hépatite B"]
                                                                                        / df["Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs"]*100,2),np.nan),

            "Taux de mise sous TARV des femmes vues en CPN testées positives à l'hépatite B":np.where(df["Femmes vues en CPN testées positives à l'hépatite B"] != 0,
                                                                                                    round(df["Femmes vues en CPN testées positives à l'hépatite B mises sous TARV"]
                                                                                                    / df["Femmes vues en CPN testées positives à l'hépatite B"]*100,2),np.nan),

            "% Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis":np.where(df["Femmes vues en CPN1"] != 0,
                                                                                                            round(df["Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis"]
                                                                                                            / df["Femmes vues en CPN1"]*100,2),np.nan),

            "Taux de positivité des femmes vues en CPN testées à la sypjilis":np.where(df["Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis"] != 0,
                                                                                    round(df["Femmes enceintes vues en CPN ayant été testées positives à la syphilis"]
                                                                                    / df["Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis"]*100,2),np.nan),

            "Taux de mise sous traitement des femmes vues en CPN testées positives à la syphilis":np.where(df["Femmes enceintes vues en CPN ayant été testées positives à la syphilis"] != 0,
                                                                                                        round(df["Femmes vues en CPN testées positives à la syphilis et mises sous traitement"] /
                                                                                                        df["Femmes enceintes vues en CPN ayant été testées positives à la syphilis"]*100,2),np.nan),
            "Taux de dépistage des MAM":np.where(df["Cas attendus de MAM"] != 0,
                                                round(df["Dépistage de la Malnutris aigüs modérés"]/df["Cas attendus de MAM"]*100,2),np.nan),

            "Taux de dépistage des MAS":np.where(df["Cas attendus de MAS"] != 0,
                                                round(df["Dépistage de la Malnutris aigüs sévères"]/df["Cas attendus de MAS"]*100,2),np.nan),

            "Taux de confirmation des MAM":np.where(df["MAM confirmé par la FS"] != 0,
                                                    round(df["MAM reçus des ASBC"]/df["MAM confirmé par la FS"]*100,2),np.nan),

            "Taux de confirmation des MAS":np.where(df["MAS confirmé par la FS"] != 0,
                                                    round(df["MAS reçus des ASBC"]/df["MAS confirmé par la FS"]*100,2),np.nan),

            "MAM Guéris (%)":np.where(df["MAM Total des Sorties"] != 0,
                                    round(df["MAM Guéris"]/ df["MAM Total des Sorties"]*100,2),np.nan),

            "MAM Décédé (%)":np.where(df["MAM Total des Sorties"] != 0,
                                    round(df["MAM Décédés"]/ df["MAM Total des Sorties"]*100,2),np.nan),

            "MAM Abandon (%)":np.where(df["MAM Total des Sorties"] != 0,
                                    round(df["MAM Abandons"] /df["MAM Total des Sorties"]*100,2),np.nan),

            "MAS ambulatoire Guéris (%)":np.where(df["MAS ambulatoire Total des Sorties"] != 0,
                                                round(df["MAS ambulatoire Guéris"] /df["MAS ambulatoire Total des Sorties"]*100,2),np.nan),

            "MAS ambulatoire Décédé (%)":np.where(df["MAS ambulatoire Total des Sorties"] != 0,
                                                round(df["MAS ambulatoire Décédés"] /df["MAS ambulatoire Total des Sorties"]*100,2),np.nan),

            "MAS ambulatoire Abandon (%)":np.where(df["MAS ambulatoire Total des Sorties"] != 0,
                                                round(df["MAS ambulatoire Abandons"] /df["MAS ambulatoire Total des Sorties"]*100,2),np.nan),

            "MAS interne Guéris (%)":np.where(df["MAS interne Total des Sorties"] != 0,
                                            round(df["MAS interne Guéris"] /df["MAS interne Total des Sorties"]*100,2),np.nan),

            "MAS interne Décédé (%)":np.where(df["MAS interne Total des Sorties"] != 0,
                                            round(df["MAS interne Décédés"] /df["MAS interne Total des Sorties"]*100,2),np.nan),

            "MAS interne Abandon (%)":np.where(df["MAS interne Total des Sorties"] != 0,
                                            round(df["MAS interne Abandons"] /df["MAS interne Total des Sorties"]*100,2),np.nan),

            "MAS globale Guéris (%)":np.where((df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"]) != 0,
                                            round((df["MAS ambulatoire Guéris"]+df["MAS interne Guéris"]) /
                                            (df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"])*100,2),np.nan),

            "MAS globale Décédé (%)":np.where((df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"]) != 0,
                                            round((df["MAS ambulatoire Décédés"]+df["MAS interne Décédés"]) /
                                            (df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"])*100,2),np.nan),

            "MAS globale Abandon (%)":np.where((df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"]) != 0,
                                            round((df["MAS ambulatoire Abandons"]+df["MAS interne Abandons"]) /
                                            (df["MAS ambulatoire Total des Sorties"]+df["MAS interne Total des Sorties"])*100,2),np.nan),

            "Nombre de contacts/habt/an":np.where(df["Population totale"] != 0,
                                                round(df["Nouveaux consultants Ensemble"] /df["Population totale"],2),np.nan),

            "Nombre de contacts/habt/an des 0 à 59 mois":np.where(df["Population de moins de 5 ans"] != 0,
                                                                round(df["Nouveaux consultants Moins de 5 ans"] /
                                                                df["Population de moins de 5 ans"]*1,2),np.nan),

            "Paludisme Taux de confirmation (%)":np.where(df["Cas suspect de Paludisme"] != 0,
                                                        round((df["TDR Réalisées"]+df["Gouttes épaisses Réalisées"]) /
                                                        df["Cas suspect de Paludisme"]*100,2),np.nan),

            "Paludisme Taux de positivité (%)":np.where((df["TDR Réalisées"]+df["Gouttes épaisses Réalisées"]) != 0,
                                                        round((df["TDR Positives"]+df["Gouttes épaisses Positives"]) /
                                                        (df["TDR Réalisées"]+df["Gouttes épaisses Réalisées"])*100,2),np.nan),

            "Incidence du paludisme chez les moins de 5 ans (p.1000)":np.where(df["Population de moins de 5 ans"] != 0,
                                                                            round(df["Cas de paludisme chez les moins de 5ans"] /
                                                                            df["Population de moins de 5 ans"]*1000,2),np.nan),

            "Incidence du paludisme dans la population générale (p.1000)":np.where(df["Population totale"] != 0,
                                                                                round(df["Total cas de paludisme dans la population générale"] /
                                                                                df["Population totale"]*1000,2),np.nan),

            "Incidence du paludisme confirmé chez les moins de 5 ans (p.1000)":np.where(df["Population de moins de 5 ans"] != 0,
                                                                                        round(df["Paludisme confirmé chez les moins de 5ans"] /
                                                                                        df["Population de moins de 5 ans"]*1000,2),np.nan),


            "Incidence du paludisme confirmé dans la population générale (p.1000)":np.where(df["Population totale"] != 0,
                                                                                                round(df["Paludisme confirmé dans la population générale"] /
                                                                                                df["Population totale"]*1000,2),np.nan),

            "Proportion des cas de paludisme simple traités aux ACT":np.where(df["Paludisme simple (confirmés+présumés) Total"] != 0,
                                                                                round(df["Paludisme simple traité avec ACT"] /
                                                                                df["Paludisme simple (confirmés+présumés) Total"]*100,2),np.nan),

            "Proportion des cas de paludisme simple confirmé traités aux ACT":np.where(df["Paludisme simple confirmés Total"] != 0,
                                                                                        round(df["Paludisme simple confirmé traité avec ACT"] /
                                                                                        df["Paludisme simple confirmés Total"]*100,2),np.nan),

            "Létalité du paludisme globale (%)":np.where(df["Paludisme grave (confirmés+présumés) Total"] != 0,
                                                            round(df["Décès paludisme global"] /
                                                            df["Paludisme grave (confirmés+présumés) Total"]*100,2),np.nan),

            "Létalité chez les moins de 5 ans (%)":np.where(df["Paludisme grave (confirmés+présumés) Moins de 5 ans"] != 0,
                                                                round(df["Décès paludisme chez les moins de 5 ans"] /
                                                                df["Paludisme grave (confirmés+présumés) Moins de 5 ans"]*100,2),np.nan),

            "Létalité chez les femmes enceintes (%)":np.where(df["Paludisme grave (confirmés+présumés) femmes enceintes"] != 0,
                                                                round(df["Décès paludisme chez les femmes enceintes"] /
                                                                df["Paludisme grave (confirmés+présumés) femmes enceintes"]*100,2),np.nan),

            "Couverture en MILDA femmes enceintes":np.where(df["Femmes vues en CPN1"] != 0,
                                                                round(df["MILDA femmes enceintes"] /
                                                                df["Femmes vues en CPN1"]*100,2),np.nan),

            "Couverture en MILDA chez les 0 à 11 mois":np.where(df["Total Naissance vivante"] != 0,
                                                                    round(df["MILDA chez les 0 à 11 mois"] /
                                                                    df["Total Naissance vivante"]*100,2),np.nan),

            "Incidence cumulée des IST (p1000)":np.where(df["Population totale"] != 0,
                                                            round(df["Total IST (Notification syndromique)"] /
                                                            df["Population totale"]*1000,2),np.nan),

            "Taux de positivité du VIH en milieu de soins":np.where(df["Dépistage VIH Personnes testées"] != 0,
                                                                        round(df["Dépistage VIH Personnes testées positives"] /
                                                                        df["Dépistage VIH Personnes testées"]*100,2),np.nan),

            "Taux de positivité pour HVB":np.where(df["Dépistage des hépatites Personnes testées pour HVB"] != 0,
                                                    round(df["Dépistage des hépatites Personnes testées positives pour HVB"] /
                                                    df["Dépistage des hépatites Personnes testées pour HVB"]*100,2),np.nan),

            "Taux de positivité pour HVC":np.where(df["Dépistage des hépatites Personnes testées pour HVC"] != 0,
                                                    round(df["Dépistage des hépatites Personnes testées positives pour HVC"] /
                                                    df["Dépistage des hépatites Personnes testées pour HVC"]*100,2),np.nan),

            "% de PVVIH sous TAR avec CV indétectable à 12 mois":np.where(df["PVVIH sous TAR ayant bénéficié du dosage de la CV à 12 mois de traitement"] != 0,
                                                                            round(df["PVVIH sous TAR dont la CV à 12 mois est indétectable"] /
                                                                            df["PVVIH sous TAR ayant bénéficié du dosage de la CV à 12 mois de traitement"]*100,2),np.nan),

            "Taux de notification TB (p.100000)":np.where(df["Population totale"] != 0,
                                                            round(df["TB Total notifiés (Nouveaux cas et Rechutes)"] /
                                                            df["Population totale"]*100000,2),np.nan),

            "% cas TB (tous les cas, toutes formes) testés au VIH":np.where(df["TB Total notifiés (Nouveaux cas et Rechutes)"] != 0,
                                                                                round(df["Nbre de cas TB (tous les cas, toutes formes) testés au VIH"] /
                                                                                df["TB Total notifiés (Nouveaux cas et Rechutes)"]*100,2),np.nan),

            "% cas TB (tous les cas, toutes formes) testé VIH+":np.where(df["Nbre de cas TB (tous les cas, toutes formes) testés au VIH"] != 0,
                                                                            round(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"] /
                                                                            df["Nbre de cas TB (tous les cas, toutes formes) testés au VIH"]*100,2),np.nan),

            "% de cas TB (tous les cas, toutes formes) VIH+ sous Cotri":np.where(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"] != 0,
                                                                                    round(df["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri"] /
                                                                                    df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"]*100,2),np.nan),

            "% de cas TB (tous les cas, toutes formes) VIH+ sous ARV":np.where(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"] != 0,
                                                                                round(df["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV"] /
                                                                                df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"]*100,2),np.nan),

            "Taux de positivité (%) au VIH Cohorte":np.where(df["Nbre de cas TB (tous les cas, toutes formes) testés au VIH Cohorte"] != 0,
                                                                round(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"] /
                                                                df["Nbre de cas TB (tous les cas, toutes formes) testés au VIH Cohorte"]*100,2),np.nan),

            "% de cas TB (tous les cas, toutes formes) VIH+ sous Cotri Cohorte":np.where(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"] != 0,
                                                                                            round(df["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri Cohorte"] /
                                                                                            df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"]*100,2),np.nan),

            "% de cas TB (tous les cas, toutes formes) VIH+ sous ARV Cohorte":np.where(df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"] != 0,
                                                                                        round(df["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV Cohorte"] /
                                                                                        df["Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"]*100,2),np.nan),

            "Résultas de traitement TB (tous les cas, toutes formes): Taux de succes au traitement":np.where(df["Malades TB (tous les cas, toutes formes) analysés"] != 0,
                                                                                                                round((df["Résultas de traitement TB (tous les cas, toutes formes): Guéri"]+df["Résultas de traitement TB (tous les cas, toutes formes):Traitement terminé"]) /
                                                                                                                df["Malades TB (tous les cas, toutes formes) analysés"]*100,2),np.nan),

            "Résultas de traitement TB (tous les cas, toutes formes): Taux d'échec":np.where(df["Malades TB (tous les cas, toutes formes) analysés"] != 0,
                                                                                                round(df["Résultas de traitement TB (tous les cas, toutes formes): Echec"] /
                                                                                                df["Malades TB (tous les cas, toutes formes) analysés"]*100,2),np.nan),

            "Résultas de traitement TB (tous les cas, toutes formes): Taux de décès":np.where(df["Malades TB (tous les cas, toutes formes) analysés"] != 0,
                                                                                                round(df["Résultas de traitement TB (tous les cas, toutes formes): Décédé"] /
                                                                                                df["Malades TB (tous les cas, toutes formes) analysés"]*100,2),np.nan),

            "Résultas de traitement TB (tous les cas, toutes formes): Taux de perdue de vue":np.where(df["Malades TB (tous les cas, toutes formes) analysés"] != 0,
                                                                                                        round(df["Résultas de traitement TB (tous les cas, toutes formes): Perdu de vue"] /
                                                                                                        df["Malades TB (tous les cas, toutes formes) analysés"]*100,2),np.nan),

            "Résultats de traitement TB (tous les cas, toutes formes): Taux de non évalué":np.where(df["Malades TB (tous les cas, toutes formes) analysés"] != 0,
                                                                                                        round(df["Résultas de traitement TB (tous les cas, toutes formes): Non évalué"] /
                                                                                                        df["Malades TB (tous les cas, toutes formes) analysés"]*100,2),np.nan),
            "(%) de patients mis en observation":np.where(df["Patients reçus en urgence"] != 0,
                                                            round(df["Patients reçus en urgence et mis en observation (MEO)"] /
                                                            df["Patients reçus en urgence"]*100,2),np.nan),

            "% de Perdues de vue recherchées et retrouvées pour CPN _com":np.where(df["Perdues de vue recherchées pour CPN_com"] != 0,
                                                                                    round(df["Perdues de vue retrouvées pour CPN _com"] /
                                                                                    df["Perdues de vue recherchées pour CPN_com"]*100,2),np.nan),

            "Taux de présences GASPA-Femmes Enceintes":np.where(df["GASPA Femmes enceintes-Femmes inscrites"] != 0,
                                                                    round(df["GASPA Femmes enceintes-Femmes présentes"] /
                                                                    df["GASPA Femmes enceintes-Femmes inscrites"]*100,2),np.nan),

            "Taux de présences GASPA-mères d’enfants de moins de 6 mois":np.where(df["GASPA mères d'enfants de moins de 6 mois-Femmes inscrites"] != 0,
                                                                                    round(df["GASPA mères d'enfants de moins de 6 mois-Femmes présentes"] /
                                                                                    df["GASPA mères d'enfants de moins de 6 mois-Femmes inscrites"]*100,2),np.nan),

            "Taux de présences GASPA-mères d’enfants de 6-23 mois":np.where(df["GASPA mères d’enfants de 6-23 mois -Femmes inscrites"] != 0,
                                                                                round(df["GASPA mères d’enfants de 6-23 mois -Femmes présentes"] /
                                                                                df["GASPA mères d’enfants de 6-23 mois -Femmes inscrites"]*100,2),np.nan)

        })
        """
        """

        #df["Prévalence (p10000)"]=np.where(df["Population totale"] != 0,
        #                                  (df["Lèpre-malades en fin d'année Total masculin"]+df["Lèpre-malades en fin d'année Total feminin"]) /
        #                                   df["Population totale"]*10000,np.nan)
        df["Proportion d'infirmité(%)"]=np.where((df["Nouveaux cas de lèpre Total masculin"]+df["Nouveaux cas de lèpre Total feminin"]) != 0,
                                                (df["Lèpre-infirmité degré 2 Total masculin"]+df["Lèpre-infirmité degré 2 Total feminin"]) /
                                                (df["Nouveaux cas de lèpre Total masculin"]+df["Nouveaux cas de lèpre Total feminin"])*100,np.nan)
        df["Taux de détection (p100000)"]=np.where(df["Population totale"] != 0,
                                                (df["Nouveaux cas de lèpre Total masculin"]+df["Nouveaux cas de lèpre Total feminin"]) /
                                                df["Population totale"]*100000,np.nan)
        df["Taux d'infirmité degré 2 (p1 000 000)"]=np.where(df["Population totale"] != 0,
                                                            (df["Lèpre-infirmité degré 2 Total masculin"]+df["Lèpre-infirmité degré 2 Total feminin"]) /
                                                            df["Population totale"]*100000,np.nan)
        df["Couverture (%) thérapeutique du TDM Schisto"]=np.where(df["Population Cible Schisto"] != 0,
                                                                df["Population traitée pour Schisto Total"] /
                                                                df["Population Cible Schisto"]*100,np.nan)
        df["Couverture (%) thérapeutique oncho Tournée 1"]=np.where(df["Population totale nécessitant traitement Onchocercose"] != 0,
                                                                    df["Population traitée pour l'onchocercose Tournée 1 total"] /
                                                                    df["Population totale nécessitant traitement Onchocercose"]*100,np.nan)

        df = pd.concat([df, indicateurs], axis=1)
        df_final=df.copy()

        #----------------------DATASET FINALE (DF)------------------
        df_ds=df_final[df_final["type_uo"].isin(["DS","DRS","Total_bfa"])]
        df=df_final.drop(["structures","type_uo"], axis=1) # df final des indicateurs avec les niveaux région, ch et ds
        df_ds=df_ds.drop(["structures","type_uo"], axis=1) # df final des indicateurs avec les niveaux région et ds


        # CONSTRUCTION DES TABLEAUX DE DONNEES ANNUAIRE
        # Tableau de complétude
        completude_fs["Rapports_saisis_a_temps"]=completude_fs['Publique-Rapports_saisis_a_temps']+completude_fs["Privee-Rapports_saisis_a_temps"]
        completude_pub=completude_fs[['Publique-Rapports_attendus','Publique-Rapports_saisis']]
        completude_privee=completude_fs[['Privee-Rapports_attendus','Privee-Rapports_saisis']]
        promptitude=completude_fs[["Rapports_saisis_a_temps"]]
        df_completude_asbc=dfc_asbc_fusion[['ASBC-Rapports_attendus','ASBC-Rapports_saisis', 'ASBC-Rapports_saisis_a_temps']]


        ## creation des tableaux du document
        populationTableau1_05=data_populationfinal[["Population de moins de 5 ans","Population 5-14 ans","Population 15 ans et plus",
                                                    "Population totale"]]
        population_cibleTableau1_06=data_populationfinal[["Femmes en Age de procréer","Grossesses attendues","Accouchements attendus",
                                                        "Naissances vivantes attendues","Population 0 à 11 mois","Population 6 à 11 Mois",
                                                                                                        "Population 12 à 59 Mois","Cas attendus de MAM","Cas attendus de MAS",
                                                                                    "Population 12 à 23 Mois","Fille de 9 ans"]]

        Vaccination_Tableau_4_01=df[["Nombre d'enfants de 0-11 mois ayant reçu le BCG","Nombre d'enfants de 0-11 mois ayant reçu le HepB",
                                    "Nombre d'enfants de 0-11 mois ayant reçu le Hep B dans les 24H","Nombre d'enfants de 0-11 mois ayant reçu le VPO 1",
                                    "Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib1","Nombre d'enfants de 0-11 mois ayant reçu le Pneumo 1",
                                    "Nombre d'enfants de 0-11 mois ayant reçu le Rota 1","Nombre d'enfants de 0-11 mois ayant reçu le VPO 3",
                                    "Nombre d'enfants de 0-11 mois ayant reçu le DTC-HepB-Hib3"]]

        Vaccination_Tableau_4_02=df[["Nombre d'enfants de 0-11 mois ayant reçu le Pneumo 3","Nombre d'enfants de 0-11 mois ayant reçu le Rota 3",
                                "Nombre d'enfants de 0-11 mois ayant reçu le VPI1","Nombre d'enfants de 0-11 mois ayant reçu le VPI2",
                                "Nombre d'enfants de 0-11 mois ayant reçu le RR1","Nombre d'enfants de 0-11 mois ayant reçu le VAA",
                                "Nombre d'enfants de 12-23 mois ayant reçu le MenA","Nombre d'enfants de 12-23 mois ayant reçu le RR2","Nombre d'enfants de 0-11 mois ayant reçu le VTC"]]

        Vaccination_Tableau_4_03=df[["Nombre de filles de 9 ans ayant reçu le HPV1",
                                    "Nombre de femmes enceintes ayant recu le Td 2 et plus et femmes completement vaccinées"]]



        PF_Tableau_4_06=df[["Nouvelles utilisatrices COC","Nouvelles utilisatrices COP","Nouvelles utilisatrices Depoprovéra",
                            "Nouvelles utilisatrices Sayanapress","Nouvelles utilisatrices implanon","Nouvelles utilisatrices jadelle",
                            "Nouvelles utilisatrices Levoplant","Nouvelles utilisatrices Levonorgestrel","Nouvelles utilisatrices Sterilisation feminine",
                            "Nouveaux utilisateurs Sterilisation masculine","Nouvelles utilisatrices DIU","Nouveaux utilisateurs Condoms masculin",
                            "Nouvelles utilisatrices Condoms féminin","Nouvelles utilisatrices Collier","Nouvelles utilisatrices MAMA","Nouvelles utilisatrices Totales"]]

        PF_Tableau_4_07=df[["Anciennes utilisatrices COC","Anciennes utilisatrices COP","Anciennes utilisatrices Depoprovéra","Anciennes utilisatrices Sayanapress",
                            "Anciennes utilisatrices implanon","Anciennes utilisatrices jadelle","Anciennes utilisatrices Levoplant","Anciennes utilisatrices Levonorgestrel",
                            "Anciennes utilisatrices Sterilisation feminine","Anciens utilisateurs Sterilisation masculine","Anciennes utilisatrices DIU","Anciens utilisateurs Condoms masculin",
                            "Anciennes utilisatrices Condoms féminin","Anciennes utilisatrices Collier","Anciennes utilisatrices MAMA","Anciennes utilisatrices Totales"]]

        PF_Tableau_4_08=df[["Nouvelles utilisatrices PF Moins de 19 ans","Nouvelles utilisatrices PF 20-24 ans","Nouvelles utilisatrices PF 25 ans et plus",
                            "Anciennes utilisatrices PF Moins de 19 ans","Anciennes utilisatrices PF 20-24 ans","Anciennes utilisatrices PF 25 ans et plus",
                            "Utilisatrices PF totales"]]

        PF_Tableau_4_09A=df[["Quanitité utilisée de COC","Quanitité utilisée de COP","Quanitité utilisée de DMPA IM","Quanitité utilisée de DMPA sous-cutanés",
                            "Quanitité utilisée d'implant pour 5 ans (Jadelle)","Quanitité utilisée d'implant pour 3 ans (Implanon)",
                            "Quanitité utilisée de Levonorgestrel","Quanitité utilisée de Levoplant"]]

        PF_Tableau_4_09B=df[["Quanitité utilisée de DIU","Quanitité utilisée de Préservatif masculin","Quanitité utilisée de Préservatif féminin",
                            "Stérilisation féminine (Ligature)","Stérilisation masculine (Vasectomie)","Quanitité utilisée de Collier du Cycle","Méthode MAMA"]]

        SMI_Tableau_4_11=df[["Femmes vues en CPN1","Dont vues au premier trimestre de grossesse","Femmes vues en CPN4","Nombre de femmes ayant bénéficié d'un TPI3"]]

        SMI_Tableau_4_12=df[["femmes vues en CPN","Femmes vues en CPN ayant reçu un counseling nutritionnel",
                            "Proportion de femmes vues en CPN ayant reçu un counseling nutritionnel (%)",
                            "Enfants (nouveaux +anciens) vus en consultation nourrissons sains 0 à 23 mois",
                            "Enfant de 0 à 23 mois dont la mère/accompagnant a reçu un counseling sur l’ANJE",
                            "Proportion de CNS de 0 à 23 mois dont la mère/accompagnant a reçu un counseling sur l’ANJE"]]

        SMI_Tableau_4_13=df[["Grossesses référées"]]

        SMI_Tableau_4_14=df[["Accouchement normal","Accouchement assisté à l'aide d'instrument et/ou de produits",
                            "Accouchement par césarienne",
                            "Total accouchement"]]
        SMI_Tableau_4_15=df[["Accouchement avec partogramme","Proportion (%) des accouchements réalisés avec partogramme",
                            "Mort né Frais","Mort né Macéré","Proportion (%) de mort-nés frais parmi les mort-nés"]]

        SMI_Tableau_4_16=df[["Total Naissance vivante","Moins de 2500 g","Enfants ayant bénéficié de la méthode Kangourou",
                        "Enfants mis au sein à la première heure","Proportion de faible poids de naissance",
                        "Proportion (%) d'enfants mis au sein à la première heure"]]
        ################################################
        SMI_Tableau_4_16=df[["Total Naissance vivante","Moins de 2500 g","Enfants ayant bénéficié de la méthode Kangourou",
                                        "Enfants mis au sein à la première heure"]]

        SMI_Tableau_4_17=df[["décès néo-natal d'enfant de 0 à 6 jours",
                            "décès néo-natal d'enfant de 7 à 28 jours","Total décès néo-natal"]]

        SMI_Tableau_4_18=df[["Femmes ayant accouché", "Femmes vues en consultation postnatale 7 ème jour",
                            "Femmes vues en consultation postnatale 42 ème jour"]]
        """
        SMI_Tableau_4_19=df[["Décès maternels pour cause de disproportion","Décès maternels pour cause de présentation vicieuse",
                            "Décès maternels pour cause d'éclampsie","Décès maternels pour cause de rétention placentaire",
                            "Décès maternels pour cause de rupture utérine","Décès maternels pour cause d'hémorragie",
                            "Décès maternels pour cause d'infection","Décès maternels pour cause de GEU",
                            "Décès maternels pour cause de complications d'avortement","Décès maternels pour d'autres causes",
                            "Total décès maternels toutes causes confondues"]]
        """
        SMI_Tableau_4_20=df[["Nombre de décès maternel audité"]] #col4

        SMI_Tableau_4_21= df[["Avortements spontanés reçus", "Avortements clandestins reçus", "Avortements thérapeutiques reçus",
                            "Soins après avortements avec AMIU", "Soins après avortements avec Médicaments",
                            "Autres soins après avortements"]]

        SMI_Tableau_4_22=df[["Cas de fistules obstetricales reçus","Fistules obstetricales operes",
                            "Cas de complications ou sequelles de l’excision reçus",
                            "Complications ou sequelles d'excision pris en charge"]]



        PTME_VIHTableau_4_23=df[["Nombre de femmes vues en CPN au cours du mois et ayant bénéficié d’un test VIH",
                            "Taux de dépistage PTME VIH", "Nombre de femmes enceintes dépistées VIH+",
                            "Taux de séropositivité PTME","Femmes vues en CPN testées VIH+ nouvellement mises sous TARV"]]

        PTME_VIHTableau_4_25=df[["Enfants nés vivant de mères séropositives",
                                "Enfants nés vivant qui reçoivent une prophylaxie par ARV",
                                "% Enfants nés vivant qui reçoivent une prophylaxie par ARV",
                                "Nourrissons nés de mères séropositives ayant subi un test virologique du VIH (PCR) à la naissance",
                                "Nourrissons ayant un test virologique du VIH (PCR) positif à la naissance",
                                "Taux de positivité à la PCR des nourrissons nés de mère séropositive à la naissance",
                                "Enfants nés de mères séropositives ayant subi un test sérologique du VIH (TDR)",
                                "Enfants ayant un test sérologique du VIH (TDR) positif",
                                "Taux de positivité au TDR des enfants nés de mère séropositive"]]

        PTME_Hep_BTableau_4_25=df[["Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs",
                                    "Femmes vues en CPN testées positives à l'hépatite B",
                                    "Femmes vues en CPN testées positives à l'hépatite B mises sous TARV",
                                    "Partenaires de femmes enceintes positives à l'hépatite B ayant bénéficié d’un dépistage de l'hépatite",
                                    "Partenaires de femmes enceintes positives à l'hépatite B testés positifs à l’hépatite B",
                                    "% Femmes enceintes vues en CPN ayant bénéficié d’un test AgHBs",
                                    "Taux de positivité des femmes vues en CPN testées à l'hépatite B",
                                    "Taux de mise sous TARV des femmes vues en CPN testées positives à l'hépatite B"]]

        PTME__SyphilisTableau_4_25=df[["Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis",
                                    "Femmes enceintes vues en CPN ayant été testées positives à la syphilis",
                                    "Femmes vues en CPN testées positives à la syphilis et mises sous traitement",
                                    "Enfants nés vivants de mères positives à l'hépatite B",
                                    "Enfants nés vivants de mères positives à l'hépatite B ayant bénéficié du vaccin Hep B dans les 24 heures de vie",
                                    "% Femmes enceintes vues en CPN ayant bénéficié d’un test de dépistage de la syphilis",
                                    "Taux de positivité des femmes vues en CPN testées à la sypjilis",
                                    "Taux de mise sous traitement des femmes vues en CPN testées positives à la syphilis"]]

        NutritionTableau_4_26=df[["Dépistage de la Malnutris aigüs modérés",
                                    "Dépistage de la Malnutris aigüs sévères","Taux de dépistage des MAM",
                                    "Taux de dépistage des MAS","Total malnutris aigüs"]]

        NutritionTableau_4_27=df_ds[["MAM reçus des ASBC", "MAM confirmé par la FS", "MAS reçus des ASBC",
                                    "MAS confirmé par la FS", "Taux de confirmation des MAM",
                                    "Taux de confirmation des MAS"]]

        NutritionTableau_4_29= df[["MAM Total des Sorties","MAM Guéris","MAM Décédés","MAM Abandons",
                                    "MAS ambulatoire Total des Sorties","MAS ambulatoire Guéris","MAS ambulatoire Décédés","MAS ambulatoire Abandons",
                                    "MAS interne Total des Sorties","MAS interne Guéris","MAS interne Décédés","MAS interne Abandons"]]
        #JVAP1Tableau_4_31

        ConsultationTableau_4_33=df[["Nouveaux consultants Moins de 5 ans","Nouveaux consultants 5-14 ans",
                                    "Nouveaux consultants 15 ans et plus","Nouveaux consultants Ensemble"]]
        PaludismeTableau_4_53=df[["Cas suspect de Paludisme","Gouttes épaisses Réalisées","Gouttes épaisses Positives",
                                    "TDR Réalisées","TDR Positives"]]
        PaludismeTableau_4_54=df[["Paludisme simple (confirmés+présumés) Total",
                                "Paludisme simple (confirmés+présumés) Moins de 5 an)","Paludisme simple (confirmés+présumés) femmes enceinte)",
                                "Paludisme grave (confirmés+présumés) Total","Paludisme grave (confirmés+présumés) Moins de 5 ans",
                                "Paludisme grave (confirmés+présumés) femmes enceintes"]]

        PaludismeTableau_4_55=df[["Paludisme simple confirmés Total","Paludisme simple confirmés Moins de 5 ans",
                                    "Paludisme simple confirmés femmes enceintes","Paludisme grave confirmés Total",
                                    "Paludisme grave confirmés Moins de 5 ans","Paludisme grave confirmés femmes enceintes"]]

        PaludismeTableau_4_57=df[["Paludisme simple traité avec ACT","Paludisme simple confirmé traité avec ACT"]]

        PaludismeTableau_4_58=df[["Décès paludisme global","Décès paludisme chez les moins de 5 ans",
                                    "Décès paludisme chez les femmes enceintes"]]

        PaludismeTableau_4_59=df[["MILDA femmes enceintes","MILDA chez les 0 à 11 mois"]]

        VIHTableau_4_65=df[["Total IST (Notification syndromique)"]]

        VIHTableau_4_66=df[["Dépistage VIH Personnes testées","Dépistage VIH Personnes testées positives"]]

        VIHTableau_4_67=df[["Dépistage des hépatites Personnes testées pour HVB","Dépistage des hépatites Personnes testées pour HVC",
                                "Dépistage des hépatites Personnes testées positives pour HVB",
                                "Dépistage des hépatites Personnes testées positives pour HVC"]]

        File_active_VIHTableau_4_68=df[["PVVIH sous TAR ayant bénéficié du dosage de la CV à 12 mois de traitement",
                                            "PVVIH sous TAR dont la CV à 12 mois est indétectable"]]

        VIHTableau_4_69=df[["Nombre de cas d'AELB notifiés","Nombre de cas d'AELB ayant nécessité une prophylaxie par ARV",
                                "Nombre de cas d'AELB positifs à l'issue de la phase de suivi"]]

        TBTableau_4_70=df[["Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) Nouveaux patients",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) Rechutes",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) déjà traités (hors rechutes) Echecs",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) déjà traités (hors rechutes) Reprises",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) déjà traités (hors rechutes) Autres",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) Antécédents de traitement inconnus",
                        "Dépistage TB ( Forme pulmonaire, confirmée bactériologiquement) Total",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Nouveaux patients",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Rechutes",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Reprises",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Autres",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Antécédents de traitement inconnus",
                        "Dépistage TB (Forme pulmonaire, diagnostiquée cliniquement) Total"]]

        TBTableau_4_71=df[["Dépistage TB (Forme extrapulmonaire) Nouveaux patients",
                        "Dépistage TB (Forme extrapulmonaire) Rechutes",
                        "Dépistage TB (Forme extrapulmonaire) Reprises",
                        "Dépistage TB (Forme extrapulmonaire) Autres",
                        "Dépistage TB (Forme extrapulmonaire) Antécédents de traitement inconnus",
                        "Dépistage TB (Forme extrapulmonaire) Total",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Nouveaux patients",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Rechutes",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Déjà traités (hors rechutes) Echecs",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Déjà traités (hors rechutes) Reprises",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Déjà traités (hors rechutes) Autres",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Antécédents de traitement inconnus",
                        "Dépistage TB (Synthèse cas de TB toutes formes) Total"]]

        TBTableau_4_76=df[["Malades TB (tous les cas, toutes formes) analysés",
                            "Résultas de traitement TB (tous les cas, toutes formes): Guéri",
                            "Résultas de traitement TB (tous les cas, toutes formes):Traitement terminé",
                            "Résultas de traitement TB (tous les cas, toutes formes): Echec",
                            "Résultas de traitement TB (tous les cas, toutes formes): Décédé",
                            "Résultas de traitement TB (tous les cas, toutes formes): Perdu de vue",
                            "Résultas de traitement TB (tous les cas, toutes formes): Non évalué"]]

        LèpreTableau_4_77=df[["Nouveaux cas de lèpre MB masculin",
                                    "Nouveaux cas de lèpre MB feminin",
                                    "Nouveaux cas de lèpre PB masculin",
                                    "Nouveaux cas de lèpre PB feminin",
                                    "Nouveaux cas de lèpre Total masculin",
                                    "Nouveaux cas de lèpre Total feminin",
                                    "Enfants nouveaux cas de lèpre MB masculin",
                                    "Enfants nouveaux cas de lèpre MB feminin",
                                    "Enfants nouveaux cas de lèpre PB masculin",
                                    "Enfants nouveaux cas de lèpre PB feminin",
                                    "Enfants nouveaux cas de lèpre Total masculin",
                                    "Enfants nouveaux cas de lèpre Total feminin",
                                    "Lèpre-infirmité degré 2 MB masculin",
                                    "Lèpre-infirmité degré 2 MB feminin",
                                    "Lèpre-infirmité degré 2 PB masculin",
                                    "Lèpre-infirmité degré 2 PB feminin",
                                    "Lèpre-infirmité degré 2 Total masculin",
                                    "Lèpre-infirmité degré 2 Total feminin"]]

        OdontoTableau_4_95=df[["Detartrage","Curetage ou attouchement de poche parodontale",
                                "Obturation dent définitive","Dents traitees avec composites",
                                "Dent traitee avec ciment","Dents traitées avec amalgames",
                                "Confection de prothese dentaire","Ablation frein de langue",
                                "Reduction des ATM","Reduction des luxations alveolaires",
                                "Extraction de corps etranger","Autres actes d’odontostomatologie"]]

        ORLTableau_4_96=df[["Extraction de corps etranger du CAE","Extraction de corps etranger du nez",
                                "Lavage de oreilles ORL","Acoumetrie","Actes-Nasofibroscopie",
                                "Actes-Tympanométrie/reflexes stapédiens","Aerosoltherapie",
                                "Aspiration-mechage","Audiometrie","Cauterisation des cornets inferieurs",
                                "Extraction corps etranger oropharynge","Extraction de corps etranger des voies respiratoires inferieures",
                                "Autres actes en ORL"]]

        AnesthesieTableau_4_98=df[["Anesthesie generale","Anesthesie locale",
                                    "Anesthesie loco regionale","Total actes anesthésiques"]]

        Chirurgie_ophtalmoTableau_4_99=df[["Chirurgie-Cataracte","Interventions de Glaucome",
                                            "Interventions pour Evisceration","Chirurgie-Trichiasis trachomateux",
                                            "Intervention-Autres trichiasis","Intervention pour Cure de Pterygion",
                                            "Intervention pour Cure de Chalazion","Interventions pour autres Ophtalmologie"]]

        Chirurgie_ORLTableau_4_99=df[["Trachéotomie","Cheloidectomie cou et face",
                                        "Polypeptectomie des fosses nasales","Chirurgie-Adenoidectomie",
                                        "Chirurgie-Adeno-amygdalectomie","Chirurgie-Amygdalectomie",
                                        "Chirurgie-Thyroidectomie","Chirurgie-Autres interventions en ORL"]]

        Imagerie_RadioTableau_4_105=df[["Nombre d'Examens de radio des os","Nombre d'Examens de radios pulmonaires",
                                        "Nombre d'Examens d'Autres Radios","Echographies abdominales","Imagérie-Pelvienne",
                                        "Echographies obstetricales","Imagérie-Echo Fibroscan Ano rectoscopie",
                                        "Imagérie-Echographie cardiaque","Imagérie-ECG","Imagérie-Echodoppler transcranien",
                                        "Echographies Autres"]]

        Imagerie_Scan_irm_endosTableau_4_105=df[["Imagérie-Scanner cérébral","Imagérie-Scanner abdomino -pelvien",
                                                    "Imagérie-Scanner pulmonaire","Imagérie-Autres scanners",
                                                    "Imagérie-IRM Cérébrale et TSA (Tronc Supra Aortique)",
                                                    "Imagérie-IRM Abdominale","Imagérie-IRM Rachis",
                                                    "Imagérie-IRM Articulations","Imagérie-IRM Membres",
                                                    "Imagérie-IRM bassins et pelvis","Imagérie-Autres IRM",
                                                    "Nombre d'Endoscopies de la voie digestive haute",
                                                    "Nombre d'Endoscopies de la voie digestive basse",
                                                    "Nombre d'Endoscopies (Bronchoscopie)"]]

        LaboratoireTableau_4_106=df[["bactériologie-examen réalisé","bactériologie-résultat anormal ou positif",
                                        "Hématologie Immunologie-examen réalisé","Hématologie Immunologie-résultat anormal ou positif",
                                        "Anatomo-pathologie-examen réalisé","Anatomo-pathologie-résultat anormal ou positif",
                                        "Biochimie-examen réalisé","Biochimie-résultat anormal ou positif",
                                        "Parasitologie et mycologie-examen réalisé","Parasitologie et mycologie-résultat anormal ou positif"]]

        Sante_comTableau_5_03=df_ds[["Séances d'IEC réalisés-CPN","Séances d'IEC réalisés-Post natale",
                                "Séances d'IEC réalisés-accouchement dans les FS","Séances d'IEC réalisés-Lavage des mains",
                                "Séances d'IEC réalisés-Consultation infantile","Séances d'IEC réalisés-Planning familial",
                                "Séances d'IEC réalisés-Surveillance basée sur les évènements","Séances d'IEC réalisés-Maladies cibles du PEV",
                                "Séances d'IEC réalisés-hygiène et l'assainissement","Séances d'IEC réalisés-IST et VIH/SIDA",
                                "Séances d'IEC réalisés-Diarrhée","Séances d'IEC réalisés-Paludisme",
                                "Séances d'IEC réalisés-Nutrition (autres que ANJE)"]]

        Sante_comTableau_5_04=df_ds[["Séances d'IEC réalisés-Alimentation du nourrisson et du jeune enfant",
                                        "Séances d'IEC réalisés-Tuberculose","Séances d'IEC réalisés-IRA",
                                        "Séances d'IEC réalisés-Enregistrement des naissances",
                                        "Séances d'IEC réalisés-Prévention de la transmission mère enfant du VIH",
                                        "Séances d'IEC réalisés-Facteurs de risque des maladies non transmissibles",
                                        "Séances d'IEC réalisés-Tabagisme","Séances d'IEC réalisés-Alcoolisme",
                                        "Séances d'IEC réalisés-Covid-19","Séances d'IEC réalisés- Don de sang",
                                        "Séances d'IEC réalisés-Autres thèmes"]]
        Sante_comTableau_5_06=df_ds[["Personnes touchées lors des séances d'IEC réalisées-CPN",
                                    "Personnes touchées lors des séances d'IEC réalisées-Post natale",
                                    "Personnes touchées lors des séances d'IEC réalisées-accouchement dans les FS",
                                    "Personnes touchées lors des séances d'IEC réalisées-Lavage des mains",
                                    "Personnes touchées lors des séances d'IEC réalisées-Consultation infantile",
                                    "Personnes touchées lors des séances d'IEC réalisées-Planning familial",
                                    "Personnes touchées lors des séances d'IEC réalisées-Surveillance basée sur les évènements",
                                    "Personnes touchées lors des séances d'IEC réalisées-Maladies cibles du PEV",
                                    "Personnes touchées lors des séances d'IEC réalisées-hygiène et l'assainissement",
                                    "Personnes touchées lors des séances d'IEC réalisées-IST et VIH/SIDA",
                                    "Personnes touchées lors des séances d'IEC réalisées-Diarrhée",
                                    "Personnes touchées lors des séances d'IEC réalisées-Paludisme"]]

        Sante_comTableau_5_07=df_ds[["Personnes touchées lors des séances d'IEC réalisées-Nutrition (autres que ANJE)",
                                        "Personnes touchées lors des séances d'IEC réalisées-Alimentation du nourrisson et du jeune enfant",
                                        "Personnes touchées lors des séances d'IEC réalisées-Tuberculose",
                                        "Personnes touchées lors des séances d'IEC réalisées-IRA",
                                        "Personnes touchées lors des séances d'IEC réalisées-Enregistrement des naissances",
                                        "Personnes touchées lors des séances d'IEC réalisées-Prévention de la transmission mère enfant du VIH",
                                        "Personnes touchées lors des séances d'IEC réalisées-Facteurs de risque des maladies non transmissibles",
                                        "Personnes touchées lors des séances d'IEC réalisées-Tabagisme",
                                        "Personnes touchées lors des séances d'IEC réalisées-Alcoolisme",
                                        "Personnes touchées lors des séances d'IEC réalisées-Covid-19",
                                        "Personnes touchées lors des séances d'IEC réalisées- Don de sang",
                                        "Personnes touchées lors des séances d'IEC réalisées-Autres thèmes"]]

        Sante_comTableau_5_10=df_ds[["Nouvelles utilisatrices PF_com-COC",
                                        "Nouvelles utilisatrices PF_com- COP","Nouvelles utilisatrices PF_com- (Sayana press)",
                                        "Nouvelles utilisatrices PF_com-Condoms masculin","Nouvelles utilisatrices PF_com-Condoms féminin",
                                        "Anciennes utilisatrices PF_com-COC","Anciennes utilisatrices PF_com- COP",
                                        "Anciennes utilisatrices PF_com- (Sayana press)","Anciennes utilisatrices PF_com-Condoms masculin",
                                        "Anciennes utilisatrices PF_com-Condoms féminin","Réapprovisionnement PF_com-COC",
                                        "Réapprovisionnement PF_com- COP","Réapprovisionnement PF_com- (Sayana press)",
                                        "Réapprovisionnement PF_com-Condoms masculin","Réapprovisionnement PF_com-Condoms féminin",
                                        "Clientes PF référées par les ASBC"]]

        Sante_comTableau_5_11=df_ds[["Femmes enceintes ayant reçu du fer + acide folique_com",
                                        "Femmes ayant accouché à domicile_com","Perdues de vue recherchées pour CPN_com",
                                        "Perdues de vue retrouvées pour CPN _com",
                                        "% de Perdues de vue recherchées et retrouvées pour CPN _com",
                                        "Femmes en post partum ayant reçu du fer +acide folique_com",
                                        "Femmes référées pour fistules obstétricales_com"]]

        Sante_comTableau_5_12=df_ds[["Enfants de 0-11 moisPerdus de vue_com",
                                        "Enfants de 0-11 moisPerdus de vue retrouvés_com",
                                        "Enfants de 15-18 moisPerdus de vue_com",
                                        "Enfants de 15-18 moisPerdus de vue retrouvés_com"]]

        Sante_comTableau_5_13=df_ds[["GASPA nouvellement mis en place-Femmes Enceintes",
                                        "GASPA nouvellement mis en place-mères d’enfants de moins de 6 mois",
                                        "GASPA nouvellement mis en place-mères d’enfants de 6-23 mois",
                                        "GASPA Femmes enceintes-Femmes inscrites",
                                        "GASPA Femmes enceintes-Femmes présentes",
                                        "GASPA mères d'enfants de moins de 6 mois-Femmes inscrites",
                                        "GASPA mères d'enfants de moins de 6 mois-Femmes présentes",
                                        "GASPA mères d’enfants de 6-23 mois -Femmes inscrites",
                                        "GASPA mères d’enfants de 6-23 mois -Femmes présentes",
                                        "Taux de présences GASPA-Femmes Enceintes",
                                        "Taux de présences GASPA-mères d’enfants de moins de 6 mois",
                                        "Taux de présences GASPA-mères d’enfants de 6-23 mois"]]

        Sante_comTableau_5_14=df_ds[["MAM Nouveaux cas _com","MAM Anciens cas_com",
                                    "MAS Nouveaux cas _com","MAS Anciens cas_com"]]

        Sante_comTableau_5_15=df_ds[["Cas suspects de paludisme moins de 5 ans_com",
                                        "Cas suspects de paludisme 5 ans et plus _com","Total Cas suspects de paludisme_com",
                                        "TDR Réalisé moins de 5 ans_com","TDR Réalisé 5 ans et plus _com",
                                        "Total TDR Réalisé_com","TDR positif moins de 5 ans_com",
                                        "TDR positif 5 ans et plus _com","Total TDR positif_com"]]

        Sante_comTableau_5_16=df_ds[["Cas de paludisme simple confirmé traités aux ACT moins de 5 ans_com",
                                        "Cas de paludisme simple confirmé traités aux ACT 5 ans et plus _com",
                                        "Total Cas de paludisme simple confirmé traités aux ACT_com",
                                        "Cas de paludisme grave confirmé ayant reçu de l’artésunate suppositoire moins de 5 ans_com",
                                        "Cas de paludisme grave confirmé ayant reçu de l’artésunate suppositoire 5 ans et plus _com",
                                        "Total Cas de paludisme grave confirmé ayant reçu de l’artésunate suppositoire_com",
                                        "Cas de paludisme référés moins de 5 ans_com","Cas de paludisme référés 5 ans et plus _com",
                                        "Total Cas de paludisme référés_com"]]

        Sante_comTableau_5_17=df_ds[["Total cas de diarrhée reçus_com","Cas de diarrhée reçus < 5 ans_com",
                                        "Total cas de diarrhee traites avec le SRO+Zinc-Com_com",
                                        "Cas de diarrhee traites avec le SRO+Zinc-Com < 5 ans_com",
                                        "Total cas de diarrhee referes_com","Cas de diarrhee referes < 5 ans_com"]]

        Sante_comTableau_5_18=df_ds[["Total cas de toux reçus_com","Cas de toux reçus < 5 ans_com",
                                    "Cas de toux simples traites avec amoxicilline dispersible-Com < 5 ans_com",
                                    "Total cas de toux/difficultes respiratoires referes_com",
                                    "Cas de toux/difficultes respiratoires referes-Com < 5 ans_com"]]

        Sante_comTableau_5_19=df_ds[["Femmes vues en CPN ayant reçu TPI2-Com",
                                        "Femmes vues en CPN ayant reçu TPI3-Com"]]

        Sante_comTableau_5_20=df_ds[["Deces Enfants de 0 à 28 jours_Com",
                                    "Deces de Femmes suite a un accouchement ou dans les 42 jrs après accouchement_Com"]]

        df_consultant_prive=df_prive[["Nouveaux consultants Ensemble"]]
        df_rupture_dmeg=df_specifique_ds[["Nombre de DMEG","Nombre de DMEG n'ayant pas connu de rupture de médicaments traceurs"]]
        df_norme_personnel=df_specifique_ds[["Effectif CSPS publiques remplissant la norme minimale en personnel"]]
        df_lepre_fin_annee=df_specifique[["Lèpre-malades en fin d'année MB masculin","Lèpre-malades en fin d'année MB feminin",
                                            "Lèpre-malades en fin d'année PB masculin","Lèpre-malades en fin d'année PB feminin",
                                            "Lèpre-malades en fin d'année Total masculin","Lèpre-malades en fin d'année Total feminin"]]
        df_vih_sous_arv=df_specifique["PvVIH sous ARV"]
        ## creation des tableaux annuels du document
        df_vap=df_vap_deces_mat[["Nombre d'enfants de 0-11 mois ayant reçu le VAP1",
            "Nombre d'enfants de 0-11 mois ayant reçu le VAP2",
            "Nombre d'enfants de 0-11 mois ayant reçu le VAP3",
            "Nombre d'enfants de 12-23 mois ayant reçu le VAP4"]]
        df_deces_maternel=df_vap_deces_mat[["Décès maternels pour cause de disproportion", "Décès maternels pour cause de présentation vicieuse",
                                        "Décès maternels pour cause d'éclampsie","Décès maternels pour cause de rétention placentaire",
                                                                            "Décès maternels pour cause de rupture utérine", "Décès maternels pour cause d'hémorragie",
                                                                        "Décès maternels pour cause d'infection", "Décès maternels pour cause de GEU",
                                                                            "Décès maternels pour cause de complications d'avortement", "Décès maternels pour d'autres causes",
                                                                            "Total décès maternels toutes causes confondues"]]

        df_jva_p1=df_jva_p1_merge.drop("UO_annuaire", axis=1)
        df_jva_p1=df_jva_p1[["Enfants 6 – 11 mois ayant reçu la vitamine A","Enfants 12 – 59 mois ayant reçu la vitamine A",
                "Enfants 12 – 59 mois ayant été déparasités","Enfants Total ayant reçu la vitamine A"]]

        df_jva_p2=df_jva_p2_merge.drop("UO_annuaire", axis=1)
        df_jva_p2=df_jva_p2[["Enfants 6 – 11 mois ayant reçu la vitamine A","Enfants 12 – 59 mois ayant reçu la vitamine A",
                "Enfants 12 – 59 mois ayant été déparasités","Enfants Total ayant reçu la vitamine A"]]


        df_chirurgie_ch1=df_chirurgie_ch[['CHR Banfora','CHR Dédougou','CHU Bogodogo', 'CHU Pédiatrique CDG','CHU Tengandogo','CHU Yalgado','Hôpital Paul VI',
            'Hôpital Saint Camille de Ouagadougou', 'Hôpital SCHIPHRA','CHR Tenkodogo', 'CHR Kaya']]


        df_chirurgie_ch2=df_chirurgie_ch[[ 'CHR Koudougou',"CHR Manga","CHR Fada N'Gourma",'CHU Sanou Souro','CHUR Ouahigouya',
                                        'CHR Ziniaré','CHR Dori', 'CHR Gaoua','total_chirurgie_ch']]

        df_soins_urgence1=df_soins_urgence[['Patients reçus en urgence','Patients reçus en urgence et mis en observation (MEO)',
                                            'patients mis en observation (%)','Nombre de journees de mise en observation']]

        df_soins_urgence2=df_soins_urgence[['Patients référés/reçus en urgence',"Dont patients reçus en urgence/évacués d'une formation sanitaire privée",
                                        'Nombre total de patients reçus en urgence: transféré', 'Autres (circuit non spécifié)']]

        df_soins_urgence3=df_soins_urgence[["Normale (exéat)",'Contre avis médical', 'Sans avis médical', 'Décédé en urgence','Sortie Transféré dans un autre service',
                                            'Sortie Evacué vers un autre hôpital']]

        # ------------Donnees de Basciloscopie et coinfection TB---------------------------------------
        # Donnees de Basciloscopie et coinfection TB
        df_tb_region_Basciloscopie=df_tb_region_merge[["Malades soumis au dépistage de TB","Malades dépistés positifs TB","Malades orientés par les acteurs communautaires soumis au dépistage de TB",
        "Malades orientés par les acteurs communautaires dépistés positifs TB","TB Lames lues pour le dépistage: Négatif","TB Lames lues pour le dépistage: Positif 1-9 B",
        "TB Lames lues pour le dépistage: Positif 1 à 3+","TB Lames lues pour le contrôle M2-M3: Négatif","TB Lames lues pour le contrôle M2-M3: Positif 1-9 B",
        "TB Lames lues pour le contrôle M2-M3: Positif 1 à 3+","TB Lames lues pour le contrôle M5-M6: Négatif","TB Lames lues pour le contrôle M5-M6: Positif 1-9 B",
        "TB Lames lues pour le contrôle M5-M6: Positif 1 à 3+","Total lames lues TB"]]

        # Donnees de coinfection TB
        df_tb_region_CoInfectionTB_VIH_1=df_tb_region_merge[["TB Total notifiés (Nouveaux cas et Rechutes)","Nbre de cas TB (tous les cas, toutes formes) testés au VIH"]]
        df_tb_region_CoInfectionTB_VIH_2=df_tb_region_merge[["Nombre de cas TB (tous les cas, toutes formes) testé VIH+"]]
        df_tb_region_CoInfectionTB_VIH_3=df_tb_region_merge[["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri"]]
        df_tb_region_CoInfectionTB_VIH_4=df_tb_region_merge[["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV Cohorte"]]
        df_tb_region_CoInfectionTB_VIH_cohorte1=df_tb_region_merge[["Nbre de cas TB (tous les cas, toutes formes) testés au VIH Cohorte",
                                        "Nombre de cas TB (tous les cas, toutes formes) testé VIH+ Cohorte"]]
        df_tb_region_CoInfectionTB_VIH_cohorte2=df_tb_region_merge[["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous Cotri"]]
        df_tb_region_CoInfectionTB_VIH_cohorte3=df_tb_region_merge[["Nbre de cas TB (tous les cas, toutes formes) VIH+ sous ARV Cohorte"]]


        #--------------------Tableau integration des services------------------------------------------------------
        df_integration_services=df_integration_whopen.copy()
        df_integration_services=df_integration_services[["Nombre de CPN au cours desquelles un counseling PFPP et des conseils nutritionnels ont été offerts",
                                                        "Nombre de femmes ayant bénéficié PFPP dans les 48 heures avec GATPA dont le nouveau-né a été mis au sein dans l'heure qui suit la naissance",
                                                        "Nombre de femmes ayant accouche au moins un enfant vivant",
                                                        "Nombre de femmes reçues en CPoN qui ont bénéficié de services PF et de conseils sur l’AE",
                                                        "Nombre de femmes vues en CPoN",
                                                        "Nombre de nourrissons de 0-6 mois reçus en CNRS qui sont à jour de leurs vaccins dont la mère a bénéficié de services PF et de conseils sur l’AE",
                                                        "Nombre de nourrissons de 0_6 mois recus en CNS","Nombre d'tilisatrices de PFPP dans les 48 heures",
                                                        "Nombre d'tilisatrices de PFPP en CPoN",
                                                        "Nombre d'utilisatrices de PFPP en CNS"]]
        #--------------------Tableau indicateurs WHOPEN------------------------------------------------------
        df_whopen=df_integration_whopen.copy()
        df_whopen=df_whopen[["Nombre de personnes asthmatiques sous traitement depuis plus de 6 mois dont l’asthme est maitrisé 15 ans et +",
                            "Nombre de personnes asthmatiques sous traitement depuis plus de 6 mois",
                            "Nombre de consultants de 15 ans et plus ayant bénéficiés de la prise de la TA en consultation",
                            "Nombre de consultants auto orientés 15 ans et plus",
                            "Nombre de patients ayant bénéficiés du dépistage du diabète en consultation",
                            "Nombre de nouveaux consultants de 40 ans et plus chez qui le risque de survenu d’un évènement cardiovasculaire a été évalué",
                            "Nombre de nouveaux consultants de 40 ans et plus"]]
        #--------------------Tableau indicateurs MPR------------------------------------------------------
        df_mpr=df_indicateurs_additionnels.copy()
        df_mpr=traitement_donnees_ds_ch(df_mpr,code_annuaire_ch,code_annuaire_unite_ch,correspondanceUID)
        df_mpr=df_mpr[["Nombre total de jours d'attente des patients avant la 1ère séance",
                    "Nombre de patients ayant bénéficier de la première séance pour la réadaptation",
                    "Nombre de séances de réadaptation réalisée",
                    "Nombre de patients ayant bénéficié de produits d'assistance",
                    "Nombre de patients qui sont dans le besoin de produits d'assistance",
                    "Nombre de sortie récupérés par groupe d'age",
                    "Nombre de sortie par groupe d'age"]]

        df_prive_accouchement=df_prive[['Total accouchement']]



        #------------------------------------------------------------------------------------
        #--CONSTRUCTION DES TABLEAUX DE L'ANNUAIRE STATISTIQUE-----------------------------

        # Copie du canevas temporaire
        temp_path = "Canevas_temp.xlsx"
        with open(canevas_path, "rb") as src, open(temp_path, "wb") as dst:
            dst.write(src.read())

        # Barre de progression
        progress = st.progress(0)
        steps = 131  # nombre de feuilles à écrire
        current = 0

        # Écriture dans le canevas
        with pd.ExcelWriter(temp_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            populationTableau1_05.to_excel(writer, sheet_name="Population", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            populationTableau1_05.to_excel(writer, sheet_name="Population", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            completude_pub.to_excel(writer, sheet_name="Complétude", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            completude_privee.to_excel(writer, sheet_name="Complétude", index=False,header=False, startrow=3, startcol=5)
            current += 1; progress.progress(current / steps)
            promptitude.to_excel(writer, sheet_name="Complétude", index=False,header=False, startrow=3, startcol=11)
            current += 1; progress.progress(current / steps)
            df_fs_publique.to_excel(writer, sheet_name="Infrast_publique", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_fs_privee1.to_excel(writer, sheet_name="Infrastructure_privées1", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            df_fs_privee2.to_excel(writer, sheet_name="Infrastructure_privées2", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            population_cibleTableau1_06.to_excel(writer, sheet_name="Pop_cible", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Vaccination_Tableau_4_01.to_excel(writer, sheet_name="Enfants_FE vaccinés1", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Vaccination_Tableau_4_02.to_excel(writer, sheet_name="Enfants vaccinés2", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Vaccination_Tableau_4_03.to_excel(writer, sheet_name="filles_femmes vaccinés", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PF_Tableau_4_06.to_excel(writer, sheet_name="PF NouvelleUtilisatrice", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PF_Tableau_4_07.to_excel(writer, sheet_name="PF AncienneUtilisatrice", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PF_Tableau_4_08.to_excel(writer, sheet_name="UtilisatricePF ", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            PF_Tableau_4_09A.to_excel(writer, sheet_name="Produits contraceptifs1", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PF_Tableau_4_09B.to_excel(writer, sheet_name="Produits contraceptifs2", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_11.to_excel(writer, sheet_name="Couverture CPN", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_12.to_excel(writer, sheet_name="Counseling", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_13.to_excel(writer, sheet_name="Grossesses référées", index=False,header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_14.to_excel(writer, sheet_name="Accouchement", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_prive_accouchement.to_excel(writer, sheet_name="Accouchement", index=False,header=False, startrow=2, startcol=6)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_15.to_excel(writer, sheet_name="Ac_Partogramme_Mortnés", index=False,header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_16.to_excel(writer, sheet_name="Naissances", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_17.to_excel(writer, sheet_name="Deces_neonatal", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_18.to_excel(writer, sheet_name="Cons_post_nat", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_20.to_excel(writer, sheet_name="Taux de décès maternel", index=False,header=False, startrow=2, startcol=4)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_21.to_excel(writer, sheet_name="PEC Avortement", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            SMI_Tableau_4_22.to_excel(writer, sheet_name="PEC excision_Fistules", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PTME_VIHTableau_4_23.to_excel(writer, sheet_name="PTME_VIH", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PTME_VIHTableau_4_25.to_excel(writer, sheet_name="PTME_VIH-Enfant", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PTME_Hep_BTableau_4_25.to_excel(writer, sheet_name="PTME_Hep B", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PTME__SyphilisTableau_4_25.to_excel(writer, sheet_name="PTME_Hep B_ Syphilis", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            NutritionTableau_4_26.to_excel(writer, sheet_name="Dépistage malnut", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            NutritionTableau_4_27.to_excel(writer, sheet_name="Confirmation MA ASBC", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            NutritionTableau_4_29.to_excel(writer, sheet_name="Nutrition 1", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            ConsultationTableau_4_33.to_excel(writer, sheet_name="Consultant_age", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_53.to_excel(writer, sheet_name="Confirmation palu", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_54.to_excel(writer, sheet_name="PaludismeGlobale", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_55.to_excel(writer, sheet_name="Paludisme Confirme", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_57.to_excel(writer, sheet_name="Traitement Paludisme", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_58.to_excel(writer, sheet_name="Palu_létalité", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            PaludismeTableau_4_59.to_excel(writer, sheet_name="Milda_Routine", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            VIHTableau_4_65.to_excel(writer, sheet_name="IST", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            VIHTableau_4_66.to_excel(writer, sheet_name="Conseil_dépist_VIH", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            VIHTableau_4_67.to_excel(writer, sheet_name="dépist Hépatite virale", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            File_active_VIHTableau_4_68.to_excel(writer, sheet_name="File active", index=False,header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            VIHTableau_4_69.to_excel(writer, sheet_name="AELB", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            TBTableau_4_70.to_excel(writer, sheet_name="Depistage TB", index=False,header=False, startrow=4, startcol=2)
            current += 1; progress.progress(current / steps)
            TBTableau_4_71.to_excel(writer, sheet_name="Depistage TB_suite", index=False,header=False, startrow=4, startcol=2)
            current += 1; progress.progress(current / steps)
            TBTableau_4_76.to_excel(writer, sheet_name="Traitement TB", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            LèpreTableau_4_77.to_excel(writer, sheet_name="Lèpre1", index=False,header=False, startrow=4, startcol=2)
            current += 1; progress.progress(current / steps)
            OdontoTableau_4_95.to_excel(writer, sheet_name="Acte odonto", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            ORLTableau_4_96.to_excel(writer, sheet_name="Acte ORL", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            AnesthesieTableau_4_98.to_excel(writer, sheet_name="Anesthésie", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Chirurgie_ophtalmoTableau_4_99.to_excel(writer, sheet_name="Chirurgie_Ophtalmo", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Chirurgie_ORLTableau_4_99.to_excel(writer, sheet_name="Chirurgie_HD_orl", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Imagerie_RadioTableau_4_105.to_excel(writer, sheet_name="Activité imagerie", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Imagerie_Scan_irm_endosTableau_4_105.to_excel(writer, sheet_name="Activité imagerie2", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            LaboratoireTableau_4_106.to_excel(writer, sheet_name="Examen_Labo", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            df_completude_asbc.to_excel(writer, sheet_name="Complétude ASBC", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_03.to_excel(writer, sheet_name="Seance IEC_ASBC_1", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_04.to_excel(writer, sheet_name="Seance IEC_ASBC_2", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_06.to_excel(writer, sheet_name="Participants IEC ASBC_1", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_07.to_excel(writer, sheet_name="Participants IEC ASBC_2", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_10.to_excel(writer, sheet_name="PF ASBC", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_11.to_excel(writer, sheet_name="Accompagnement ASBC", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_12.to_excel(writer, sheet_name="PDV Vaccination", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_13.to_excel(writer, sheet_name="GASPA", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_14.to_excel(writer, sheet_name="Dépistage communautaire MA", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_15.to_excel(writer, sheet_name="Palu ASBC_1", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_16.to_excel(writer, sheet_name="Palu ASBC_2", index=False,header=False, startrow=3, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_17.to_excel(writer, sheet_name="DIARRHEE ASBC", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_18.to_excel(writer, sheet_name="Toux ASBC", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_19.to_excel(writer, sheet_name="TPI com", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            Sante_comTableau_5_20.to_excel(writer, sheet_name="Décès en communauté", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            # atelier
            df_drd_final.to_excel(writer, sheet_name="Rupture_DRD", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_rupture_dmeg.to_excel(writer, sheet_name="Rupture_DMEG", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_norme_personnel.to_excel(writer, sheet_name="Norme_personnel", index=False,header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            df_lepre_fin_annee.to_excel(writer, sheet_name="Lèpre2", index=False,header=False, startrow=4, startcol=3)
            current += 1; progress.progress(current / steps)
            df_vih_sous_arv.to_excel(writer, sheet_name="File active", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_vap.to_excel(writer, sheet_name=" vaccination VAP", index=False,header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            df_deces_maternel.to_excel(writer, sheet_name="Décès maternel par causes", index=False,header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_jva_p1.to_excel(writer, sheet_name="JVAP1", index=False,header=False, startrow=2, startcol=5)
            current += 1; progress.progress(current / steps)
            df_jva_p2.to_excel(writer, sheet_name="JVAP2", index=False,header=False, startrow=2, startcol=5)
            current += 1; progress.progress(current / steps)
            df_consultant_prive.to_excel(writer, sheet_name="Consultant_age", index=False, header=False, startrow=2, startcol=6)
            current += 1; progress.progress(current / steps)
            df_consultant_chr_chu.to_excel(writer, sheet_name="consultant_CH", index=False, header=False, startrow=3, startcol=1)
            current += 1; progress.progress(current / steps)
            df_pcime.to_excel(writer, sheet_name="PCIME_TETU", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_tetu_cma_ch.to_excel(writer, sheet_name="PCIME_TETU", index=False, header=False, startrow=2, startcol=5)
            current += 1; progress.progress(current / steps)
            df_ref_contreref_echelon1.to_excel(writer, sheet_name="Réf_Evac_cont ref_1er échelon", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_ref_contreref_cma_ch.to_excel(writer, sheet_name="Références 2et3echelon", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_evacuation_cma_ch.to_excel(writer, sheet_name="Evacuation CH_CMA", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_mise_en_observation.to_excel(writer, sheet_name="MEO_décès", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_entre_hospi.to_excel(writer, sheet_name="Mode_entrée_hospi", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_sortie_hospi.to_excel(writer, sheet_name="Sortie Hospi", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_chirurgie_ds.to_excel(writer, sheet_name="Chirurgie_HD", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_chirurgie_ch1.to_excel(writer, sheet_name="Chirurgie_CH1", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_chirurgie_ch2.to_excel(writer, sheet_name="Chirurgie_CH2", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_soins_urgence1.to_excel(writer, sheet_name="activ_soins_urgence", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_soins_urgence2.to_excel(writer, sheet_name="activ_soins_urgence", index=False, header=False, startrow=25, startcol=1)
            current += 1; progress.progress(current / steps)
            df_soins_urgence3.to_excel(writer, sheet_name="sortie urgence", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_tb_region_Basciloscopie.to_excel(writer, sheet_name="Basciloscopie", index=False, header=False, startrow=4, startcol=1)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_1.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_2.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=2, startcol=4)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_3.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=2, startcol=6)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_4.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=2, startcol=8)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_cohorte1.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=24, startcol=1)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_cohorte2.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=24, startcol=4)
            current += 1; progress.progress(current / steps)
            df_tb_region_CoInfectionTB_VIH_cohorte3.to_excel(writer, sheet_name="CoInfectionTB_VIH", index=False, header=False, startrow=24, startcol=6)
            current += 1; progress.progress(current / steps)
            df_lits_cma_ch_chu.to_excel(writer, sheet_name="Lits", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_hospitalisation_cm_ch.to_excel(writer, sheet_name="Occupation lit", index=False, header=False, startrow=3, startcol=3)
            current += 1; progress.progress(current / steps)
            df_integration_services.to_excel(writer, sheet_name="Integration_service", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_whopen.to_excel(writer, sheet_name="WHOPEN", index=False, header=False, startrow=2, startcol=2)
            current += 1; progress.progress(current / steps)
            df_mpr.to_excel(writer, sheet_name="MPR", index=False, header=False, startrow=2, startcol=3)
            current += 1; progress.progress(current / steps)
            df_nosologie_consul_tranche_age.to_excel(writer, sheet_name="C externe_Age", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_nosologie_consul_region1.to_excel(writer, sheet_name="C externe_Région 1", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            df_nosologie_consul_region2.to_excel(writer, sheet_name="C externe_Région_2", index=False, header=False, startrow=2, startcol=1)
            current += 1; progress.progress(current / steps)
            principauxmotif_consultation.to_excel(writer, sheet_name="Principaux motifs", index=False, header=False, startrow=2, startcol=0)
            current += 1; progress.progress(current / steps)
            data_noso_hospit_age.to_excel(writer, sheet_name="Hospi_CM_CMA_CH_Age", index=False, header=False, startrow=3, startcol=0)
            principauxmotif_hospitalisation.to_excel(writer, sheet_name="Principaux motifs", index=False, header=False, startrow=15, startcol=0)
            current += 1; progress.progress(current / steps)
            principauxmotif_hospi_deces.to_excel(writer, sheet_name="Principaux motifs", index=False, header=False, startrow=28, startcol=0)
            current += 1; progress.progress(current / steps)
            hospitregion_cas_deces1.to_excel(writer, sheet_name="Hospi_CM_CMA_CH_Région 1", index=False, header=False, startrow=3, startcol=0)
            current += 1; progress.progress(current / steps)
            hospitregion_cas_deces2.to_excel(writer, sheet_name="Hospi_CM_CMA_CH_Région 2", index=False, header=False, startrow=3, startcol=0)
            current += 1; progress.progress(current / steps)
            hospitregion_cas_deces3.to_excel(writer, sheet_name="Hospi_CM_CMA_CH_Région 3", index=False, header=False, startrow=3, startcol=0)
            current += 1; progress.progress(current / steps)
            df_personnes_agees_sexe.to_excel(writer, sheet_name="Personnes_ages", index=False, header=False, startrow=2, startcol=0)
            df_personnes_agees_region_cas_deces1.to_excel(writer, sheet_name="Personnes_ages_region", index=False, header=False, startrow=3, startcol=0)
            current += 1; progress.progress(current / steps)
            df_personnes_agees_region_cas_deces2.to_excel(writer, sheet_name="Personnes_ages_region", index=False, header=False, startrow=26, startcol=0)
            current += 1; progress.progress(min(current / steps, 1.0))
            df_personnes_agees_region_cas_deces3.to_excel(writer, sheet_name="Personnes_ages_region", index=False, header=False, startrow=49, startcol=0)
            current += 1; progress.progress(min(current / steps, 1.0))



        #remplissage des titres
        titres_tableaux = {
        "indicateursPNDS_ODD": (f"Tableau 0.01: indicateurs PNDS et ODD","A1"),
        "synthèse indicateurs": (f"Tableau 0.02: Synthèse de quelques indicateurs ","A1"),
        "IndicateursEconomiques": (f"Tableau 0.03: Quelques indicateurs économiques","A1"),
        "SARA_QoC_RDQA_CS_HHFA": (f"Tableau 1.01: quelques résultats de l'audit de la qualité des données de routine (RDQA*)","A1"),
        "SARA_QoC_RDQA_CS_HHFA": (f"Tableau 1.02: quelques résultats de l'enquête SARA/HHFA* ","A11"),
        "SARA_QoC_RDQA_CS_HHFA": (f"Tableau 1.03: quelques résultats de l'enquête qualité des prestations de soins du paludisme, de la tuberculose et de l’infection à vih (QoC*) ","A20"),
        "PMABF": (f"Tableau 1.04: Quelques indicateurs clés de planification familiale issus de l’enquête PMABF au Burkina Faso désagrégés par milieu de résidence ","A1"),
        "Population": (f"Tableau 1.05: répartition de la population du Burkina Faso par district et par groupe d'âge en {annee_annuaire}","A1"),
        "Pop_cible": (f"Tableau 1.06: Effectifs des populations cibles  par district en {annee_annuaire}","A1"),
        "Pop_distance ": (f"Tableau 1.07: Répartition de la population selon la distance à la formation sanitaire de référence en {annee_annuaire}","A1"),
        "Complétude": (f"Tableau 1.08: Complétude des rapports mensuels des formations sanitaires en {annee_annuaire}","A1"),
        "Sécurité précaire_Publique": (f"Tableau 1.09: Rapports non transmis pour raison d'insécurité par mois des FS publiques en {annee_annuaire}","A1"),
        "Sécurité précaire_Privee": (f"Tableau 1.10: Rapports non transmis pour raison d'insécurité par mois des FS privées en {annee_annuaire}","A1"),
        "ComplétudeCorrige": (f"Tableau 1.11: complétude des rapports mensuels des formations sanitaires  prenant en compte la non transmission de rapport pour fait d'insecurité en {annee_annuaire}","A1"),
        "Bilan PA _1": (f"Tableau 2.01: Bilan physique et financier par effet du PNDS {annee_annuaire}","A1"),
        "Bilan PA _1": (f"Tableau 2.02: Bilan physique et financier par programme budgétaire du PNDS {annee_annuaire}","A12"),
        "Bilan PA _2": (f"Tableau 2.03: Bilan physique et financier par structure {annee_annuaire}","A1"),
        "Rupture_DMEG": (f"Tableau 2.04: Disponibilité des médicament traceurs dans les formations sanitaires publiques en {annee_annuaire}","A1"),
        "Rupture_DRD": (f"Tableau 2.05: Disponibilité des MEG traceurs dans les DRD par régions en {annee_annuaire}","A1"),
        "TX rupture mensuel DRD traceurs": (f"Tableau 2.06: Taux (%) de rupture de stock des 50 médicaments traceurs dans les DRD en {annee_annuaire}","A1"),
        "Rupture_DMEG_Temoins": (f"Tableau 2.07: Proportion des DMEG témoins ayant connu une ruptures de stock des 25 médicaments traceurs par région en {annee_annuaire}","A1"),
        "Gratuité_des soins": (f"Tableau 2.08: Activité de la gratuité des soins en {annee_annuaire}","A1"),
        "Personnel Région_1 DRH": (f"Tableau 3.01: Effectif de quelques emplois du Ministère de la santé au 31 décembre {annee_annuaire}","A1"),
        "Personnel Région_2 DRH": (f"Tableau 3.02: Effectif de quelques emplois du Ministère de la santé au 31 décembre {annee_annuaire}(suite)","A1"),
        "Personnel Région 3 DRH": (f"Tableau 3.03: Effectif de quelques emplois du Ministère de la santé au 31 décembre {annee_annuaire} (suite)","A1"),
        "Effectif par niveau": (f"Tableau 3.04: Effectif du personnel pour quelques emplois dans les structures publiques en {annee_annuaire}","A1"),
        "Effectif par niveau": (f"Tableau 3.05: Evolution des effectifs du personnel pour quelques emplois  de 2018 à {annee_annuaire}","A22"),
        "Evolution Effectif personnel": (f"Tableau 3.06: Evolution des effectifs du personnel pour quelques emplois  de 2019 à {annee_annuaire}","A1"),
        "Evolution Effectif personnel": (f"Tableau 3.07: Ratio habitants par type de personnel de santé de 2019 à 2023 (np le secteur privé)","A10"),
        "Infrast_publique": (f"Tableau 3.08: Nombre de formations sanitaires publiques selon le type en {annee_annuaire}","A1"),
        "Norme_personnel": (f"Tableau 3.09: Pourcentage des CSPS remplissant la norme minimale en personnel en {annee_annuaire}","A1"),
        "Infrastructure_privées1": (f"Tableau 3.10: Nombre de formations sanitaires privées par district et selon le type en {annee_annuaire}","A1"),
        "Infrastructure_privées2": (f"Tableau 3.11: Nombre de formations sanitaires privées par district et selon le type en {annee_annuaire} (suite)","A1"),
        "OBC": (f"Tableau 3.12: situation des OBC par district en {annee_annuaire}","A1"),
        "Lits": (f"Tableau 3.13: Nombre de lits dans les hopitaux (CMA/HD, CHR, CHU) en {annee_annuaire}","A1"),
        "Occupation lit": (f"Tableau 3.14: Taux d'occupation des lits dans les hôpitaux (CMA/HD, CHR, CHU) en {annee_annuaire}","A1"),
        "RMAT": (f"Tableau 3.15: rayon moyen d'action théorique en {annee_annuaire}","A1"),
        "Ratio_hbt_Csps": (f"Tableau 3.16: Ratio habitant/CSPS* en {annee_annuaire}","A1"),
        "Enfants_FE vaccinés1": (f"Tableau 4.01: Enfants 0-23 mois vaccinés par antigène en {annee_annuaire}","A1"),
        "Enfants vaccinés2": (f"Tableau 4.02: Enfants 0-23 mois vaccinés par antigène en {annee_annuaire} (suite)","A1"),
        "Couverture vaccinale1": (f"Tableau 4.03: Couverture vaccinale (%) par antigène en {annee_annuaire}","A1"),
        "Couverture vaccinale2": (f"Tableau 4.04: couverture vaccinale (%) par antigène en {annee_annuaire} (suite)","A1"),
        "filles_femmes vaccinés": (f"Tableau 4.05: Filles et femmes enceintes vaccinées par antigène en {annee_annuaire}","A1"),
        " vaccination VAP": (f"Tableau 4.06: Vaccination contre le paludisme en {annee_annuaire}","A1"),
        "PF NouvelleUtilisatrice": (f"Tableau 4.07: Nouvelles utilisatrices de PF par méthode  contraceptive en {annee_annuaire}","A1"),
        "PF AncienneUtilisatrice": (f"Tableau 4.08: Anciennes  utilisatrices de PF par méthode  contraceptive par structure en {annee_annuaire}","A1"),
        "UtilisatricePF ": (f"Tableau 4.09: Utilisatrices de PF par tranche d'age en {annee_annuaire}","A1"),
        "Produits contraceptifs1": (f"Tableau 4.10: Quantités de produits utilisés en PF par méthode en {annee_annuaire}","A1"),
        "Produits contraceptifs2": (f"Tableau 4.11: Quantités de produits utilisés en PF par méthode en {annee_annuaire}","A1"),
        "Couple années par structure1": (f"Tableau 4.12: Couple année protection par méthode  en {annee_annuaire}","A1"),
        "Couple années par structure2": (f"Tableau 4.13: Couple année protection par méthode  en {annee_annuaire}","A1"),
        "Couverture CPN": (f"Tableau 4.14: Consultations prénatales et TPI en {annee_annuaire}","A1"),
        "Counseling": (f"Tableau 4.15: Counseling nutritionnel en {annee_annuaire}","A1"),
        "Grossesses référées": (f"Tableau 4.16: Grossesses référées en {annee_annuaire}","A1"),
        "Accouchement": (f"Tableau 4.17: Accouchements réalisés dans les formations sanitaires en {annee_annuaire}","A1"),
        "Ac_Partogramme_Mortnés": (f"Tableau 4.18: Accouchements réalisés avec partogramme et morts nés en {annee_annuaire}","A1"),
        "Naissances": (f"Tableau 4.19: Naissances vivantes en {annee_annuaire}","A1"),
        "Deces_neonatal": (f"Tableau 4.20: Décès néo-natal en {annee_annuaire}","A1"),
        "Integration_service": (f"Tableau 4.21: Integration des services de santé en {annee_annuaire}","A1"),
        "Integration_service_Couverture": (f"Tableau 4.22: Integration des services de santé en {annee_annuaire}","A1"),
        "WHOPEN": (f"Tableau 4.23: Activités de prise en charge des MNT (Asthme, Diabète, MCV) en {annee_annuaire}","A1"),
        "WHOPEN_Couverture": (f"Tableau 4.24: Activités de prise en charge des MNT (Asthme, Diabète, MCV) en {annee_annuaire}","A1"),
        "Cons_post_nat": (f"Tableau 4.25: Consultation postnatale en {annee_annuaire}","A1"),
        "Décès maternel par causes": (f"Tableau 4.26: Décès maternels par cause dans les formations sanitaires en {annee_annuaire}","A1"),
        "Taux de décès maternel": (f"Tableau 4.27: Décès maternels audités et taux de décès maternel en {annee_annuaire}","A1"),
        "PEC Avortement": (f"Tableau 4.28: Soins après avortement en {annee_annuaire}","A1"),
        "PEC excision_Fistules": (f"Tableau 4.29: Prise en charge des fistules, séquelles et complications d’excision en {annee_annuaire}","A1"),
        "PTME_VIH": (f"Tableau 4.30: Dépistage du VIH chez les femmes enceinte et la mise sous traitement en {annee_annuaire}","A1"),
        "PTME_VIH-Enfant": (f"Tableau 4.31: Suivi des enfants nés de mères seropositives en {annee_annuaire}","A1"),
        "PTME_Hep B": (f"Tableau 4.32: Dépistage de l'hépatite B chez les femmes enceinte et la mise sous traitement en {annee_annuaire}","A1"),
        "PTME_Hep B_ Syphilis": (f"Tableau 4.33: Dépistage du syphilis et de l'hépatite B chez les femmes enceinte et la mise sous traitement en {annee_annuaire}","A1"),
        "Dépistage malnut": (f"Tableau 4.34: Dépistage de la malnutrition* aigu chez les enfants de moins de 5 ans en {annee_annuaire}","A1"),
        "Confirmation MA ASBC": (f"Tableau 4.35: Dépistage de la malnutrition aigue par les ASBC en {annee_annuaire}","A1"),
        "Prévalence Malnutrition": (f"Tableau 4.36: Prévalence de la malnutrition par région en {annee_annuaire}","A1"),
        "Nutrition 1": (f"Tableau 4.37: Enfants malnutris sortis de la prise en charge en {annee_annuaire}","A1"),
        "Nutrition 2": (f"Tableau 4.38: Performance de la PEC de la malnutrition aigue en {annee_annuaire}","A1"),
        "JVAP1": (f"Tableau 4.39: Résultats de la campagne JVA+ premier passage {annee_annuaire}","A1"),
        "JVAP2": (f"Tableau 4.40: Résultats de la campagne JVA+ deuxième passage {annee_annuaire}","A1"),
        "Consultant_age": (f"Tableau 4.41: Nouveaux consultants (auto orientés) par tranche  d'âge en {annee_annuaire}","A1"),
        "consultant_CH": (f"Tableau 4.42: Consultants par centre hospitalier en {annee_annuaire}","A1"),
        "Nb_nouveau contact": (f"Tableau 4.43: Contacts par habitant par an en {annee_annuaire}","A1"),
        "PCIME_TETU": (f"Tableau 4.44: Enfants pris en charge selon l'approche PCIME et TETU en {annee_annuaire}","A1"),
        "Réf_Evac_cont ref_1er échelon": (f"Tableau 4.45: Références/contre-références des FS du premier échelon* en {annee_annuaire}","A1"),
        "Références 2et3echelon": (f"Tableau 4.46: Références réalisées et contre référence recus des CMA et CHR/CHU en {annee_annuaire}","A1"),
        "Evacuation CH_CMA": (f"Tableau 4.47: Références reçues, Contre reference réalisées CMA et CHR/CHU en {annee_annuaire}","A1"),
        "MEO_décès": (f"Tableau 4.48: Mise en observation et décès* en {annee_annuaire}","A1"),
        "Principaux motifs": (f"Tableau 4.49: les 10 principaux motifs de  consultations externes  dans les FS en {annee_annuaire}","A1"),
        "Principaux motifs": (f"Tableau 4.50: Les 10 principaux motifs d'hospitalisation dans les centres médicaux et les hôpitaux en {annee_annuaire}","A14"),
        "Principaux motifs": (f"Tableau 4.51: Les 10 principales causes de décès dans les centres médicaux et les hôpitaux en {annee_annuaire}","A27"),
        "C externe_Age": (f"Tableau 4.52: Nosologie des consultations externes dans les FS par tranche d'âge en {annee_annuaire}","A1"),
        "C externe_Région 1": (f"Tableau 4.53: Nosologie des consultations externes dans les formations sanitaires par région en {annee_annuaire}","A1"),
        "C externe_Région_2": (f"Tableau 4.54: Nosologie des consultations externes dans les formations sanitaires par région en {annee_annuaire} (suite)","A1"),
        "Hospi_CM_CMA_CH_Age": (f"Tableau 4.55: Nosologie des hospitalisations par tranche d'age en {annee_annuaire}","A1"),
        "Hospi_CM_CMA_CH_Région 1": (f"Tableau 4.56: Nosologie des hospitalisations par région en {annee_annuaire}","A1"),
        "Hospi_CM_CMA_CH_Région 2": (f"Tableau 4.57: Nosologie des hospitalisations par région en {annee_annuaire} (suite1)","A1"),
        "Hospi_CM_CMA_CH_Région 3": (f"Tableau 4.58: Nosologie des hospitalisations par région en {annee_annuaire} (suite2)","A1"),
        "Personnes_ages": (f"Tableau 4.59: Nosologie des personnes agées (60 ans et plus) en {annee_annuaire}","A1"),
        "Personnes_ages_region": (f"Tableau 4.60: Nosologie des personnes agées (60 ans et plus) en {annee_annuaire}","A1"),
        "Mode_entrée_hospi": (f"Tableau 4.61: Mode d'entrée des patients en hospitalisation en {annee_annuaire}","A1"),
        "Sortie Hospi": (f"Tableau 4.62: Mode de sortie en hospitalisation en {annee_annuaire}","A1"),
        "Confirmation palu": (f"Tableau 4.63: Confirmation du paludisme dans les formations sanitaires en {annee_annuaire}","A1"),
        "PaludismeGlobale": (f"Tableau 4.64: Paludisme simple et grave (confirmés+présumés) en {annee_annuaire}","A1"),
        "Incidence paludisme": (f"Tableau 4.65: Incidence du paludisme en {annee_annuaire}","A1"),
        "Paludisme Confirme": (f"Tableau 4.66: Paludisme confirmé en {annee_annuaire}","A1"),
        "Incidence paludisme confirmé": (f"Tableau 4.67: Incidence du paludisme confirmé en {annee_annuaire}","A1"),
        "Traitement Paludisme": (f"Tableau 4.68: Traitement du paludisme simple en {annee_annuaire}","A1"),
        "Palu_létalité": (f"Tableau 4.69: Létalité du paludisme en {annee_annuaire}","A1"),
        "Milda_Routine": (f"Tableau 4.70: Distribution des MILDA de routine en {annee_annuaire}","A1"),
        "Couverture CPS Passage 0": (f"Tableau 4.71: Résultat de la campagne de chimioprophylaxie saisonnière du paludisme chez les enfants de 3 à 59 mois au passage zéro en {annee_annuaire}","A1"),
        "Couverture CPS Passage 1": (f"Tableau 4.72: Résultat de la campagne de chimioprophylaxie saisonnière du paludisme chez les enfants de 3 à 59 mois au premier passage {annee_annuaire}","A1"),
        "Couverture CPS Passage 2": (f"Tableau 4.73: Résultat de la campagne de chimioprophylaxie saisonnière du paludisme chez les enfants de 3 à 59 mois au deuxième passage {annee_annuaire}","A1"),
        "Couverture CPS Passage 3": (f"Tableau 4.74: Résultat de la campagne de chimioprophylaxie saisonnière du paludisme chez les enfants de 3 à 59 mois au troisième passage {annee_annuaire}","A1"),
        "Couverture CPS Passage 4": (f"Tableau 4.75: Résultat de la campagne de chimioprophylaxie saisonnière du paludisme chez les enfants de 3 à 59 mois au quatrième passage {annee_annuaire}","A1"),
        "IST": (f"Tableau 4.76: Notification syndromique des IST en {annee_annuaire}","A1"),
        "Conseil_dépist_VIH": (f"Tableau 4.77: Dépistage de l'infection à VIH en {annee_annuaire}","A1"),
        "dépist Hépatite virale": (f"Tableau 4.78: Dépistage des hépatites virales en {annee_annuaire}","A1"),
        "File active": (f"Tableau 4.79: Suivi des PvVIH en fin {annee_annuaire}","A1"),
        "AELB": (f"Tableau 4.80: Accidents d'exposition aux liquides biologiques en {annee_annuaire}","A1"),
        "Depistage TB": (f"Tableau 4.81: Dépistage de la tuberculose en {annee_annuaire}","A1"),
        "Depistage TB_suite": (f"Tableau 4.82: Dépistage de la tuberculose en {annee_annuaire} (Suite)","A1"),
        "Basciloscopie": (f"Tableau 4.83: Activités de bacilloscopie de la tuberculose en {annee_annuaire}","A1"),
        "Notification TB_Dépist TBVIH": (f"Tableau 4.84: Taux de notification de la tuberculose en {annee_annuaire}","A1"),
        "CoInfectionTB_VIH": (f"Tableau 4.85: Dépistage du VIH chez les patients de tuberculose en {annee_annuaire}","A1"),
        "CoInfectionTB_VIH": (f"Tableau 4.86: Situation de la réalisation des test VIH chez les patients tuberculeux et la prise en charge des co-infectés pour la cohorte 2023","A23"),
        "Traitement TB": (f"Tableau 4.87: Résultats de traitement de la cohorte 2023 des cas de tuberculose (toutes formes)","A1"),
        "Lèpre1": (f"Tableau 4.88: Lèpre en {annee_annuaire}","A1"),
        "Lèpre2": (f"Tableau 4.89: Situation de la lèpre en fin d'année en {annee_annuaire}","A1"),
        "TDM FL": (f"Tableau 4.90: Résultats de traitement de masse contre la filariose lymphatique(FL) en {annee_annuaire}","A1"),
        "TDM Schistosomiase": (f"Tableau 4.91: Résultats de traitement de masse contre la schistosomiase en {annee_annuaire}","A1"),
        "TIDC_Oncho": (f"Tableau 4.92: Résultats de traitement sous directives communautaires (TIDC) contre l'onchocercose en {annee_annuaire}","A1"),
        "Complétude TLOH_Publique ": (f"Tableau 4.93: Complétude et promptitude des télégrammes lettres officiels hebdomadaires des FS publiques en {annee_annuaire}","A1"),
        "Complétude TLOH_Privée": (f"Tableau 4.94: complétude et promptitude des télégrammes lettres officiels hebdomadaires des FS privées en {annee_annuaire}","A1"),
        "Complétude TLOH_Globale": (f"Tableau 4.95: complétude et promptitude globales des télégrammes lettres officiels hebdomadaires des FS en {annee_annuaire}","A1"),
        "MPE": (f"Tableau 4.996: Cas et décès des maladies sous surveillance en {annee_annuaire}","A1"),
        "MPE(suite1)": (f"Tableau 4.997: cas et décès des maladies sous surveillance en {annee_annuaire} (suite)","A1"),
        "MPE(suite2": (f"Tableau 4.998: cas et décès des maladies sous surveillance en {annee_annuaire} (suite)","A1"),
        "Resultat_LCR Gram ": (f"Tableau 4.99: résultats des examens de LCR en {annee_annuaire}","A1"),
        "Resultat_LCR Latex ": (f"Tableau 4.100: résultats des examens de LCR en {annee_annuaire}","A1"),
        "Résultats LCR PCR": (f"Tableau 4.101: résultats des examens de LCR PR en {annee_annuaire}","A1"),
        "Acte odonto": (f"Tableau 4.102: Actes réalisés en odontostomatologie en {annee_annuaire}","A1"),
        "Acte ORL": (f"Tableau 4.103: Actes réalisés en ORL en {annee_annuaire}","A1"),
        "Anesthésie": (f"Tableau 4.104: Activité d'anesthesie en {annee_annuaire}","A1"),
        "Chirurgie_HD": (f"Tableau 4.105: Interventions de chirurgie essentielle dans les districts sanitaires en {annee_annuaire}","A1"),
        "Chirurgie_Ophtalmo": (f"Tableau 4.106: Interventions chirurgicales d'ophtalmologie en {annee_annuaire}","A1"),
        "Chirurgie_HD_orl": (f"Tableau 4.107: Interventions chirurgicales en ORL en {annee_annuaire}","A1"),
        "Chirurgie_CH1": (f"Tableau 4.108: Interventions chirurgicales et orthopédiques réalisées dans les CH en {annee_annuaire}","A1"),
        "Chirurgie_CH2": (f"Tableau 4.109: Interventions chirurgicales et orthopédiques réalisées dans les CH en {annee_annuaire} (suite)","A1"),
        "Taux_césarienne": (f"Tableau 4.110: Taux de réalisation des césariennes en {annee_annuaire}","A1"),
        "activ_soins_urgence": (f"Tableau 4.111: Patients reçus en urgence dans les CH en {annee_annuaire}","A1"),
        "sortie urgence": (f"Tableau 4.112: modes de sortie  en urgence dans les CH en {annee_annuaire}","A1"),
        "Activité imagerie": (f"Tableau 4.113: Examens d'imagerie médicale réalisés en {annee_annuaire}","A1"),
        "Activité imagerie2": (f"Tableau 4.114: Autres examens d'imagerie médicale réalisés en {annee_annuaire}","A1"),
        "Médecine nucléaire": (f"Tableau 4.115: Examens réalisés en médecine nucléaire du CHU-YO en {annee_annuaire}","A1"),
        "Examen_Labo": (f"Tableau 4.116: Examens de laboratoire en {annee_annuaire}","A1"),
        "CNTS1": (f"Tableau 4.117: Nombre de poches de sang collectées par type de dons dans les CRTS et les CHR/CMA","A1"),
        "CNTS1": (f"Tableau 4.118: Répartition des dons selon le type de collecte par CRTS et DPD/PS","A8"),
        "CNTS1": (f"Tableau 4.119: Répartition des dons selon les tranches d’âge des donneurs et par CRTS","A23"),
        "CNTS1": (f"Tableau 4.120: aux  d’exclusion des candidats au don de sang dans les CRTS et DPD","A38"),
        "CNTS1": (f"Tableau 4.121: Type d’exclusion au don de sang par CRTS","A52"),
        "CNTS2": (f"Tableau 4.122: Taux de positivité des marqueurs infectieux* dans les  CRTS","A1"),
        "CNTS3": (f"Tableau 4.123: Pourcentage de poches séparées en PSL par CRTS","A1"),
        "CNTS3": (f"Tableau 4.124: Nombre d’unités de PSL obtenus à partir de sang total","A15"),
        "CNTS4": (f"Tableau 4.125: Motifs de rejet des  poches de sang","A1"),
        "CNTS4": (f"Tableau 4.126: Nombre de formations sanitaires publiques et privées approvisionnées par les CRTS et DPD/PS","A18"),
        "CNTS5": (f"Tableau 4.127: Satisfaction des demandes en produits sanguins labiles (PSL) dans les zones CNTS","A1"),
        "CNTS5": (f"Tableau 4.128: Répartition des receveurs de PSL selon l’âge et le sexe","A16"),
        "CNTS5": (f"Tableau 4.129: aux de retour des Fiches post-transfusionnelles d’hémovigilance (FPTH) et la traçabilité des PSL des CRTS","A26"),
        "OST": (f"Tableau 4.130: situation des visites des lieux de travail (VLT) réalisées dans les directions régionales de l'OST en {annee_annuaire}","A1"),
        "OST": (f"Tableau 4.131: situation des séances d'IEC dans les directions régionales de l'OST en {annee_annuaire}","A11"),
        "OST": (f"Tableau 4.132: situation de la visite médicale périodique (VMP) des travailleurs dans les directions régionales de l'OST en {annee_annuaire}","A20"),
        "OST": (f"Tableau 4.133: situation des autres types de visites médicales reglementaires réalisées dans les directions régionales de l'OST en {annee_annuaire}","A30"),
        "OST": (f"Tableau 4.134: situation des accidents du travail (AT) notifiés dans les directions régionales de l'OST en {annee_annuaire}","A39"),
        "OST": (f"Tableau 4.135: situation des Maladies professionnelles (MP) et maladies à caratère professionnel (MCP)","A47"),
        "MPR": (f"Tableau 4.136: Activités MPR en {annee_annuaire}","A1"),
        "MPR_Couverture": (f"Tableau 4.137: Activités MPR en {annee_annuaire}","A1"),
        "ENSP": (f"Tableau 4.138: Effectifs des etudiants de l’ensp pour l’annee {annee_annuaire}","A1"),
        "ANSSEAT": (f"Tableau 4.139: Bilan de réalisation des analyses de routine de l'année {annee_annuaire}","A1"),
        "Situation ASBC": (f"Tableau 5.01: Situation des ASBC récrutés en {annee_annuaire}","A1"),
        "Complétude ASBC": (f"Tableau 5.02: Complétude de la saisie des RMA ASBC en {annee_annuaire}","A1"),
        "Seance IEC_ASBC_1": (f"Tableau 5.03: Séance d'IEC réalisé par les ASBC selon le thème en {annee_annuaire}","A1"),
        "Seance IEC_ASBC_2": (f"Tableau 5.04: Séance d'IEC réalisé par les ASBC selon le thème en {annee_annuaire} (Suite)","A1"),
        "Participants IEC ASBC_1": (f"Tableau 5.05: Personnes touchées lors des séances d'IEC réalisées par les ASBC selon le thème en {annee_annuaire}","A1"),
        "Participants IEC ASBC_2": (f"Tableau 5.06: Personnes touchées lors des séances d'IEC réalisées par les ASBC selon le thème en {annee_annuaire} (suite1)","A1"),
        "PF ASBC": (f"Tableau 5.07: Distribution des méthodes contraceptives par les ASBC en {annee_annuaire}","A1"),
        "Accompagnement ASBC": (f"Tableau 5.08: Accompagnement de la grossesse, l’accouchement et le post partum en {annee_annuaire}","A1"),
        "PDV Vaccination": (f"Tableau 5.09: Recherche de perdus de vue à la vaccination en {annee_annuaire}","A1"),
        "GASPA": (f"Tableau 5.10: Réalisation des GASPA par les ASBC en {annee_annuaire}","A1"),
        "Dépistage communautaire MA": (f"Tableau 5.11: Dépistage  de la malnutrition aiguë au niveau communautaire en {annee_annuaire}","A1"),
        "Palu ASBC_1": (f"Tableau 5.12: Diagnostic du paludisme au niveau communautaire par les ASBC en {annee_annuaire}","A1"),
        "Palu ASBC_2": (f"Tableau 5.13: Traitement du paludisme au niveau communautaire par les ASBC en {annee_annuaire}","A1"),
        "DIARRHEE ASBC": (f"Tableau 5.14: Prise en charge de la diarrhée au niveau communautaire par les ASBC en {annee_annuaire}","A1"),
        "Toux ASBC": (f"Annuaire 5.15: Prise en charge de la toux au niveau communautaire par les ASBC en {annee_annuaire}","A1"),
        "TPI com": (f"Annuaire 5.16: TPI communautaire en {annee_annuaire}","A1"),
        "Décès en communauté": (f"Annuaire 5.17: Decès en communauté notifiés par les ASBC en 2025 {annee_annuaire}","A1")
        }

        # --- Fonction principale d'ajout de titres ---
        def ajouter_titres_et_generer(temp_path, titres_tableaux):
            try:
                st.write("🔹 **Ajout des titres dans les feuilles du canevas...**")
                wb = load_workbook(temp_path)

                for sheet_name, (titre, cellule) in titres_tableaux.items():
                    if sheet_name not in wb.sheetnames:
                        st.warning(f"⚠️ Feuille '{sheet_name}' introuvable — ignorée.")
                        continue

                    ws = wb[sheet_name]

                    try:
                        # Conversion de la référence en coordonnées (ex: "A1" → (1,1))
                        row, col = coordinate_to_tuple(cellule)

                        # Écriture directe du titre (ne pose pas de problème avec merged cells)
                        ws.cell(row=row, column=col).value = str(titre)

                    except Exception as ex:
                        st.warning(f"⚠️ Erreur lors de l'ajout du titre dans '{sheet_name}': {ex}")

                # Sauvegarde du fichier mis à jour
                wb.save(temp_path)
                wb.close()

                # --- Sauvegarde et téléchargement final ---
                date_aujourdhui = datetime.today().strftime("%d-%m-%Y")
                nom_fichier_final = f"Annuaire_Sanitaire_MS_{date_aujourdhui}.xlsx"

                # Supprimer l'ancien fichier s’il existe
                if os.path.exists(nom_fichier_final):
                    os.remove(nom_fichier_final)

                os.rename(temp_path, nom_fichier_final)

                # Afficher le bouton de téléchargement (clé unique)
                with open(nom_fichier_final, "rb") as f:
                    st.download_button(
                        label="⬇️ Télécharger le fichier généré",
                        data=f,
                        file_name=nom_fichier_final,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{date_aujourdhui}"  # ✅ évite le bug de doublon
                    )

                st.success("✅ Fichier Annuaire généré avec succès !")

            except Exception as e:
                st.error(f"⚠️ Erreur lors de la génération : {e}")

        ajouter_titres_et_generer(temp_path, titres_tableaux)
    except Exception as e:
        st.error(f"⚠️ Erreur lors de la génération : {e}")
